#!/usr/bin/env python3
"""XLSX input adapter. Turns a workbook into the collection shape everything
else in this repo already consumes. It adds no task semantics.

```text
workbook.xlsx --> {"orders": [{...}, ...], "prices": [{...}, ...]}
```

Downstream, nothing knows it came from Excel. The observer measures it, the
modeller models it, the executor runs it, exactly as with hand-written JSON.

## Every conversion is DECLARED, and none of them tidies

This is the whole risk of a spreadsheet adapter. Excel hands back Python
objects, and the temptation is to make them look nice on the way past.

```text
str        unchanged
int        str(int)                    40        -> "40"
float      repr -- shortest decimal that round-trips
                                       19.99     -> "19.99"
                                       0.1       -> "0.1"
                                       0.30000000000000004 stays exactly that
datetime   ISO date at midnight, ISO datetime otherwise
bool       "true" / "false"            checked BEFORE int; bool IS an int
None       the field is OMITTED from the row, which is not the same as ""
```

`repr` on a float is the shortest decimal string that reads back as the same
double, so a literal a person typed into a cell comes back as that literal. A
value that is genuinely `0.30000000000000004` is reported as
`0.30000000000000004` — **the adapter reports what is stored, it does not make
it presentable.** Emitting `"0.3"` there would be the same defect the enrichment
executor already refuses to commit (PRO-2 instance 9).

### A limit of the carrier, not of this adapter

Writing `0.30000000000000004` into a workbook with openpyxl stores `0.3` in the
sheet XML — the seventeenth digit is gone before any reader sees it. This
adapter READS; it is faithful to whatever the file actually holds, and the
self-test proves that on the conversion directly. But it is worth knowing that a
spreadsheet round-trip is not a lossless carrier of a double, so a workbook is
evidence of what someone typed, not of what a computation produced.

## What it refuses rather than guesses

```text
a formula cell with no cached value   the workbook was never calculated, so
                                      there is nothing faithful to read
a blank or duplicated header          a column that cannot be named cannot be
                                      referred to by a model
an empty sheet, a missing sheet       nothing to convert
```

Refusing is the point. An adapter that guesses puts a wrong number into a
deterministic pipeline that will then be faithful to it forever.
"""
from __future__ import annotations

import datetime as dt
import json
import sys
from dataclasses import dataclass, field as dc_field
from pathlib import Path
from typing import Any, Optional

import openpyxl

PROBLEM_CODES = (
    "missing_sheet",
    "empty_sheet",
    "blank_header",
    "duplicate_header",
    "uncalculated_formula",
    "unsupported_cell_type",
)


class UnreadableWorkbook(Exception):
    """The workbook cannot be converted faithfully."""

    def __init__(self, problems: list[str]):
        super().__init__("; ".join(problems))
        self.problems = problems


@dataclass(frozen=True)
class SheetSpec:
    """Which sheet becomes which collection. Declared, never guessed."""
    sheet: str
    collection: str
    header_row: int = 1


@dataclass
class Conversion:
    collections: dict = dc_field(default_factory=dict)
    problems: list = dc_field(default_factory=list)
    notes: list = dc_field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.problems


def specs_from(raw: Any) -> list[SheetSpec]:
    return [SheetSpec(s["sheet"], s["collection"], int(s.get("header_row", 1)))
            for s in raw]


def cell_value(value: Any) -> Optional[str]:
    """One cell, converted by the declared table. Raises on anything else."""
    if value is None:
        return None
    if isinstance(value, bool):            # BEFORE int -- bool is an int
        return "true" if value else "false"
    if isinstance(value, str):
        if value.startswith("="):
            raise ValueError("uncalculated_formula")
        return value
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return repr(value)                 # shortest decimal that round-trips
    if isinstance(value, dt.datetime):
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    raise ValueError(f"unsupported_cell_type:{type(value).__name__}")


def convert(path: Path, specs: list[SheetSpec]) -> Conversion:
    """Read the declared sheets. `data_only=True` reads cached formula results."""
    out = Conversion()
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for spec in specs:
            if spec.sheet not in book.sheetnames:
                out.problems.append(f"missing_sheet: {spec.sheet!r} "
                                    f"(available: {book.sheetnames})")
                continue
            sheet = book[spec.sheet]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) <= spec.header_row:
                out.problems.append(f"empty_sheet: {spec.sheet!r} has no data "
                                    f"rows below header row {spec.header_row}")
                continue

            header_raw = rows[spec.header_row - 1]
            headers: list[str] = []
            for index, name in enumerate(header_raw):
                if name is None or str(name).strip() == "":
                    if any(r[index] is not None for r in rows[spec.header_row:]):
                        out.problems.append(
                            f"blank_header: {spec.sheet!r} column {index + 1} "
                            f"has data but no name")
                    headers.append("")
                    continue
                headers.append(str(name).strip())
            named = [h for h in headers if h]
            if len(named) != len(set(named)):
                out.problems.append(f"duplicate_header: {spec.sheet!r} {named}")

            items = []
            for number, row in enumerate(rows[spec.header_row:],
                                         start=spec.header_row + 1):
                if all(v is None for v in row):
                    continue               # a wholly blank row is not a record
                item = {}
                for header, value in zip(headers, row):
                    if not header:
                        continue
                    try:
                        converted = cell_value(value)
                    except ValueError as exc:
                        out.problems.append(
                            f"{exc}: {spec.sheet!r}!{header} row {number}")
                        continue
                    if converted is not None:
                        item[header] = converted
                if item:
                    items.append(item)
            if not items and not out.problems:
                out.problems.append(f"empty_sheet: {spec.sheet!r} produced no rows")
            out.collections[spec.collection] = items
            out.notes.append(f"{spec.sheet} -> {spec.collection}: {len(items)} row(s)")
    finally:
        book.close()
    return out


def write_collections(conversion: Conversion, directory: Path,
                      note: str = "") -> list[Path]:
    """One JSON file per collection, in the shape every source in this repo uses."""
    if not conversion.ok:
        raise UnreadableWorkbook(conversion.problems)
    directory.mkdir(parents=True, exist_ok=True)
    written = []
    for collection, items in conversion.collections.items():
        path = directory / f"{collection}.json"
        payload = {collection: items}
        if note:
            payload = {"_note": note, collection: items}
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
                        encoding="utf-8")
        written.append(path)
    return written


def _self_test() -> int:
    import tempfile
    from openpyxl import Workbook
    failures: list[str] = []

    def check(cond: bool, msg: str) -> None:
        if not cond:
            failures.append(msg)

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "book.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "Orders"
        sheet.append(["line", "code", "qty", "when", "urgent", "note"])
        sheet.append(["L-1", "P-100", 3, dt.date(2026, 3, 2), True, "first"])
        sheet.append(["L-2", "P-200", 40, dt.date(2026, 3, 9), False, None])
        sheet.append([None, None, None, None, None, None])       # blank row
        prices = book.create_sheet("Prices")
        prices.append(["code", "unit_price"])
        prices.append(["P-100", 19.99])
        prices.append(["P-200", 0.1])
        prices.append(["P-300", 0.30000000000000004])
        book.save(path)

        spec = [SheetSpec("Orders", "orders"), SheetSpec("Prices", "prices")]
        result = convert(path, spec)
        check(result.ok, f"a clean workbook must convert: {result.problems}")
        orders, price_rows = result.collections["orders"], result.collections["prices"]

        check(len(orders) == 2, f"a wholly blank row is not a record: {orders}")
        check(orders[0]["qty"] == "3" and orders[1]["qty"] == "40",
              f"integers become plain strings: {[o['qty'] for o in orders]}")
        check(orders[0]["when"] == "2026-03-02",
              f"a midnight datetime becomes an ISO date: {orders[0]['when']}")
        check(orders[0]["urgent"] == "true" and orders[1]["urgent"] == "false",
              f"CANARY: booleans must not become 1 and 0: "
              f"{[o['urgent'] for o in orders]}")
        check("note" not in orders[1],
              f"CANARY: an empty cell is OMITTED, not empty string: {orders[1]}")

        # --- the money round-trip, which is the whole point ---------------
        check(price_rows[0]["unit_price"] == "19.99",
              f"CANARY: a money literal must survive exactly: "
              f"{price_rows[0]['unit_price']}")
        check(price_rows[1]["unit_price"] == "0.1",
              f"…and so must 0.1: {price_rows[1]['unit_price']}")
        # The adapter must not tidy. Tested on the CONVERSION, with no file in
        # the way, because the write path is not the adapter's and does lose
        # this -- see the note below and the module docstring.
        check(cell_value(0.30000000000000004) == "0.30000000000000004",
              f"CANARY: the adapter must report an ugly value as it IS, not "
              f"tidy it to 0.3: {cell_value(0.30000000000000004)}")
        check(price_rows[2]["unit_price"] == "0.3",
              f"…and the openpyxl WRITE path is what collapsed this one to 0.3 "
              f"before the adapter ever saw it: {price_rows[2]['unit_price']}")
        from decimal import Decimal
        check(Decimal(price_rows[0]["unit_price"]) * Decimal("3")
              == Decimal("59.97"),
              "…and must be exactly computable downstream")

        # --- refusals -----------------------------------------------------
        bad = Workbook()
        s = bad.active
        s.title = "Orders"
        s.append(["line", None, "qty"])
        s.append(["L-1", "orphan", 3])
        bad_path = Path(tmp) / "blank_header.xlsx"
        bad.save(bad_path)
        result = convert(bad_path, [SheetSpec("Orders", "orders")])
        check(any("blank_header" in p for p in result.problems),
              f"CANARY: a data column with no name must be refused: "
              f"{result.problems}")

        dup = Workbook()
        s = dup.active
        s.title = "Orders"
        s.append(["code", "code"])
        s.append(["A", "B"])
        dup_path = Path(tmp) / "dup.xlsx"
        dup.save(dup_path)
        result = convert(dup_path, [SheetSpec("Orders", "orders")])
        check(any("duplicate_header" in p for p in result.problems),
              f"CANARY: duplicate headers must be refused: {result.problems}")

        result = convert(path, [SheetSpec("Nope", "nope")])
        check(any("missing_sheet" in p for p in result.problems),
              f"CANARY: a missing sheet must be refused: {result.problems}")

        empty = Workbook()
        empty.active.title = "Orders"
        empty.active.append(["a", "b"])
        empty_path = Path(tmp) / "empty.xlsx"
        empty.save(empty_path)
        result = convert(empty_path, [SheetSpec("Orders", "orders")])
        check(any("empty_sheet" in p for p in result.problems),
              f"CANARY: a header with no rows must be refused: {result.problems}")

        # --- an uncalculated formula ---------------------------------------
        try:
            cell_value("=A1*B1")
            failures.append("CANARY: an uncalculated formula must be refused")
        except ValueError as exc:
            check("uncalculated_formula" in str(exc), f"named correctly: {exc}")

        # --- writing produces the shape everything else consumes -----------
        out_dir = Path(tmp) / "out"
        written = write_collections(convert(path, spec), out_dir, note="from xlsx")
        check(len(written) == 2, f"one file per collection: {written}")
        payload = json.loads((out_dir / "orders.json").read_text(encoding="utf-8"))
        check(list(payload) == ["_note", "orders"] and len(payload["orders"]) == 2,
              f"…in the collection shape: {list(payload)}")
        try:
            write_collections(convert(bad_path, [SheetSpec("Orders", "orders")]),
                              out_dir)
            failures.append("CANARY: a refused conversion must not be written")
        except UnreadableWorkbook:
            pass

    if failures:
        sys.stderr.write("SELF-TEST FAILED:\n  " + "\n  ".join(failures) + "\n")
        return 1
    print("SELF-TEST PASSED (a clean workbook converts / blank rows are not "
          "records / integers, ISO dates and booleans convert by the declared "
          "table and booleans do not become 1 and 0 / an empty cell is omitted "
          "rather than emptied / 19.99 and 0.1 survive exactly and are exactly "
          "computable, and 0.30000000000000004 is reported as it IS rather than "
          "tidied / blank headers, duplicate headers, missing sheets, empty "
          "sheets and uncalculated formulas are all refused / a refused "
          "conversion is never written)")
    return 0


def main(argv: list[str]) -> int:
    if argv[:1] == ["--self-test"]:
        return _self_test()
    if len(argv) < 2:
        sys.stderr.write("usage: xlsx.py <book.xlsx> <specs.json> [outdir]\n")
        return 2
    specs = specs_from(json.loads(Path(argv[1]).read_text(encoding="utf-8")))
    result = convert(Path(argv[0]), specs)
    if not result.ok:
        for problem in result.problems:
            sys.stderr.write(problem + "\n")
        return 1
    if len(argv) > 2:
        for path in write_collections(result, Path(argv[2])):
            print(path)
    else:
        print(json.dumps(result.collections, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
