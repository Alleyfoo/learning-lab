#!/usr/bin/env python3
"""Generate Experiment K's candidate workbooks.

The generator is the fixtures' source of truth: regenerating rebuilds them.
Every candidate is a full 6-sheet workbook with the same shape as
`definition_phase/fixtures/W1_multisheet.xlsx`, varying ONLY the thing named.

  C1  identical            byte-identical to W1
  C2  renamed file         identical content, different filename
  C3  more data rows       6 products instead of 4; the total row moves
  C4  column renamed       Tuote -> Tuotekoodi
  C5  column inserted      a new 'Maa' column at B (shifts the month block)
  C6  sheet added          a new 'Kampanjat' sheet
  C7  data sheet renamed   Sales -> Myynnit
  C8  silent row semantics SAME dimensions, but a product row is replaced by a
                           VÄLISUMMA subtotal row

C8 is the load-bearing one: nothing about the workbook's shape changed, so a
shape-based predicate cannot see it.
"""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
LAB = ROOT.parent
W1 = LAB / "definition_phase" / "fixtures" / "W1_multisheet.xlsx"
OUT = ROOT / "fixtures"

SALES_HEADER = ["Tuote", "Tammi", "Helmi", "Maalis", "Yhteensä", "Kommentti"]
SALES_DATA = [
    ["ART-001", 10, 12, 8, 30, None],
    ["ART-002", 7, 9, 11, 27, "korjattu"],
    ["ART-003", 5, 7, 6, 18, None],
    ["ART-004", 8, 11, 7, 26, None],
]
SALES_TOTAL = ["YHTEENSÄ", 30, 39, 32, 101, None]


def _rows(ws: Worksheet, rows: list[list[object]]) -> None:
    for row in rows:
        ws.append(row)


def _sales(ws: Worksheet, header: list, data: list, total: list) -> None:
    _rows(ws, [
        ["Myyntiraportti 2026", None, None, None, None, None],
        ["Päivitetty", "3.2.2026", None, None, None, None],
        [],
        header,
        *data,
        total,
    ])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))


def _companions(wb: Workbook, extra_sheet: str | None = None) -> None:
    ws = wb.create_sheet("Myynti 2026")
    _rows(ws, [["Tuote", "Myynti", "Kate"], ["ART-001", 100, 30], ["ART-002", 80, 25]])
    ws = wb.create_sheet("Dup")
    _rows(ws, [["Tuote", "Myynti", "Myynti"], ["ART-001", 1, 2], ["ART-002", 3, 4]])
    ws = wb.create_sheet("Notes")
    _rows(ws, [["Vapaata tekstiä. Ei taulukkoa."],
               ["Toimittaja vaihtoi järjestelmää 1.1.2026."]])
    for name, base in (("2026-01", 10), ("2026-02", 20)):
        ws = wb.create_sheet(name)
        _rows(ws, [["Tuote", "Myynti"], ["ART-001", base], ["ART-002", base + 5]])
    if extra_sheet:
        ws = wb.create_sheet(extra_sheet)
        _rows(ws, [["Kampanja", "Alennus"], ["KEVAT", "10%"], ["SYKSY", "15%"]])


def _build(sales_name: str, header: list, data: list, total: list,
           extra_sheet: str | None = None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sales_name
    _sales(ws, header, data, total)
    _companions(wb, extra_sheet)
    return wb


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)

    # C1 / C2: identical content. C2 differs ONLY in filename -- the case
    # Data-agents-demo's filename-sensitive hash gets wrong (DA-2).
    shutil.copyfile(W1, OUT / "C1_identical.xlsx")
    shutil.copyfile(W1, OUT / "C2_renamed_file.xlsx")

    # C3: next month's file -- two more products, so the total row moves.
    more = SALES_DATA + [["ART-005", 6, 8, 9, 23, None], ["ART-006", 4, 5, 3, 12, None]]
    _build("Sales", SALES_HEADER, more,
           ["YHTEENSÄ", 40, 52, 44, 136, None]).save(OUT / "C3_more_rows.xlsx")

    # C4: a header renamed -> the NAMED binding @Tuote cannot resolve.
    _build("Sales", ["Tuotekoodi"] + SALES_HEADER[1:], SALES_DATA,
           SALES_TOTAL).save(OUT / "C4_column_renamed.xlsx")

    # C5: a column inserted -> the POSITIONAL binding B:D silently shifts.
    ins_header = ["Tuote", "Maa"] + SALES_HEADER[1:]
    ins_data = [[r[0], "FI"] + r[1:] for r in SALES_DATA]
    ins_total = [SALES_TOTAL[0], None] + SALES_TOTAL[1:]
    _build("Sales", ins_header, ins_data, ins_total).save(OUT / "C5_column_inserted.xlsx")

    # C6: an extra sheet appears -- nothing breaks, but "which sheets matter?"
    # is a definition decision and there is now a new one.
    _build("Sales", SALES_HEADER, SALES_DATA, SALES_TOTAL,
           extra_sheet="Kampanjat").save(OUT / "C6_sheet_added.xlsx")

    # C7: the bound data sheet is gone under that name -> nothing carries over.
    _build("Myynnit", SALES_HEADER, SALES_DATA,
           SALES_TOTAL).save(OUT / "C7_sheet_renamed.xlsx")

    # C8: THE BLIND-SPOT CASE. Same dimensions, same headers, same sheets --
    # but a product row has become a VÄLISUMMA subtotal. Shape is unchanged, so
    # a shape-based predicate has nothing to see, and the subtotal would be
    # consumed as if it were a product.
    quiet = [
        ["ART-001", 10, 12, 8, 30, None],
        ["ART-002", 7, 9, 11, 27, "korjattu"],
        ["VÄLISUMMA", 17, 21, 19, 57, None],
        ["ART-003", 5, 7, 6, 18, None],
    ]
    _build("Sales", SALES_HEADER, quiet,
           ["YHTEENSÄ", 22, 28, 25, 75, None]).save(OUT / "C8_silent_subtotal.xlsx")

    for path in sorted(OUT.glob("C*.xlsx")):
        print(f"wrote {path.name}")


if __name__ == "__main__":
    main()
