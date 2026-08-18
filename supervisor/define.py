#!/usr/bin/env python3
"""Workspace v0.3 -- the thinnest callable interface from the supervisor to the
EXISTING modeller floor + fleet establishment.

This module does NOT reimplement observation, interpretation, proposal, preview,
or establishment. It calls `modeller/pipeline.py` + `modeller/builder.py` (the
same functions `modeller/app.py` and `modeller/journey.py` call) and
`fleet.establish` (the same function `fleet/seed.py` calls). The supervisor UI
drives these stages one at a time so the human stays in the loop:

    select incoming data -> observe (program) -> describe the job ->
    interpret (LLM) -> propose a task model (LLM) -> [answer a load-bearing
    question if asked] -> deterministic preview -> EXPLICIT Establish ->
    the new worker appears on the System Map.

Evidence boundaries are the modeller's, unchanged: OBSERVED is program-only;
INFERRED/UNKNOWN come only from the LLM through the boundary; CONFIRMED only
from a human answer (`pipeline.submit_answer`). Sufficiency gates
(`check_join_supported` / the task's own validator inside `pipeline.build`) run
before any worker is established. Establishment is an explicit human action --
this module never calls `fleet.establish` on its own; the UI calls it only when
the operator clicks "Establish worker".

The one modeller adaptation: `modeller/app.py` calls `pipeline.propose(...)`
without `task=`, so it always models as enrichment. Here the operator-chosen task
family is threaded into `propose(..., task=...)`, so reconciliation / aggregation
/ reservation models can be produced. `modeller/app.py` itself is NOT touched.

## What this module is NOT

Not a second modeller, not a new task language, not a UI. It is glue: the
smallest set of callable wrappers over machinery that already exists and is
self-tested elsewhere. The LLM `ask` is a local copy of the generate-form call
`modeller/journey.py` and `modeller/app.py` already duplicate -- factoring a
shared one would touch `modeller/` and `fleet/`, out of scope for v0.3.
"""
from __future__ import annotations

import datetime as _dt
import json
import sys
import urllib.request
from pathlib import Path
from typing import Optional

HERE = Path(__file__).resolve().parent
LAB = HERE.parent
# `import fleet` resolves to fleet/fleet.py (no fleet/__init__.py; LAB/fleet on
# path). `import pipeline`/`builder` resolve via LAB/modeller (pipeline itself
# installs inspector + boundary + builder's task harness paths). `import
# incoming` is the supervisor's read-only scanner (used only by the self-test's
# linkage check).
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(LAB / "modeller"))
sys.path.insert(0, str(LAB / "fleet"))
sys.path.insert(0, str(LAB / "adapters"))

import fleet      # noqa: E402  (fleet/fleet.py: establish, load, load_all, ROOT)
import pipeline   # noqa: E402  (modeller/pipeline.py: the staged journey)
import incoming   # noqa: E402  (supervisor/incoming.py: scan -- self-test only)
import xlsx as xlsx_adapter  # noqa: E402  (adapters/xlsx.py: the declared conversion)

# Re-exported so the UI builds declarations through the glue, not by reaching
# into the adapter module directly. A SheetSpec is a DECLARED sheet -> collection
# + header_row mapping; the program never guesses one.
SheetSpec = xlsx_adapter.SheetSpec

MODEL = "glm-5.2:cloud"
ENDPOINT = "http://localhost:11434/api/generate"
REQUEST_TIMEOUT = 900


# ---------------------------------------------------------------------------
# the LLM call -- generate form, same as modeller/journey.py:28-34
# ---------------------------------------------------------------------------

def ask(prompt: str) -> str:
    """One round-trip to local Ollama (generate form). Returns the response text.

    The supervisor's own `core._chat` uses the chat protocol and is not
    interchangeable: the modeller pipeline hands `interpret`/`propose` a single
    prompt string and expects a single response string. This is therefore a
    local copy of the call `modeller/journey.py` and `modeller/app.py` already
    hold; factoring a shared `ask` would touch `modeller/`+`fleet/` and is out of
    scope for v0.3.
    """
    payload = json.dumps({"model": MODEL, "prompt": prompt,
                          "stream": False}).encode("utf-8")
    req = urllib.request.Request(ENDPOINT, data=payload,
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=REQUEST_TIMEOUT) as resp:
        return json.loads(resp.read())["response"]


# ---------------------------------------------------------------------------
# 1. data -- resolve a selected incoming-data dir into a modeller Workspace
# ---------------------------------------------------------------------------

def workspace_for(data_dir: str) -> Optional[pipeline.Workspace]:
    """The modeller Workspace for a `data/<dir>/` name, or None if not reachable.

    `pipeline.workspaces()` already lists every `data/<dir>/` as a Workspace,
    so selecting an incoming-data dir IS selecting a workspace -- no new loader.
    """
    for ws in pipeline.workspaces():
        if ws.label == data_dir:
            return ws
    return None


def chosen_sources(ws: pipeline.Workspace) -> list[pipeline.SourceFile]:
    """All JSON collections in the workspace. The modeller needs >=2 to relate."""
    return pipeline.sources_in(ws)


def observed(ws: pipeline.Workspace, chosen: list[pipeline.SourceFile]) -> list[dict]:
    """Measured OBSERVED facts for the selected sources (program-only)."""
    return pipeline.observed_facts(ws, chosen)


def relationships(observed_facts: list[dict]) -> list[dict]:
    return pipeline.relationships(observed_facts)


def source_spec(ws: pipeline.Workspace, chosen: list[pipeline.SourceFile]) -> dict:
    """The `sources` dict that becomes the model's source bindings."""
    return pipeline.source_spec(ws, chosen)


def expressible_tasks(chosen: list[pipeline.SourceFile]) -> tuple[str, ...]:
    """Task shapes the selected sources could support. Eliminating, not choosing.

    For >=2 collections: (enrichment, reconciliation, reservation). Structure
    cannot pick among these -- purpose does -- so the operator chooses.
    """
    return pipeline.expressible(chosen)


def suggest_task(goal: str, chosen: list[pipeline.SourceFile],
                 ask_fn=ask) -> tuple[Optional[str], Optional[dict]]:
    """The LLM task-shape choice (`pipeline.choose_task`), as a default suggestion.

    Returns (task, None) when the LLM settles the shape, or (None, question) when
    it cannot. Available for the UI to offer as a suggestion; v0.3 defaults to
    the expressible set and lets the operator pick (the LLM task-choice is a
    later slice). Thin wrapper -- not wired into the v0.3 UI by default.
    """
    return pipeline.choose_task(goal, expressible_tasks(chosen), ask_fn)


# ---------------------------------------------------------------------------
# 2-3. interpret (LLM) + propose (LLM) -- the one adaptation threads `task`
# ---------------------------------------------------------------------------

def interpret(observed_facts: list[dict], goal: str, ask_fn=ask) -> tuple[list[dict], dict]:
    """LLM inspection: INFERRED/UNKNOWN claims + the boundary ingest dict."""
    return pipeline.interpret(observed_facts, goal, ask_fn)


def propose(report: list[dict], goal: str, sources: dict, observed_facts: list[dict],
            task: str, ask_fn=ask, resumed: bool = False) -> tuple[Optional[dict], list, list]:
    """LLM task definition + triage, threading the operator-chosen `task`.

    This is the one modeller adaptation: `modeller/app.py` calls
    `pipeline.propose(...)` without `task=`, so it always models as enrichment.
    Here the chosen task family is passed through, so reconciliation /
    aggregation / reservation models can be produced from the same machinery.
    Returns (model, asked, deferred) exactly as `pipeline.propose` does.
    """
    return pipeline.propose(report, goal, sources, observed_facts, ask_fn,
                            resumed=resumed, task=task)


# ---------------------------------------------------------------------------
# 4. missing truth -- a load-bearing question the program cannot settle
# ---------------------------------------------------------------------------

def questions(asked: list, observed_facts: list[dict]) -> list:
    """Question objects for the asked block, carrying their obligation ids."""
    return pipeline.questions_from([e for e, _ in asked], observed_facts)


def build_answer(q, choice) -> str:
    """Assemble the human answer string from a UI choice, mirroring
    modeller/app.py:206-207. A join question (q.options) becomes
    `<source>.<field> matches <choice>`; a free-text question passes through."""
    if q.options:
        src = q.source[0] if isinstance(q.source, list) else q.source
        return f"{src}.{q.field} matches {choice}"
    return choice


def apply_answer(report: list[dict], q, answer: str) -> list[dict]:
    """Apply a human answer -> CONFIRMED claims (carries the obligation id back)."""
    return pipeline.submit_answer(report, q, answer)


# ---------------------------------------------------------------------------
# 5. deterministic preview + sufficiency gate
# ---------------------------------------------------------------------------

def check_join(model: dict, observed_facts: list[dict],
               report: Optional[list[dict]] = None) -> Optional[str]:
    """The program's own sufficiency check on the declared join (enrichment-
    oriented). None = OK, str = complaint. For reconciliation the authoritative
    validity gate is `preview` ok (the task's own validator via
    `builder.validate_raw`, run inside `pipeline.build`); this result is shown
    but is not the sole gate."""
    return pipeline.check_join_supported(model, observed_facts, report)


def preview(ws: pipeline.Workspace, model: dict):
    """Validate (the task's own validator) and run deterministically."""
    model = dict(model)
    model.setdefault("task", "enrichment")  # pipeline.build reads model["task"]
    return pipeline.build(ws, model)


# ---------------------------------------------------------------------------
# presentation -- the readable model. pipeline.readable is enrichment-only and
# fleet.fleet.readable covers enrichment/reservation; neither renders
# reconciliation, so render the key fields directly (presentation, not logic).
# ---------------------------------------------------------------------------

def render_model(model: dict, task: str) -> list[str]:
    """The proposed task in sentences for a person who will not read JSON."""
    if task == "enrichment":
        return pipeline.readable(model)
    lines: list[str] = []
    srcs = model.get("sources") or {}
    lines.append("Read " + " and ".join(
        f"**{name}** (`{spec.get('collection')}`)" for name, spec in srcs.items()) + ".")
    # v0.4: show the durable origin chain when present -- which workbook + sheet +
    # header_row each converted source came from. Ordinary (non-converted) JSON
    # sources carry no `origin` and show nothing (no fake provenance).
    provenance = [(name, spec.get("origin")) for name, spec in srcs.items()
                  if spec.get("origin")]
    if provenance:
        lines.append("**Origin** — where each executable source came from:")
        for name, o in provenance:
            lines.append(f"- **{name}** ← `{o.get('path')}` sheet `{o.get('sheet')}` "
                         f"(header row {o.get('header_row')})")
    if task == "reconciliation":
        m = model.get("match_on") or {}
        lines.append(f"Match **{model.get('left')}**.`{m.get('left_field')}` against "
                     f"**{model.get('right')}**.`{m.get('right_field')}`.")
        for c in model.get("compare") or []:
            how = c.get("comparison", "")
            extra = f" within {c['tolerance']}" if how == "within" else ""
            lines.append(f"Compare `{c.get('field')}` ({how}{extra}).")
        cls = model.get("classify") or {}
        if cls:
            lines.append("Classify each key as " + " / ".join(
                f"{rel}={label}" for rel, label in cls.items()) + ".")
    elif task == "aggregation":
        lines.append(f"Group **{model.get('driving_source')}** and aggregate "
                     f"(see the model JSON for the grouped outputs).")
    else:
        lines.append(f"(readable rendering for {task} is task-specific; see "
                      f"the model JSON.)")
    return lines


# ---------------------------------------------------------------------------
# 6. establishment -- the explicit human action. Wraps fleet.establish.
# ---------------------------------------------------------------------------

def establish(name: str, purpose: str, task: str, base: str, model: dict,
              trigger: str, customer: Optional[str] = None,
              destination: Optional[dict] = None,
              delivery: Optional[dict] = None,
              root: Optional[Path] = None):
    """Establish a worker in the live fleet (or `root`). Writes exactly three
    files: worker.json, versions/v1.json, history.jsonl (see fleet.fleet.establish).

    `customer`, if given, is written into worker.json post-establish (the existing
    field the other workers carry; `fleet.establish`'s signature is fixed and not
    extended). Without it the worker renders in an unscoped map band
    (`system_map.lanes` handles None) -- not a Company entity.

    `destination` / `delivery` (v0.5) are likewise written into worker.json
    post-establish, alongside `customer`. They are declared business facts (where
    the result belongs, desired delivery mode) -- NOT model fields, so they never
    reach versions/v1.json and never collide with `on_accept`/effect authority.

    `model.setdefault("task", task)` defends against a definer that omits the task
    field; the executor reads `model["task"]` via `pipeline.build`.
    """
    root = root or fleet.ROOT
    model = dict(model)
    model.setdefault("task", task)
    w = fleet.establish(root, name, purpose, task, base, model, trigger=trigger)
    if customer or destination or delivery:
        wp = w.directory / "worker.json"
        ident = json.loads(wp.read_text(encoding="utf-8"))
        if customer:
            ident["customer"] = customer
        if destination:
            ident["destination"] = destination
        if delivery:
            ident["delivery"] = delivery
        wp.write_text(json.dumps(ident, indent=2) + "\n", encoding="utf-8")
        w = fleet.load(w.directory)  # reload so identity carries the new fields
    return w


def establish_workspace(ws: pipeline.Workspace, name: str, purpose: str, task: str,
                        model: dict, customer: Optional[str] = None,
                        destination: Optional[dict] = None,
                        delivery: Optional[dict] = None,
                        root: Optional[Path] = None):
    """Establish from a Workspace, computing `base` and `trigger` from it.

    `base` is the lab-root-relative dir the model's source paths resolve against
    (= `data/` for `data/<dir>/` workspaces). `trigger` is the source data dir
    (`data/<dir>/`) -- this is what makes `incoming._link_worker` link the data
    dir to the new worker STRUCTURALLY via trigger-path containment, for free
    (the source-path provenance the incoming browser needs, without a new field).
    """
    base = str(ws.base.relative_to(LAB))
    trigger = f"{base}/{ws.rel}/"
    return establish(name, purpose, task, base, model, trigger,
                     customer=customer, destination=destination,
                     delivery=delivery, root=root)


# ===========================================================================
# v0.4 -- declared XLSX materialization with durable source provenance.
#
# An unfamiliar workbook enters the product through an explicit, refusal-safe
# declaration: discover its sheets (no LLM) -> the OPERATOR declares sheet +
# collection + header_row (the program never decides which sheets have meaning)
# -> the existing XLSX adapter validates + materializes the selected sheets into
# a separate derived area (raw workbooks untouched; a sheet that cannot be
# converted faithfully REFUSES before any LLM call) -> the materialized JSON
# feeds the UNCHANGED v0.3 journey. Each established source binding carries an
# `origin` (workbook + sheet + header_row) injected in the glue AFTER `propose`
# and BEFORE `fleet.establish` (pipeline.define() overwrites `sources` from
# source_spec, which knows only path/collection, so `origin` cannot ride through
# propose; fleet.establish writes the model VERBATIM, so the injected `origin`
# reaches disk). `path` says what the worker executes; `origin` says where that
# executable representation came from. The executor never reads `origin`.
#
# No change to adapters/xlsx.py, modeller/**, taskmodel/**, or fleet/**.
# ===========================================================================

def discover_workbook(xlsx_path: Path) -> list[dict]:
    """The structural shape of every sheet in a workbook (no LLM, no values
    interpreted): one entry per sheet with its name, row x column counts, and a
    preview of the first ~5 rows as raw lists.

    This is the program's half of the authority boundary: it may DISCOVER
    structure (sheet names, row/col counts, what the first rows look like) so the
    operator can pick a `header_row`; it never decides which sheets have business
    meaning. The preview uses the formula view (`data_only=False`) so a sheet of
    uncalculated formulas shows its `=...` strings -- the operator can see, before
    declaring, that a sheet will refuse.
    """
    import openpyxl
    out: list[dict] = []
    book = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=False)
    try:
        for name in book.sheetnames:
            sheet = book[name]
            rows = list(sheet.iter_rows(values_only=True))
            preview = [list(r) for r in rows[:5]]
            out.append({
                "name": name,
                "rows": sheet.max_row or len(rows),
                "cols": sheet.max_column or (len(rows[0]) if rows else 0),
                "preview": preview,
            })
    finally:
        book.close()
    return out


def validate_selection(xlsx_path: Path, sheet: str, header_row: int
                       ) -> tuple[bool, list[str], list[str], int]:
    """A trial conversion of one declared sheet, WITHOUT writing. Reuses the
    (now-real) adapter refusals: `uncalculated_formula`, `blank_header`,
    `duplicate_header`, `empty_sheet`, `missing_sheet`. This is the
    "validate before LLM" step -- a refused selection returns `ok=False` with the
    adapter's named problem, and the UI stops that selection before any model call.

    Returns (ok, problems, headers, rows). `headers` are the column names the
    adapter actually emitted (the named headers that produced data); `rows` is the
    record count. The adapter is the single authority for ok/refusal.
    """
    spec = xlsx_adapter.SheetSpec(sheet, "_probe", header_row)
    result = xlsx_adapter.convert(xlsx_path, [spec])
    if not result.ok:
        return False, list(result.problems), [], 0
    items = result.collections.get("_probe", [])
    headers = sorted(set().union(*[item.keys() for item in items])) if items else []
    return True, [], headers, len(items)


def materialize_selections(xlsx_path: Path, specs: list[xlsx_adapter.SheetSpec],
                            raw_rel: str, data_root: Optional[Path] = None
                            ) -> tuple[list[Path], dict]:
    """Convert + write the declared sheets into `data/_derived/<raw_rel>/` (the
    raw workbook dir stays pure). Also writes a ghost-source-safe
    `materialization.json` manifest there: its top-level keys are all non-list, so
    `pipeline.sources_in` finds no SourceFile in it (mirrors the
    `data/xlsx-purchases/adapter.json` pattern). Returns the written JSON paths +
    the collection -> `origin` map.

    `data_root` defaults to the lab `data/` dir; overridable for the self-test so
    it never pollutes the real data tree. `raw_rel` is the raw incoming dir name
    (e.g. "acme-august"); the `origin.path` points back at the RAW workbook
    (`<raw_rel>/<workbook>`), not the derived JSON.
    """
    conversion = xlsx_adapter.convert(xlsx_path, specs)
    if not conversion.ok:
        raise xlsx_adapter.UnreadableWorkbook(conversion.problems)
    base = data_root or (LAB / "data")
    derived_dir = base / "_derived" / raw_rel
    written = xlsx_adapter.write_collections(
        conversion, derived_dir,
        note=f"materialized from {xlsx_path.name} by supervisor/define.py")

    selections: dict[str, dict] = {}
    origins: dict[str, dict] = {}
    for spec, (collection, items) in zip(specs, conversion.collections.items()):
        selections[collection] = {"workbook": xlsx_path.name,
                                  "sheet": spec.sheet,
                                  "header_row": spec.header_row,
                                  "rows": len(items)}
        origins[collection] = {
            "kind": "xlsx",
            "path": f"{raw_rel}/{xlsx_path.name}",
            "sheet": spec.sheet,
            "header_row": spec.header_row,
        }
    # Merge with any prior manifest so multiple workbooks in the same incoming
    # dir accumulate their selections into one derived dir. The manifest's only
    # structural job is to be ghost-source-safe: every top-level key is non-list,
    # so `pipeline.sources_in` finds no SourceFile in it.
    manifest_path = derived_dir / "materialization.json"
    if manifest_path.is_file():
        try:
            prev = json.loads(manifest_path.read_text(encoding="utf-8"))
            selections = {**(prev.get("selections") or {}), **selections}
        except Exception:
            pass
    manifest = {
        "_note": "derived working representation; not a source collection",
        "created_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "selections": selections,
    }
    manifest_path.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return written, origins


def workspace_from_rel(rel: str, data_root: Optional[Path] = None) -> pipeline.Workspace:
    """A modeller Workspace for an explicit `data/<rel>/` (default base = the lab
    `data/` dir). The derived workspace (`rel = "_derived/<raw_rel>"`) is NOT
    produced by `pipeline.workspaces()` (one-level glob, no nesting), so it never
    appears as a ghost in the supervisor; the glue constructs it explicitly here.
    `sources_in` then resolves `data/_derived/<raw_rel>/*.json` correctly.
    """
    base = data_root or (LAB / "data")
    return pipeline.Workspace(label=rel, base=base, rel=rel)


def attach_origin(model: dict, origins: dict[str, dict]) -> dict:
    """Stamp each established source binding with its durable `origin` (workbook +
    sheet + header_row). Injected in the glue AFTER `propose` (which overwrites
    `sources` from source_spec, path/collection only) and BEFORE `fleet.establish`
    (which writes the model verbatim, so `origin` reaches `versions/v1.json`).
    The executor never reads `origin`; only the on-disk model + the map do.

    Match is by the source binding key (= the collection name). A source with no
    entry in `origins` (an ordinary JSON source, not converted) is left untouched
    -- no fake provenance. Returns a shallow-copied model; the caller's dict is
    not mutated.
    """
    out = dict(model)
    srcs = dict(out.get("sources") or {})
    for name, spec in srcs.items():
        if name in origins:
            spec = dict(spec)
            spec["origin"] = origins[name]
            srcs[name] = spec
    out["sources"] = srcs
    return out


def establish_derived(raw_rel: str, derived_ws: pipeline.Workspace, name: str,
                      purpose: str, task: str, model: dict,
                      customer: Optional[str] = None,
                      destination: Optional[dict] = None,
                      delivery: Optional[dict] = None,
                      root: Optional[Path] = None):
    """Establish a worker whose executable sources live in `_derived/` but whose
    `trigger` points at the RAW incoming dir `data/<raw_rel>/`.

    The semantic split the user froze: `path` (in the model's sources, built by
    `source_spec` on the derived workspace) says what the worker executes
    (`_derived/<raw_rel>/*.json`); `trigger` says where the operational thing that
    arrives lives (the raw workbook dir), so `incoming._link_worker` links the raw
    incoming dir to the worker -- not the derived working representation. This
    fixes the current xlsx-purchases -> june-purchases amnesia rather than
    institutionalizing it. `base = "data"` so `load_collection` resolves
    `data/_derived/<raw_rel>/*.json`.
    """
    return establish(name, purpose, task, "data", model,
                     trigger=f"data/{raw_rel}/",
                     customer=customer, destination=destination,
                     delivery=delivery, root=root)


# ---------------------------------------------------------------------------
# self-test -- deterministic spine, no LLM, no real fleet. The LLM stages
# (interpret/propose) are proven by the real acceptance run, not here.
# ---------------------------------------------------------------------------

# A valid reconciliation model over the real kesko data: match Invoice (ledger)
# against "Their ref" (statement), compare Amount within 0.01. The amounts are
# numeric strings, so on_non_numeric is required (the within comparison is
# numeric) but never fires. Keys are text, which is fine -- the policy governs
# the compared OPERAND, not the match key.
_KESKO_MODEL = {
    "model_version": 1,
    "model_id": "kesko-reconciliation-selftest",
    "task": "reconciliation",
    "sources": {
        "purchase_ledger": {"path": "kesko-reconciliation/purchase_ledger.json",
                            "collection": "purchase_ledger"},
        "supplier_statement": {"path": "kesko-reconciliation/supplier_statement.json",
                               "collection": "supplier_statement"},
    },
    "left": "purchase_ledger",
    "right": "supplier_statement",
    "match_on": {"left_field": "Invoice", "right_field": "Their ref"},
    "compare": [{"field": "Amount", "comparison": "within", "tolerance": "0.01"}],
    "on_non_numeric": "refuse_run",
    "classify": {"both_same": "SAME", "both_different": "DIFFERENT",
                 "only_left": "ONLY_LEDGER", "only_right": "ONLY_STATEMENT"},
    "output_order": "left_then_right",
    "on_duplicate_key": "refuse_run",
}


def _self_test() -> int:
    import tempfile
    failures: list[str] = []

    def check(cond: bool, msg: str) -> None:
        if not cond:
            failures.append(msg)

    # --- 1. workspace resolution + observation (no LLM) ---------------------
    ws = workspace_for("kesko-reconciliation")
    check(ws is not None, "workspace_for finds kesko-reconciliation (it is a data/ dir)")
    if ws is None:
        sys.stderr.write("SELF-TEST FAILED: kesko-reconciliation not reachable; "
                         "is data/kesko-reconciliation present?\n")
        return 1

    chosen = chosen_sources(ws)
    check(len(chosen) == 2, f"kesko has 2 collections to relate: {len(chosen)}")
    names = {c.collection for c in chosen}
    check(names == {"purchase_ledger", "supplier_statement"},
          f"collections are purchase_ledger + supplier_statement: {names}")

    obs = observed(ws, chosen)
    check(bool(obs), "observation produced OBSERVED facts")
    check(all(c["status"] == "OBSERVED" for c in obs),
          "OBSERVED is program-only -- no INFERRED/CONFIRMED from observation")

    tasks = expressible_tasks(chosen)
    check("reconciliation" in tasks,
          f"reconciliation is expressible for 2 collections: {tasks}")

    # --- 2. deterministic preview over the real kesko data -----------------
    p = preview(ws, _KESKO_MODEL)
    check(p.ok, f"the canned reconciliation model previews ok: {p.problems}")
    if p.ok:
        rows_text = json.dumps(p.rows, ensure_ascii=False)
        # PI-3301 same, PI-3303 different (119.94 vs 110.94), PI-3305 same,
        # PI-3350 only ledger, PI-3399 only statement.
        check("SAME" in rows_text and "DIFFERENT" in rows_text
              and "ONLY_LEDGER" in rows_text and "ONLY_STATEMENT" in rows_text,
              f"preview rows span all four relations: {rows_text[:300]}")

    # --- 3. render_model ----------------------------------------------------
    lines = render_model(_KESKO_MODEL, "reconciliation")
    check(bool(lines) and any("Invoice" in ln or "Their ref" in ln for ln in lines),
          f"render_model produces reconciliation sentences: {lines}")

    # --- 4. build_answer (the answer-string assembly) ----------------------
    class _FakeQ:
        options = ["supplier_statement.Their ref"]
        source = ["purchase_ledger"]
        field = "Invoice"
    check(build_answer(_FakeQ, "supplier_statement.Their ref")
          == "purchase_ledger.Invoice matches supplier_statement.Their ref",
          "a join answer assembles as <source>.<field> matches <choice>")
    class _FreeQ:
        options = []
        source = "purchase_ledger"
        field = "Booked"
    check(build_answer(_FreeQ, "a booking date") == "a booking date",
          "a free-text answer passes through unchanged")

    # --- 5. establish writes exactly 3 files + structural linkage -----------
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        w = establish_workspace(ws, "kesko-reconciliation",
                                "Reconcile the purchase ledger against the "
                                "supplier statement.", "reconciliation",
                                _KESKO_MODEL, customer="kesko", root=root)
        check(w.task == "reconciliation", f"established worker task: {w.task}")
        check(w.current_version == 1, "established at v1")
        check(len(w.history) == 1, "one history line (established)")
        check(w.trigger == "data/kesko-reconciliation/",
              f"trigger is the source data dir: {w.trigger}")
        check(w.identity.get("customer") == "kesko",
              f"customer written into worker.json: {w.identity.get('customer')}")

        # exactly three files, no more (no state/inbox/runs/investigation)
        files = sorted(p2.name for p2 in w.directory.rglob("*") if p2.is_file())
        check(files == ["history.jsonl", "v1.json", "worker.json"],
              f"establish writes exactly worker.json + versions/v1.json + "
              f"history.jsonl: {files}")

        # the model is stored verbatim in v1.json
        stored = json.loads((w.directory / "versions" / "v1.json").read_text(encoding="utf-8"))
        check(stored["task"] == "reconciliation" and stored["match_on"]["left_field"] == "Invoice",
              "the established model is stored verbatim in v1.json")

        # structural linkage: incoming.scan links the data dir to this worker
        # via trigger-path containment (the provenance that falls out of
        # setting trigger = the data dir).
        scan = incoming.scan([w], LAB / "data")
        kesko = next((e for e in scan["data_library"] if e["dir"] == "kesko-reconciliation"), None)
        check(kesko is not None, "incoming.scan lists kesko-reconciliation")
        check(kesko and kesko["worker"] == "kesko-reconciliation",
              f"the data dir links STRUCTURALLY to the new worker via trigger: "
              f"{kesko['worker'] if kesko else None}")

    # --- 6. v0.4 xlsx spine: discover / validate-refuses / materialize /
    #         derived-workspace-resolves / attach_origin / establish_derived -----
    # All in a TEMP data tree (data_root override) so the real data/ dir is never
    # touched. The LLM stages (interpret/propose) and the real incoming.scan
    # linkage of the raw dir are proven by the real acceptance run, not here;
    # this proves the deterministic spine and the provenance semantic split.
    with tempfile.TemporaryDirectory() as tmp:
        from openpyxl import Workbook
        data_root = Path(tmp) / "data"
        raw_rel = "selftest-xlsx"
        raw_dir = data_root / raw_rel
        raw_dir.mkdir(parents=True)
        xlsx_path = raw_dir / "book.xlsx"

        book = Workbook()
        clean = book.active
        clean.title = "Statement"
        clean.append(["Acme Oy -- supplier statement"])   # row 1 (title)
        clean.append([None])                                # row 2 (blank)
        clean.append(["Invoice", "Amount"])                # row 3 (header)
        clean.append(["PI-100", "119.94"])
        clean.append(["PI-101", "40"])
        broken = book.create_sheet("Broken formulas")
        broken.append(["Invoice", "Amount"])               # row 1 (header)
        broken.append(["PI-200", "=A2*2"])                 # uncalculated formula
        book.save(str(xlsx_path))
        book.close()

        # (1) discover_workbook lists both sheets with row/col counts + preview
        sheets = discover_workbook(xlsx_path)
        names = [s["name"] for s in sheets]
        check(names == ["Statement", "Broken formulas"],
              f"discover_workbook lists both sheets: {names}")
        stmt = next(s for s in sheets if s["name"] == "Statement")
        check(stmt["rows"] == 5 and stmt["cols"] == 2,
              f"Statement shape 5x2: rows={stmt['rows']} cols={stmt['cols']}")
        check(any("Invoice" in str(stmt["preview"]) for _ in [0]),
              "Statement preview carries the header row")

        # (2) validate_selection: clean sheet ok; formula sheet REFUSES
        ok, problems, headers, rows = validate_selection(xlsx_path, "Statement", 3)
        check(ok, f"clean sheet validates (header_row 3): {problems}")
        check("Invoice" in headers and "Amount" in headers,
              f"clean sheet emits the named headers: {headers}")
        check(rows == 2, f"clean sheet emits 2 data rows: {rows}")
        ok2, problems2, _, _ = validate_selection(xlsx_path, "Broken formulas", 1)
        check(not ok2, f"formula sheet must REFUSE (was ok={ok2}): {problems2}")
        check(any("uncalculated_formula" in p for p in problems2),
              f"refusal is named uncalculated_formula: {problems2}")

        # (3) materialize the clean sheet into _derived/<raw_rel>/ + manifest
        clean_spec = xlsx_adapter.SheetSpec("Statement", "statement", 3)
        written, origins = materialize_selections(
            xlsx_path, [clean_spec], raw_rel, data_root=data_root)
        check(len(written) == 1 and written[0].name == "statement.json",
              f"one derived JSON per clean collection: {written}")
        check((data_root / "_derived" / raw_rel / "materialization.json").is_file(),
              "ghost-safe materialization.json manifest written")
        # the derived workspace resolves and chosen_sources sees ONLY the clean
        # collection (no ghost SourceFile from the manifest)
        derived_ws = workspace_from_rel(f"_derived/{raw_rel}", data_root=data_root)
        derived_chosen = chosen_sources(derived_ws)
        check([c.collection for c in derived_chosen] == ["statement"],
              f"derived workspace has exactly the clean collection, no manifest "
              f"ghost: {[c.collection for c in derived_chosen]}")

        # (4) attach_origin stamps origin on the matching source, leaves others
        canned = {
            "task": "reconciliation",
            "sources": {
                "statement": {"path": f"_derived/{raw_rel}/statement.json",
                               "collection": "statement"},
                "ledger": {"path": "selftest-xlsx/ledger.json",
                           "collection": "ledger"},   # ordinary JSON, no origin
            },
        }
        stamped = attach_origin(canned, origins)
        check(stamped["sources"]["statement"]["origin"]["sheet"] == "Statement"
              and stamped["sources"]["statement"]["origin"]["header_row"] == 3
              and stamped["sources"]["statement"]["origin"]["path"]
              == f"{raw_rel}/book.xlsx",
              f"attach_origin stamps the exact workbook/sheet/header_row origin: "
              f"{stamped['sources']['statement'].get('origin')}")
        check("origin" not in stamped["sources"]["ledger"],
              "an ordinary (non-converted) source gets no fake provenance")
        check("origin" not in canned["sources"]["statement"],
              "attach_origin does not mutate the caller's model")

        # (5) establish_derived: trigger points at the RAW dir (not _derived),
        #     and the injected origin survives verbatim to versions/v1.json.
        #     v0.5: a declared destination/delivery ride on worker.json identity
        #     (like customer), NOT on the model, so v1.json stays pure task
        #     semantics (acceptance C + the task-semantics != destination split).
        fleet_root = Path(tmp) / "fleet"
        w = establish_derived(raw_rel, derived_ws, "selftest-xlsx-recon",
                              "Reconcile the supplier statement.", "reconciliation",
                              stamped, customer="acme",
                              destination={"system": "finance", "area": "reskontra"},
                              delivery={"mode": "review"}, root=fleet_root)
        check(w.trigger == f"data/{raw_rel}/",
              f"trigger points at the RAW dir, not _derived: {w.trigger}")
        check(w.identity.get("destination") ==
              {"system": "finance", "area": "reskontra"},
              f"destination written into worker.json: {w.identity.get('destination')}")
        check(w.identity.get("delivery") == {"mode": "review"},
              f"delivery written into worker.json: {w.identity.get('delivery')}")
        stored = json.loads((w.directory / "versions" / "v1.json")
                            .read_text(encoding="utf-8"))
        check("destination" not in stored and "delivery" not in stored,
              "CANARY: destination/delivery are identity fields, NOT model "
              "fields -- v1.json must carry only task semantics")
        stmt_path = stored["sources"]["statement"]["path"]
        check(stmt_path == f"_derived/{raw_rel}/statement.json",
              f"the executable source path points at the derived JSON: {stmt_path}")
        check(stored["sources"]["statement"]["origin"]["sheet"] == "Statement"
              and stored["sources"]["statement"]["origin"]["path"]
              == f"{raw_rel}/book.xlsx",
              f"the exact XLSX origin survives verbatim to v1.json: "
              f"{stored['sources']['statement'].get('origin')}")

    if failures:
        sys.stderr.write("SELF-TEST FAILED:\n  " + "\n  ".join(failures) + "\n")
        return 1
    print("SELF-TEST PASSED (workspace resolution + OBSERVED-only observation / "
          "reconciliation expressible / canned model previews ok with all four "
          "relations / render_model / build_answer join + free-text / establish "
          "writes exactly 3 files + customer / trigger links the data dir "
          "structurally / v0.4 xlsx spine: discover lists sheets+shape / "
          "validate REFUSES an uncalculated formula by name / materialize writes "
          "derived JSON + ghost-safe manifest and the derived workspace sees only "
          "the clean collection / attach_origin stamps exact workbook+sheet+"
          "header_row provenance and leaves ordinary sources alone / "
          "establish_derived trigger points at the RAW dir while the source path "
          "points at _derived and the origin survives verbatim to v1.json). "
          "LLM stages (interpret/propose) + the real incoming.scan linkage of the "
          "raw dir are proven by the real acceptance run, not here.")
    return 0


if __name__ == "__main__":
    raise SystemExit(_self_test() if sys.argv[1:2] == ["--self-test"] else 2)