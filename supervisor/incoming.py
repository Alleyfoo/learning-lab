#!/usr/bin/env python3
"""Workspace v0.2 -- the read-only incoming-data browser.

The company's actual incoming data, in two parts:

  data_library -- the `data/` tree: one entry per top-level subdir, with its files
                  (workbooks -> sheet names, JSON, other), `adapter.json` /
                  `established_model.json` badges, and a best-effort *worker link*.
  inboxes      -- each worker's inbox/processed/exceptions files (the granular file
                  view behind the System Map's `input:` nodes), with xlsx sheet names.

Read-only: never writes. It does **not** import `fleet` -- it duck-types the worker
objects the app passes in (`.directory: Path`, `.identity: dict`, `.name: str`), so the
inherited fleet package stays untouched. Uses `openpyxl` (already a dep via
`adapters/xlsx.py`) in read-only mode to enumerate xlsx sheet names.

The worker link is best-effort and honest. A `data/` dir is linked to a worker if
**either** (a) the worker's `trigger` folder (any ` (*.xlsx)` glob stripped, resolved
against the lab root) is equal to or inside that dir, **or** (b) a file name in the dir
also appears in that worker's `inbox/`/`processed/`/`exceptions/`. `worker = None`
means "no link found" -- rendered by the app as a neutral **"no worker link"**, NOT a
confident "not yet modelled". Some unlinked dirs (e.g. `data/april-invoicing/`,
`data/timesheets/`) are seed/reference JSON for modelled workers whose structural link
is not machine-findable. The genuine "model exists but not deployed as a worker" gap
(`data/xlsx-purchases/`, `data/xlsx-statement/`) is surfaced honestly via the
`has_model` badge on a dir with no worker link -- the point of the browser is to
*include* not-yet-modelled data, not to perfectly classify every file.
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

HERE = Path(__file__).resolve().parent
LAB = HERE.parent


def _sheet_names(path: Path) -> list[str]:
    """The sheet names of an xlsx workbook (read-only), or [] on any failure."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return []


def _kind(path: Path) -> str:
    s = path.suffix.lower()
    if s == ".xlsx":
        return "xlsx"
    if s == ".json":
        return "json"
    return "other"


def _file_entry(path: Path) -> dict:
    kind = _kind(path)
    return {
        "name": path.name,
        "kind": kind,
        "sheets": _sheet_names(path) if kind == "xlsx" else [],
    }


def _trigger_path(identity: dict) -> Optional[Path]:
    """The worker's trigger folder as a Path: strip any ` (*.xlsx)` glob suffix,
    resolve relative paths against the lab root. None if no trigger is declared."""
    trig = identity.get("trigger")
    if not trig:
        return None
    raw = str(trig).split("(")[0].strip().rstrip("/\\")
    if not raw:
        return None
    p = Path(raw)
    if not p.is_absolute():
        p = LAB / p
    try:
        return p.resolve()
    except Exception:
        return p


def _stage_files(w_dir: Path, stage: str) -> list[dict]:
    """File entries (with sheets for xlsx) in one inbox-stage folder, sorted by name."""
    d = w_dir / stage
    if not d.is_dir():
        return []
    return [_file_entry(p) for p in sorted(d.iterdir()) if p.is_file()]


def _inbox_file_names(w_dir: Path) -> set[str]:
    """All file names across a worker's inbox/processed/exceptions (for cross-ref)."""
    out: set[str] = set()
    for stage in ("inbox", "processed", "exceptions"):
        d = w_dir / stage
        if d.is_dir():
            out.update(p.name for p in d.iterdir() if p.is_file())
    return out


def _tokens(name: str) -> set[str]:
    """Lowercase alphanumeric tokens of a name (for the cross-ref tiebreak)."""
    import re
    return set(re.findall(r"[a-z0-9]+", (name or "").lower()))


def _worker_dir(w) -> Optional[Path]:
    wd = getattr(w, "directory", None)
    try:
        wd = Path(wd) if wd is not None else None
    except Exception:
        wd = None
    return wd if (wd is not None and wd.is_dir()) else None


def _link_worker(data_dir: Path, data_file_names: set[str],
                  workers: list) -> Optional[str]:
    """The name of a worker linked to this `data/` dir, or None.

    Link if (a) the worker's trigger path is equal to or inside `data_dir`
    (the worker's incoming folder is this dir or lives within it), or (b) a file name
    in the dir also appears in that worker's inbox/processed/exceptions -- in which case
    the worker with the MOST matching file names wins; a tie is broken by name-token
    overlap with the dir name, and a still-tied/empty result is None (genuinely
    ambiguous, rendered as a neutral "no worker link" rather than a guessed one).
    """
    try:
        data_resolved = data_dir.resolve()
    except Exception:
        data_resolved = data_dir

    # (a) trigger-path containment (reliable -- declared in worker.json)
    for w in workers:
        tp = _trigger_path((getattr(w, "identity", None) or {}))
        if tp is not None and (tp == data_resolved or data_resolved in tp.parents):
            return w.name

    # (b) filename cross-reference: most matches wins; tiebreak by name-token overlap
    counts: list[tuple[int, object]] = []
    for w in workers:
        wd = _worker_dir(w)
        if wd is None:
            continue
        n = len(_inbox_file_names(wd) & data_file_names)
        if n:
            counts.append((n, w))
    if not counts:
        return None
    top = max(n for n, _ in counts)
    winners = [w for n, w in counts if n == top]
    if len(winners) == 1:
        return winners[0].name
    # tie -> name-token overlap with the dir name
    dir_tokens = _tokens(data_dir.name)
    best: object = None
    best_overlap = 0
    for w in winners:
        ov = len(dir_tokens & _tokens(w.name))
        if ov > best_overlap:
            best, best_overlap = w, ov
    return best.name if best is not None and best_overlap > 0 else None


def scan(workers: list, data_root: Path) -> dict:
    """The company's incoming data.

    Returns:
      data_library: one entry per top-level subdir of `data_root`:
        {dir, files: [{name, kind, sheets}], has_adapter, has_model, worker: str|None}
      inboxes: one entry per worker that actually has inbox/processed/exceptions files:
        {worker, customer, files: [{name, kind, stage, sheets}]}
    """
    data_library: list[dict] = []
    if data_root.is_dir():
        for d in sorted(p for p in data_root.iterdir() if p.is_dir()):
            files = [_file_entry(p) for p in sorted(d.iterdir()) if p.is_file()]
            data_library.append({
                "dir": d.name,
                "files": files,
                "has_adapter": (d / "adapter.json").is_file(),
                "has_model": (d / "established_model.json").is_file(),
                "worker": _link_worker(d, {f["name"] for f in files}, workers),
            })

    inboxes: list[dict] = []
    for w in workers:
        w_dir = _worker_dir(w)
        if w_dir is None:
            continue
        files: list[dict] = []
        for stage in ("inbox", "processed", "exceptions"):
            for fe in _stage_files(w_dir, stage):
                files.append({"name": fe["name"], "kind": fe["kind"],
                               "stage": stage, "sheets": fe["sheets"]})
        if files:
            inboxes.append({
                "worker": getattr(w, "name", w_dir.name),
                "customer": (getattr(w, "identity", None) or {}).get("customer"),
                "files": files,
            })

    return {"data_library": data_library, "inboxes": inboxes}


# --- self-test (no model, no real fleet, no real data/) ---------------------

class _FakeWorker:
    """Duck-typed stand-in for fleet.Worker: .directory, .identity, .name."""
    def __init__(self, name: str, directory: Path, identity: dict):
        self.name = name
        self.directory = directory
        self.identity = identity


def _self_test() -> int:
    import shutil
    import tempfile

    failures: list[str] = []

    def check(cond: bool, msg: str) -> None:
        if not cond:
            failures.append(msg)

    tmp = Path(tempfile.mkdtemp())
    data_root = tmp / "data"
    workers_root = tmp / "workers"
    try:
        # --- build the data/ tree -------------------------------------------
        from openpyxl import Workbook

        # linked-by-trigger/: a json file; a worker's trigger will point here
        (data_root / "linked-by-trigger").mkdir(parents=True)
        (data_root / "linked-by-trigger" / "orders.json").write_text("[]", encoding="utf-8")

        # linked-by-filename/: an xlsx with 2 named sheets + a json
        (data_root / "linked-by-filename").mkdir(parents=True)
        wb = Workbook()
        wb.active.title = "Order lines"
        ws2 = wb.create_sheet("Price list")
        wb.save(str(data_root / "linked-by-filename" / "shared.xlsx"))
        wb.close()
        (data_root / "linked-by-filename" / "note.json").write_text("{}", encoding="utf-8")

        # unlinked/: a json file, no worker references it
        (data_root / "unlinked").mkdir(parents=True)
        (data_root / "unlinked" / "raw.json").write_text("[]", encoding="utf-8")

        # has-model/: adapter.json + established_model.json + a json, no worker
        (data_root / "has-model").mkdir(parents=True)
        (data_root / "has-model" / "adapter.json").write_text("{}", encoding="utf-8")
        (data_root / "has-model" / "established_model.json").write_text("{}", encoding="utf-8")
        (data_root / "has-model" / "data.json").write_text("[]", encoding="utf-8")

        # --- build fake workers ---------------------------------------------
        # trig-worker: trigger points at linked-by-trigger/ (no inbox files)
        tw_dir = workers_root / "trig-worker"
        (tw_dir).mkdir(parents=True)
        trig_worker = _FakeWorker(
            "trig-worker", tw_dir,
            {"trigger": str(data_root / "linked-by-trigger"), "customer": "Acme"})

        # file-worker: processed/shared.xlsx (filename cross-ref to linked-by-filename)
        fw_dir = workers_root / "file-worker"
        (fw_dir / "processed").mkdir(parents=True)
        # a real xlsx so _stage_files enumerates sheets without error
        wb2 = Workbook(); wb2.active.title = "Order lines"
        wb2.save(str(fw_dir / "processed" / "shared.xlsx")); wb2.close()
        file_worker = _FakeWorker(
            "file-worker", fw_dir,
            {"trigger": str(fw_dir / "inbox"), "customer": "Fazerish"})

        # empty-worker: no inbox/processed/exceptions files -> must NOT appear in inboxes
        ew_dir = workers_root / "empty-worker"
        (ew_dir).mkdir(parents=True)
        empty_worker = _FakeWorker(
            "empty-worker", ew_dir, {"trigger": str(ew_dir / "inbox"), "customer": "Demo"})

        # two workers that BOTH match the same file count on tied-dir/ -> the one
        # whose name shares a token with the dir wins; a no-token-overlap tie -> None
        (data_root / "tied-fazerish").mkdir(parents=True)
        (data_root / "tied-fazerish" / "may-order-lines.xlsx").write_text("x", encoding="utf-8")
        (data_root / "tied-ambiguous").mkdir(parents=True)
        (data_root / "tied-ambiguous" / "may-order-lines.xlsx").write_text("x", encoding="utf-8")

        apr_dir = workers_root / "april-invoicing"
        (apr_dir / "processed").mkdir(parents=True)
        (apr_dir / "processed" / "may-order-lines.xlsx").write_text("x", encoding="utf-8")
        april_worker = _FakeWorker("april-invoicing", apr_dir,
                                   {"trigger": str(apr_dir / "inbox"), "customer": "Demo"})

        faz_dir = workers_root / "fazerish-invoicing"
        (faz_dir / "processed").mkdir(parents=True)
        (faz_dir / "processed" / "may-order-lines.xlsx").write_text("x", encoding="utf-8")
        fazerish_worker = _FakeWorker("fazerish-invoicing", faz_dir,
                                      {"trigger": str(faz_dir / "inbox"), "customer": "Fazerish"})

        workers = [trig_worker, file_worker, empty_worker, april_worker, fazerish_worker]
        result = scan(workers, data_root)

        # --- data_library: all dirs, sorted --------------------------------
        lib = result["data_library"]
        dirs = [e["dir"] for e in lib]
        expected = sorted(["linked-by-trigger", "linked-by-filename", "unlinked",
                            "has-model", "tied-fazerish", "tied-ambiguous"])
        check(dirs == expected,
              f"data_library lists all top-level data/ dirs sorted: {dirs}")

        by_dir = {e["dir"]: e for e in lib}

        # linked-by-trigger -> worker via trigger containment
        check(by_dir["linked-by-trigger"]["worker"] == "trig-worker",
              f"trigger containment links linked-by-trigger to trig-worker: "
              f"{by_dir['linked-by-trigger']['worker']}")

        # linked-by-filename -> worker via filename cross-ref; xlsx sheets enumerated
        lbf = by_dir["linked-by-filename"]
        check(lbf["worker"] == "file-worker",
              f"filename cross-ref links linked-by-filename to file-worker: {lbf['worker']}")
        xlsx_entry = next(f for f in lbf["files"] if f["name"] == "shared.xlsx")
        check(xlsx_entry["kind"] == "xlsx" and xlsx_entry["sheets"] == ["Order lines", "Price list"],
              f"xlsx sheets enumerated: {xlsx_entry}")
        check(not lbf["has_adapter"] and not lbf["has_model"],
              "linked-by-filename has no adapter/model badges")

        # unlinked -> worker None
        check(by_dir["unlinked"]["worker"] is None,
              f"unlinked dir has no worker link (None): {by_dir['unlinked']['worker']}")

        # has-model -> has_adapter + has_model True, worker None
        hm = by_dir["has-model"]
        check(hm["has_adapter"] and hm["has_model"] and hm["worker"] is None,
              f"has-model dir: badges True, no worker link: {hm}")

        # tied-fazerish: april and fazerish both match 1 file -> name-token tiebreak
        # picks fazerish-invoicing (shares "fazerish" with the dir name)
        check(by_dir["tied-fazerish"]["worker"] == "fazerish-invoicing",
              f"cross-ref tie broken by name-token overlap -> fazerish-invoicing: "
              f"{by_dir['tied-fazerish']['worker']}")

        # tied-ambiguous: same 1-file tie but no token overlap with either worker -> None
        check(by_dir["tied-ambiguous"]["worker"] is None,
              f"cross-ref tie with no name-token overlap -> None (ambiguous): "
              f"{by_dir['tied-ambiguous']['worker']}")

        # --- inboxes: only workers with inbox/processed/exceptions files -----
        ib = result["inboxes"]
        names = [e["worker"] for e in ib]
        check(names == ["file-worker", "april-invoicing", "fazerish-invoicing"],
              f"inboxes list only workers with inbox files, in worker order: {names}")
        check("empty-worker" not in names and "trig-worker" not in names,
              "workers with no inbox files (empty-worker, trig-worker) are absent")
        fw_ib = ib[0]
        check(fw_ib["customer"] == "Fazerish", f"inbox carries customer: {fw_ib['customer']}")
        fe = fw_ib["files"][0]
        check(fe["name"] == "shared.xlsx" and fe["stage"] == "processed"
              and fe["kind"] == "xlsx" and fe["sheets"] == ["Order lines"],
              f"inbox file carries stage + sheets: {fe}")

        # --- empty data_root -> empty data_library, no crash -----------------
        empty_result = scan(workers, tmp / "no-such-data")
        check(empty_result["data_library"] == []
              and [e["worker"] for e in empty_result["inboxes"]] == names,
              "scanning a missing data_root yields empty data_library, inboxes still work")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    if failures:
        sys.stderr.write("SELF-TEST FAILED:\n  " + "\n  ".join(failures) + "\n")
        return 1
    print("SELF-TEST PASSED (data_library lists data/ dirs sorted / "
          "worker link via trigger containment AND filename cross-ref / "
          "xlsx sheets enumerated / unlinked dir -> worker None / "
          "has-model badges True with no worker link / "
          "inboxes list only workers with files + carry stage + sheets / "
          "missing data_root -> empty data_library, no crash)")
    return 0


if __name__ == "__main__":
    raise SystemExit(_self_test() if sys.argv[1:2] == ["--self-test"] else 2)