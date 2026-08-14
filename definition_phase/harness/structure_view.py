#!/usr/bin/env python3
"""The structure view: a workbook rendered as cell TYPES, never cell text.

One implementation, imported by both the browser and Experiment N. Experiment M
found the cost of the alternative: a capability living in one layer and not
matching another, silently.

    #  number      A  text      ·  blank      ƒ  formula      ?  other

The security property is not measured, it is **structural**: the output alphabet
is exactly `# A · ƒ ?` plus column letters, row numbers and layout punctuation.
No cell content can appear in it, so no payload in any cell can reach a reader of
this view — including an agent. `contains_no_content()` asserts that against a
real workbook rather than leaving it as a claim.

What survives the projection is shape: where text stops and numbers begin, which
is what locating a header row and a data region actually depends on.
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Iterable

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from referents import index0_to_col  # noqa: E402

GLYPHS = {"#": "number", "A": "text", "·": "blank", "ƒ": "formula", "?": "other"}
ALPHABET = set("#A·ƒ?")


def glyph(value: object) -> str:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return "·"
    if isinstance(value, str) and value.startswith("="):
        return "ƒ"
    if isinstance(value, bool):
        return "?"
    if isinstance(value, (int, float)):
        return "#"
    if isinstance(value, str):
        return "A"
    return "?"


def sheet_rows(ws) -> list[list[object]]:
    return [[c.value for c in row] for row in ws.iter_rows()]


def render_structure(rows: Iterable[Iterable[object]], n_cols: int) -> str:
    """Glyph grid with A1 row numbers and column letters."""
    rows = [list(r) for r in rows]
    header = "     " + " ".join(index0_to_col(c) for c in range(n_cols))
    out = [header]
    for i, row in enumerate(rows):
        cells = [glyph(row[c] if c < len(row) else None) for c in range(n_cols)]
        out.append(f"{i + 1:>4} " + " ".join(cells))
    return "\n".join(out)


def render_values(rows: Iterable[Iterable[object]], n_cols: int) -> str:
    """Raw values, exactly as an agent reading the workbook would see them --
    including anything invisible on screen. UNTRUSTED."""
    rows = [list(r) for r in rows]
    out = ["     " + " | ".join(index0_to_col(c) for c in range(n_cols))]
    for i, row in enumerate(rows):
        cells = []
        for c in range(n_cols):
            v = row[c] if c < len(row) else None
            cells.append("" if v is None else str(v))
        out.append(f"{i + 1:>4} " + " | ".join(cells))
    return "\n".join(out)


def contains_no_content(structure: str, rows: Iterable[Iterable[object]]) -> bool:
    """No cell's text appears in the structure rendering.

    Checked against the actual cell values rather than asserted: every token of
    every string cell (length >= 3, to skip fragments that collide with column
    letters) must be absent from the rendered view.
    """
    body = "\n".join(structure.split("\n")[1:])      # drop the column-letter row
    body = "".join(ch for ch in body if not ch.isdigit())
    for row in rows:
        for value in row:
            if not isinstance(value, str):
                continue
            for token in value.split():
                if len(token) >= 3 and token.lower() in body.lower():
                    return False
    return True


def _self_test() -> int:
    from openpyxl import load_workbook

    failures: list[str] = []
    root = HERE.parent
    for name in ("W1_multisheet.xlsx", "W2_hidden_content.xlsx"):
        path = root / "fixtures" / name
        if not path.exists():
            failures.append(f"missing fixture {name}")
            continue
        wb = load_workbook(path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = sheet_rows(ws)
            view = render_structure(rows, ws.max_column or 0)
            body = "\n".join(view.split("\n")[1:])
            stray = {ch for ch in body
                     if not ch.isdigit() and ch not in ALPHABET and ch not in " \n"}
            if stray:
                failures.append(f"{name}:{sheet} structure view leaked {stray!r}")
            if not contains_no_content(view, rows):
                failures.append(f"{name}:{sheet} structure view contains cell content")

    # The load-bearing case: W2 carries an injection payload in white text, and
    # it must not survive the projection.
    w2 = load_workbook(root / "fixtures" / "W2_hidden_content.xlsx")["Sales"]
    rows = sheet_rows(w2)
    view = render_structure(rows, w2.max_column or 0)
    if "ignore previous instructions" in view.lower():
        failures.append("the injection payload survived into the structure view")
    if "ignore previous instructions" not in render_values(rows, w2.max_column or 0).lower():
        failures.append("the values view should contain the payload (it is the "
                        "untrusted view) -- the contrast is the point")

    if failures:
        sys.stderr.write("SELF-TEST FAILED:\n  " + "\n  ".join(failures) + "\n")
        return 1
    sys.stdout.write(
        "SELF-TEST PASSED (structure view alphabet is exactly '# A · ƒ ?' on every "
        "sheet of W1 and W2 / no cell content survives the projection / the "
        "injection payload is absent from the structure view and present in the "
        "values view)\n")
    return 0


if __name__ == "__main__":
    if sys.argv[1:2] == ["--self-test"]:
        raise SystemExit(_self_test())
    if sys.argv[1:]:
        from openpyxl import load_workbook

        wb = load_workbook(sys.argv[1])
        ws = wb[sys.argv[2]] if len(sys.argv) > 2 else wb[wb.sheetnames[0]]
        print(render_structure(sheet_rows(ws), ws.max_column or 0))
        raise SystemExit(0)
    sys.stderr.write("usage: structure_view.py --self-test | <workbook.xlsx> [sheet]\n")
    raise SystemExit(2)
