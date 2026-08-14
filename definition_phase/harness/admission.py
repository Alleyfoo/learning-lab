#!/usr/bin/env python3
"""The admission boundary — what the system is allowed to lose on the way in.

Pinning the observation boundary for No Undeclared Interpretation exposed
something larger than the whitespace bug it was built to catch. That invariant
compares:

```text
typed value admitted by ingestion  ->  typed value emitted by execution
```

which is the right boundary for *executor* interpretation, and cannot by
construction detect semantics destroyed **before** admission. The formula case
said so plainly:

```text
source cell:      formula "=SUM(A1:A2)"
cached value:     unavailable
ingestion admits: None
execution emits:  ""
```

Given `None`, the executor did nothing wrong and the invariant correctly stayed
quiet. The larger system still collapsed *a formula whose result is unavailable*
into *an empty cell*, and those are not the same source fact.

So there are two conservation boundaries, not one:

```text
SOURCE  ->  INGESTION  ->  EXECUTION
        A              B
```

**A — No Silent Loss on Admission.** Source properties within the language's
semantic budget must be preserved, explicitly normalised, or explicitly declared
unavailable/unsupported.

**B — No Undeclared Interpretation.** Admitted values may acquire only semantics
authorised by the accepted recipe.

Without A, cross-sheet testing would inherit lossy ingestion, and every
downstream conservation invariant could truthfully report green over an already
corrupted representation.

## What this module does NOT decide

It deliberately does not answer *what a formula means to this language*. Three
contracts are coherent:

```text
A. recipes operate only on cached values; a formula without cache is
   explicitly UNAVAILABLE
B. formulas are source semantics; the expression is preserved separately
   from the cached value
C. formulas are unsupported; a source region containing one is refused
```

What is *not* coherent is deciding this accidentally, via whatever `openpyxl`
happens to return. `SourceCell` exists so the distinction survives long enough
for the architecture to make that choice deliberately. None of it is exposed to
the recipe language yet.
"""
from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

CELL_KINDS = ("literal", "formula", "empty")


@dataclass(frozen=True)
class SourceCell:
    """What the source actually contains, before any construct interprets it.

    Not all of this is exposed to the recipe language, and that is the point:
    ingestion must not destroy a distinction before the architecture has decided
    whether it matters.
    """
    kind: str                                  # literal | formula | empty
    typed_value: Any                           # the cached/literal value, or None
    formula_expression: Optional[str] = None   # present iff kind == "formula"
    cached_value_available: bool = True

    @property
    def semantically_unavailable(self) -> bool:
        """A source fact exists that the system does not possess."""
        return self.kind == "formula" and not self.cached_value_available


def read_source_cells(path: Path, sheet: str, row0: int) -> list[SourceCell]:
    """Read one row at the ADMISSION boundary, losing nothing.

    Two passes are required and this is the whole reason the function exists:
    `data_only=True` yields cached results and *silently discards the formula*,
    `data_only=False` yields formulas and never the results. Either pass alone
    destroys a distinction; only together do they establish one.
    """
    from openpyxl import load_workbook

    cached = load_workbook(path, data_only=True)[sheet]
    formulas = load_workbook(path, data_only=False)[sheet]
    one_based = row0 + 1
    n_cols = max(int(cached.max_column or 0), int(formulas.max_column or 0))

    out: list[SourceCell] = []
    for col in range(1, n_cols + 1):
        value = cached.cell(row=one_based, column=col).value
        raw = formulas.cell(row=one_based, column=col).value
        is_formula = isinstance(raw, str) and raw.startswith("=")
        if is_formula:
            out.append(SourceCell(kind="formula", typed_value=value,
                                  formula_expression=raw,
                                  cached_value_available=value is not None))
        elif value is None and raw is None:
            out.append(SourceCell(kind="empty", typed_value=None))
        else:
            out.append(SourceCell(kind="literal", typed_value=value))
    return out


def admission_loss(source: SourceCell, admitted: Any) -> Optional[str]:
    """Boundary A. Did admission destroy a source property without saying so?

    The first implementation distinguishes only literal / formula /
    value-unavailable. That is deliberately the smallest thing that makes the
    boundary observable at all; a richer budget (number formats, merged ranges,
    data validation) can be added later without moving the boundary.
    """
    if source.semantically_unavailable and admitted is None:
        return (f"ADMISSION LOSS: source is a formula {source.formula_expression!r} "
                f"whose result is unavailable, and admission collapsed it to "
                f"{admitted!r} — indistinguishable from an empty cell. A source "
                f"fact exists that the system does not possess, and nothing "
                f"records that.")
    if source.kind == "formula" and source.formula_expression and admitted is not None:
        # Cached value present: the VALUE survives. The expression does not, but
        # contract A ("recipes operate on cached values") permits that, so it is
        # not a loss until the language says formulas are source semantics.
        return None
    if source.kind == "literal" and admitted is None:
        return (f"ADMISSION LOSS: literal source {source.typed_value!r} was "
                f"admitted as None")
    return None


def _canary_admission(tmp: Path) -> tuple[bool, bool, str]:
    """Known violation: a formula with no cached result, admitted as None.

    Reachability matters as much as firing — a canary that never delivers its
    stimulus to the observation point proves nothing about the detector.
    """
    from openpyxl import Workbook, load_workbook

    path = tmp / "canary_admission.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B"])
    ws.append([1, "=SUM(A1:A2)"])
    wb.save(path)

    cells = read_source_cells(path, "S", 1)
    cell = cells[1]
    reached = cell.kind == "formula" and not cell.cached_value_available
    if not reached:
        return False, False, f"CANARY UNREACHABLE: source read as {cell.kind}"

    admitted = load_workbook(path, data_only=True)["S"].cell(row=2, column=2).value
    why = admission_loss(cell, admitted)
    fired = why is not None

    # INVERSE CONTROL: a plain literal must NOT be reported as a loss, otherwise
    # the detector would only be proving that cells exist.
    control = admission_loss(cells[0], 1)
    if control is not None:
        return reached, False, f"INVERSE CONTROL FAILED: literal reported as loss: {control}"
    return reached, fired, (why or "detector did not fire")[:96]


def _self_test() -> int:
    import tempfile

    with tempfile.TemporaryDirectory() as td:
        reached, fired, detail = _canary_admission(Path(td))
    sys.stdout.write(f"  canary admission_loss   reached={reached}  fired={fired}\n"
                     f"    {detail}\n\n")
    if not (reached and fired):
        sys.stdout.write("ADMISSION BOUNDARY FAILED\n")
        return 1
    sys.stdout.write(
        "ADMISSION BOUNDARY established — a formula whose result is unavailable is\n"
        "distinguishable from an empty cell at the point of admission.\n\n"
        "  NOT decided here: what a formula MEANS to this language. Three contracts\n"
        "  are coherent; deciding by accident is the only incoherent option.\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(_self_test())
