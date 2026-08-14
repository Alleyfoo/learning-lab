#!/usr/bin/env python3
"""Referent grammar v1 — parser, canonical renderer, comparison key, resolver.

Spec: `design/referent_grammar_v1.md`. The spec is authority; if this file and
the spec disagree, this file is wrong.

A referent is a deterministic address into a workbook — the primitive every
definition-phase layer speaks. Surface form is A1-style; ALL INDICES ARE
0-BASED, so they drop straight into pandas/openpyxl consumers without
arithmetic.

    workbook:                      the workbook under definition
    sheetset:Months                a declared set of sheets
    sheet:Sales                    a whole sheet
    sheet:Sales!D5                 cell          sheet:Sales!D5:P96   region
    sheet:Sales!5                  row           sheet:Sales!5:96     row range
    sheet:Sales!D                  column        sheet:Sales!D:P      column range
    sheet:Sales!@Myynti            column BY HEADER NAME
    sheet:'Myynti 2026'!A1         quoted sheet name

TWO NUMBER SPACES, ONE BOUNDARY (spec sec.5)
--------------------------------------------
A1 notation is 1-based by definition: `D5` is what Excel shows in row 5,
column D. Python, pandas, and this project's recipes are 0-based. Both are kept,
and the conversion happens in exactly two functions -- `parse()` and `render()`:

    "sheet:Sales!D5"  --parse-->  row0=4, col0=3   --render-->  "sheet:Sales!D5"

Every field and parameter carrying an index is suffixed `0` (`row0`, `col0`,
`header_rows0`) so the convention is visible at each call site rather than
remembered. Nothing between parse and render ever sees a 1-based number.

Binding mode is explicit and load-bearing: `!D` survives a rename and breaks on
an inserted column; `!@Myynti` survives an insert and breaks on a rename. The
grammar refuses to hide which one a recipe depends on (see spec sec.3 and
repo_reuse_map.md defect D1 -- positional mapping that silently shifts).

Resolution failures are a frozen enum and are HARD ERRORS, never warnings. Two
of them are judgement calls made deliberately:
  * `header_ambiguous` -- two columns share the header, so the reference is
    refused rather than resolved to the first match. Picking one would assert
    what the evidence does not establish.
  * `header_row_not_declared` -- `@Myynti` is meaningless until someone says
    which row the headers are on, so the recipe is forced to have declared it.
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Mapping, Optional, Sequence

ROOT = Path(__file__).resolve().parents[1]

MAX_COL = 16384  # XFD, Excel's limit (a COUNT, so valid col0 is 0..MAX_COL-1)
KINDS = (
    "workbook", "sheetset", "sheet",
    "cell", "cellrange", "row", "rowrange", "col", "colrange", "namedcol",
)
REASONS = (
    "malformed",
    "sheet_not_found",
    "sheetset_not_declared",
    "out_of_bounds",
    "header_row_not_declared",
    "header_not_found",
    "header_ambiguous",
)
# Sheet names containing any of these must be quoted.
_NEEDS_QUOTE = set(" !:'")


class ReferentSyntaxError(ValueError):
    """Raised by parse(). A malformed referent never becomes an object."""


# ---------------------------------------------------------------------------
# Column letters <-> 0-based index.  A->0, Z->25, AA->26, XFD->16383
# ---------------------------------------------------------------------------

def col_to_index0(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        if not ("A" <= ch <= "Z"):
            raise ReferentSyntaxError(f"bad column letters: {letters!r}")
        n = n * 26 + (ord(ch) - 64)
    if not 1 <= n <= MAX_COL:
        raise ReferentSyntaxError(f"column out of range: {letters!r}")
    return n - 1


def index0_to_col(index0: int) -> str:
    if not 0 <= index0 < MAX_COL:
        raise ReferentSyntaxError(f"column index out of range: {index0}")
    n = index0 + 1
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


# ---------------------------------------------------------------------------
# The object — every index is 0-based and inclusive
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class Referent:
    kind: str
    sheet: Optional[str] = None
    name: Optional[str] = None            # sheetset name, or header text
    row0: Optional[int] = None            # first row, 0-based
    row0_last: Optional[int] = None       # last row, 0-based INCLUSIVE
    col0: Optional[int] = None
    col0_last: Optional[int] = None

    def render(self) -> str:
        """Canonical A1 surface form (converts 0-based back to 1-based)."""
        if self.kind == "workbook":
            return "workbook:"
        if self.kind == "sheetset":
            return f"sheetset:{self.name}"
        head = f"sheet:{_render_sheet_name(self.sheet or '')}"
        if self.kind == "sheet":
            return head
        return f"{head}!{self._render_span()}"

    def _render_span(self) -> str:
        k = self.kind
        if k == "namedcol":
            return f"@{self.name}"
        if k == "cell":
            return f"{index0_to_col(self.col0)}{self.row0 + 1}"
        if k == "cellrange":
            return (f"{index0_to_col(self.col0)}{self.row0 + 1}:"
                    f"{index0_to_col(self.col0_last)}{self.row0_last + 1}")
        if k == "row":
            return str(self.row0 + 1)
        if k == "rowrange":
            return f"{self.row0 + 1}:{self.row0_last + 1}"
        if k == "col":
            return index0_to_col(self.col0)
        if k == "colrange":
            return f"{index0_to_col(self.col0)}:{index0_to_col(self.col0_last)}"
        raise ValueError(f"unrenderable kind: {k}")

    def key(self) -> str:
        """Comparison key: two referents are the SAME referent iff keys match."""
        return " ".join(self.render().casefold().split())

    def __str__(self) -> str:  # pragma: no cover - convenience
        return self.render()


def _render_sheet_name(name: str) -> str:
    if any(ch in _NEEDS_QUOTE for ch in name) or name == "":
        return "'" + name.replace("'", "''") + "'"
    return name


# ---------------------------------------------------------------------------
# Parsing — the only place a 1-based number exists
# ---------------------------------------------------------------------------

_CELLRANGE = re.compile(r"([A-Za-z]{1,3})([1-9]\d*):([A-Za-z]{1,3})([1-9]\d*)")
_CELL = re.compile(r"([A-Za-z]{1,3})([1-9]\d*)")
_ROWRANGE = re.compile(r"([1-9]\d*):([1-9]\d*)")
_ROW = re.compile(r"[1-9]\d*")
_COLRANGE = re.compile(r"([A-Za-z]{1,3}):([A-Za-z]{1,3})")
_COL = re.compile(r"[A-Za-z]{1,3}")
_NAMED = re.compile(r"@(.+)", re.DOTALL)


def parse(text: str) -> Referent:
    raw = text.strip()
    if raw == "workbook:":
        return Referent(kind="workbook")
    if raw.startswith("sheetset:"):
        name = raw[len("sheetset:"):].strip()
        if not name:
            raise ReferentSyntaxError("empty sheetset name")
        return Referent(kind="sheetset", name=name)
    if not raw.startswith("sheet:"):
        raise ReferentSyntaxError(f"unknown referent prefix: {text!r}")

    sheet, span = _split_sheet_and_span(raw[len("sheet:"):])
    if span is None:
        return Referent(kind="sheet", sheet=sheet)
    return _parse_span(sheet, span)


def _split_sheet_and_span(rest: str) -> tuple[str, Optional[str]]:
    if rest.startswith("'"):
        buf: list[str] = []
        i = 1
        closed = False
        while i < len(rest):
            ch = rest[i]
            if ch == "'":
                if i + 1 < len(rest) and rest[i + 1] == "'":
                    buf.append("'")
                    i += 2
                    continue
                closed = True
                i += 1
                break
            buf.append(ch)
            i += 1
        if not closed:
            raise ReferentSyntaxError("unterminated quoted sheet name")
        name = "".join(buf)
        remainder = rest[i:]
        if remainder == "":
            span = None
        elif remainder.startswith("!"):
            span = remainder[1:]
        else:
            raise ReferentSyntaxError(f"expected '!' after quoted sheet name: {rest!r}")
    else:
        if "!" in rest:
            name, span = rest.split("!", 1)
        else:
            name, span = rest, None
        if any(ch in _NEEDS_QUOTE for ch in name):
            raise ReferentSyntaxError(
                f"sheet name {name!r} must be quoted: contains one of space ! : '"
            )
    if not name:
        raise ReferentSyntaxError("empty sheet name")
    if span is not None and span == "":
        raise ReferentSyntaxError("empty span after '!'")
    return name, span


def _parse_span(sheet: str, span: str) -> Referent:
    """A1 text (1-based) -> Referent (0-based). The conversion lives here."""
    m = _NAMED.fullmatch(span)
    if m:
        header = m.group(1).strip()
        if not header:
            raise ReferentSyntaxError("empty header name after '@'")
        return Referent(kind="namedcol", sheet=sheet, name=header)

    m = _CELLRANGE.fullmatch(span)
    if m:
        c_a, r_a = col_to_index0(m.group(1)), int(m.group(2)) - 1
        c_b, r_b = col_to_index0(m.group(3)), int(m.group(4)) - 1
        return Referent(kind="cellrange", sheet=sheet,
                        row0=min(r_a, r_b), row0_last=max(r_a, r_b),
                        col0=min(c_a, c_b), col0_last=max(c_a, c_b))

    m = _CELL.fullmatch(span)
    if m:
        return Referent(kind="cell", sheet=sheet,
                        row0=int(m.group(2)) - 1, col0=col_to_index0(m.group(1)))

    m = _ROWRANGE.fullmatch(span)
    if m:
        a, b = int(m.group(1)) - 1, int(m.group(2)) - 1
        return Referent(kind="rowrange", sheet=sheet,
                        row0=min(a, b), row0_last=max(a, b))

    if _ROW.fullmatch(span):
        return Referent(kind="row", sheet=sheet, row0=int(span) - 1)

    m = _COLRANGE.fullmatch(span)
    if m:
        a, b = col_to_index0(m.group(1)), col_to_index0(m.group(2))
        return Referent(kind="colrange", sheet=sheet,
                        col0=min(a, b), col0_last=max(a, b))

    if _COL.fullmatch(span):
        return Referent(kind="col", sheet=sheet, col0=col_to_index0(span))

    raise ReferentSyntaxError(f"unparseable span: {span!r}")


def from_legacy_pointer(sheet: str, pointer: Mapping[str, object]) -> Referent:
    """Convert a Data-agents-demo `manual_recipe` source_pointer to a Referent.

    Legacy pointers are 0-based and so is this grammar, so there is NO
    arithmetic here -- the indices carry across unchanged. That is the point of
    choosing 0-based internals: the adapter cannot introduce an off-by-one.
    """
    if "column" in pointer:
        return Referent(kind="namedcol", sheet=sheet, name=str(pointer["column"]))
    if "row" in pointer and "col" in pointer:
        return Referent(kind="cell", sheet=sheet,
                        row0=int(pointer["row"]), col0=int(pointer["col"]))
    if "column_index" in pointer:
        return Referent(kind="col", sheet=sheet, col0=int(pointer["column_index"]))
    raise ReferentSyntaxError(f"unrecognised legacy pointer: {dict(pointer)!r}")


# ---------------------------------------------------------------------------
# Resolution — 0-based inclusive bounds, plus half-open slices for pandas
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class Resolution:
    ok: bool
    referent: Optional[Referent] = None
    sheet: Optional[str] = None                 # the workbook's ACTUAL spelling
    sheets: Optional[tuple[str, ...]] = None    # for workbook:/sheetset:
    row0: Optional[int] = None
    row0_last: Optional[int] = None             # INCLUSIVE
    col0: Optional[int] = None
    col0_last: Optional[int] = None             # INCLUSIVE
    reason: Optional[str] = None
    detail: Optional[str] = None

    def row_slice(self) -> tuple[int, int]:
        """Half-open (start, stop) for df.iloc — no +1 at the call site."""
        return self.row0, self.row0_last + 1

    def col_slice(self) -> tuple[int, int]:
        return self.col0, self.col0_last + 1


class WorkbookView:
    """Read-only view: sheet names, used-range COUNTS, header values.

    `dims()` returns counts (n_rows, n_cols), not max indices, so a bounds check
    is `row0 < n_rows` with no off-by-one to get wrong. openpyxl is 1-based; the
    conversion is confined to this class.
    """

    def __init__(self, path: str | Path):
        from openpyxl import load_workbook  # parsing itself needs no deps

        self.path = Path(path)
        self._wb = load_workbook(self.path, data_only=True)
        self.sheet_names: tuple[str, ...] = tuple(self._wb.sheetnames)
        self._by_key = {n.casefold(): n for n in self.sheet_names}

    def actual_sheet(self, name: str) -> Optional[str]:
        return self._by_key.get(name.casefold())

    def dims(self, actual_sheet: str) -> tuple[int, int]:
        ws = self._wb[actual_sheet]
        return int(ws.max_row or 0), int(ws.max_column or 0)

    def row_values(self, actual_sheet: str, row0: int) -> list[str]:
        """Cell text, PRESERVED. Normalisation is the caller's declaration.

        This used to `.strip()`, which meant every construct reading a cell
        inherited a trim nobody asked for — PRO-2 instance 9. Constructs now
        declare what they want via `executor_contract.normalize_for`.
        """
        ws = self._wb[actual_sheet]
        one_based = row0 + 1
        values = next(ws.iter_rows(min_row=one_based, max_row=one_based,
                                   values_only=True), ())
        return ["" if v is None else str(v) for v in values]


def resolve(
    referent: Referent | str,
    wb: WorkbookView,
    header_rows0: Optional[Mapping[str, int]] = None,
    sheetsets: Optional[Mapping[str, Sequence[str]]] = None,
) -> Resolution:
    """Resolve a referent against a workbook. Failure is a hard error.

    `header_rows0` maps sheet name -> 0-BASED header row index. A header on the
    row Excel labels 4 is `3` here.
    """
    if isinstance(referent, str):
        try:
            referent = parse(referent)
        except ReferentSyntaxError as exc:
            return Resolution(ok=False, reason="malformed", detail=str(exc))

    if referent.kind == "workbook":
        return Resolution(ok=True, referent=referent, sheets=wb.sheet_names)

    if referent.kind == "sheetset":
        declared = (sheetsets or {}).get(referent.name or "")
        if declared is None:
            return Resolution(ok=False, referent=referent,
                              reason="sheetset_not_declared", detail=referent.name)
        members: list[str] = []
        for member in declared:
            actual = wb.actual_sheet(member)
            if actual is None:
                return Resolution(ok=False, referent=referent,
                                  reason="sheet_not_found", detail=member)
            members.append(actual)
        return Resolution(ok=True, referent=referent, sheets=tuple(members))

    actual = wb.actual_sheet(referent.sheet or "")
    if actual is None:
        return Resolution(ok=False, referent=referent, reason="sheet_not_found",
                          detail=referent.sheet)
    n_rows, n_cols = wb.dims(actual)

    if referent.kind == "sheet":
        return Resolution(ok=True, referent=referent, sheet=actual,
                          row0=0, row0_last=n_rows - 1,
                          col0=0, col0_last=n_cols - 1)

    if referent.kind == "namedcol":
        header_row0 = _header_row0_for(actual, header_rows0)
        if header_row0 is None:
            return Resolution(ok=False, referent=referent, sheet=actual,
                              reason="header_row_not_declared", detail=actual)
        if header_row0 >= n_rows:
            return Resolution(ok=False, referent=referent, sheet=actual,
                              reason="out_of_bounds",
                              detail=f"header row0 {header_row0} >= {n_rows} rows")
        from executor_contract import normalize_for  # no project deps upward

        wanted = normalize_for("header_label", referent.name or "")
        hits = [i for i, v in enumerate(wb.row_values(actual, header_row0))
                if normalize_for("header_label", v) == wanted]
        if not hits:
            return Resolution(ok=False, referent=referent, sheet=actual,
                              reason="header_not_found", detail=referent.name)
        if len(hits) > 1:
            # Refuse rather than take the first: picking one asserts what the
            # evidence does not establish.
            return Resolution(ok=False, referent=referent, sheet=actual,
                              reason="header_ambiguous",
                              detail=f"{referent.name} at col0 {hits}")
        return Resolution(ok=True, referent=referent, sheet=actual,
                          row0=0, row0_last=n_rows - 1,
                          col0=hits[0], col0_last=hits[0])

    row0 = referent.row0
    row0_last = referent.row0_last if referent.row0_last is not None else referent.row0
    col0 = referent.col0
    col0_last = referent.col0_last if referent.col0_last is not None else referent.col0
    if referent.kind in ("row", "rowrange"):
        col0, col0_last = 0, n_cols - 1
    if referent.kind in ("col", "colrange"):
        row0, row0_last = 0, n_rows - 1

    if (row0_last is not None and row0_last >= n_rows) or \
       (col0_last is not None and col0_last >= n_cols):
        return Resolution(ok=False, referent=referent, sheet=actual,
                          reason="out_of_bounds",
                          detail=f"used range is {n_rows} rows x {n_cols} cols")
    return Resolution(ok=True, referent=referent, sheet=actual,
                      row0=row0, row0_last=row0_last,
                      col0=col0, col0_last=col0_last)


def _header_row0_for(actual_sheet: str,
                     header_rows0: Optional[Mapping[str, int]]) -> Optional[int]:
    if not header_rows0:
        return None
    for name, row0 in header_rows0.items():
        if name.casefold() == actual_sheet.casefold():
            return int(row0)
    return None


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------

def _self_test() -> int:
    failures: list[str] = []
    seen_reasons: set[str] = set()

    def check(cond: bool, msg: str) -> None:
        if not cond:
            failures.append(msg)

    # --- round-trip: parse -> render is the identity on canonical forms -------
    canonical = [
        "workbook:", "sheetset:Months", "sheet:Sales", "sheet:Sales!D5",
        "sheet:Sales!D5:P96", "sheet:Sales!5", "sheet:Sales!5:96",
        "sheet:Sales!D", "sheet:Sales!D:P", "sheet:Sales!@Myynti",
        "sheet:'Myynti 2026'!A1",
    ]
    for text in canonical:
        try:
            r = parse(text)
        except ReferentSyntaxError as exc:
            failures.append(f"canonical {text!r} failed to parse: {exc}")
            continue
        check(r.render() == text, f"round-trip {text!r} -> {r.render()!r}")
        check(r.kind in KINDS, f"unknown kind for {text!r}: {r.kind}")

    # --- THE BOUNDARY: A1 text is 1-based, the object is 0-based -------------
    a1 = parse("sheet:Sales!D5")
    check((a1.row0, a1.col0) == (4, 3),
          f"A1 'D5' must be row0=4 col0=3, got ({a1.row0}, {a1.col0})")
    check(parse("sheet:Sales!A1").row0 == 0 and parse("sheet:Sales!A1").col0 == 0,
          "A1 'A1' must be row0=0 col0=0")
    check(parse("sheet:Sales!5").row0 == 4, "A1 row 5 must be row0=4")
    check(parse("sheet:Sales!D").col0 == 3, "A1 col D must be col0=3")
    rng = parse("sheet:Sales!B2:D5")
    check((rng.row0, rng.row0_last, rng.col0, rng.col0_last) == (1, 4, 1, 3),
          f"B2:D5 must be rows 1..4 cols 1..3 (0-based), got {rng}")
    # Rendering converts back, so the string a human sees never shifts.
    check(Referent(kind="cell", sheet="S", row0=0, col0=0).render() == "sheet:S!A1",
          "row0=0/col0=0 must render as A1")

    # --- comparison key ------------------------------------------------------
    check(parse("sheet:sales!d5").key() == parse("sheet:Sales!D5").key(),
          "key should be case-insensitive")
    check(parse("sheet:Sales!D5").key() != parse("sheet:Sales!D6").key(),
          "different cells must not share a key")
    check(parse("  sheet:Sales!D5  ").key() == parse("sheet:Sales!D5").key(),
          "surrounding whitespace must not matter")

    # --- quoting -------------------------------------------------------------
    check(parse("sheet:'Myynti 2026'!A1").sheet == "Myynti 2026", "quoted name lost")
    try:
        parse("sheet:Myynti 2026!A1")
        failures.append("a bare sheet name containing a space must be rejected")
    except ReferentSyntaxError:
        pass
    apostrophe = Referent(kind="sheet", sheet="O'Brien")
    check(apostrophe.render() == "sheet:'O''Brien'", f"apostrophe render: {apostrophe.render()}")
    check(parse(apostrophe.render()).sheet == "O'Brien", "apostrophe round-trip")

    # --- range ordering normalises -------------------------------------------
    check(parse("sheet:Sales!D5:B2").render() == "sheet:Sales!B2:D5", "cellrange must sort")
    check(parse("sheet:Sales!96:5").render() == "sheet:Sales!5:96", "rowrange must sort")
    check(parse("sheet:Sales!P:D").render() == "sheet:Sales!D:P", "colrange must sort")

    # --- A1 has no row 0 / no column 0 ---------------------------------------
    for bad in ("sheet:Sales!A0", "sheet:Sales!0", "sheet:Sales!0:5",
                "sheet:Sales!@", "sheet:!A1"):
        try:
            parse(bad)
            failures.append(f"{bad!r} must be rejected (A1 has no row 0)")
        except ReferentSyntaxError:
            pass

    # --- column letter maths, 0-based ----------------------------------------
    for letters, idx0 in (("A", 0), ("Z", 25), ("AA", 26), ("XFD", 16383)):
        check(col_to_index0(letters) == idx0, f"col_to_index0({letters}) != {idx0}")
        check(index0_to_col(idx0) == letters, f"index0_to_col({idx0}) != {letters}")

    # --- legacy pointers are ALREADY 0-based: no arithmetic ------------------
    legacy = from_legacy_pointer("Sales", {"row": 0, "col": 0})
    check((legacy.row0, legacy.col0) == (0, 0), "legacy (0,0) must stay (0,0)")
    check(legacy.render() == "sheet:Sales!A1", "legacy (0,0) renders as A1")
    legacy2 = from_legacy_pointer("Sales", {"row": 1, "col": 1})
    check((legacy2.row0, legacy2.col0) == (1, 1), "legacy (1,1) must stay (1,1)")
    check(legacy2.render() == "sheet:Sales!B2", "legacy (1,1) renders as B2")
    check(from_legacy_pointer("Sales", {"column": "Myynti"}).render() == "sheet:Sales!@Myynti",
          "legacy named column")

    # --- resolution against the real multi-sheet workbook --------------------
    fixture = ROOT / "fixtures" / "W1_multisheet.xlsx"
    if not fixture.exists():
        failures.append(f"fixture missing: {fixture} (run make_w1.py)")
    else:
        wb = WorkbookView(fixture)
        # Sales headers sit on the row Excel labels 4 -> row0 = 3.
        headers0 = {"Sales": 3, "Dup": 0, "Myynti 2026": 0}

        def res(text: str, **kw) -> Resolution:
            r = resolve(text, wb, header_rows0=kw.pop("headers0", headers0), **kw)
            if r.reason:
                seen_reasons.add(r.reason)
            return r

        ok = res("sheet:Sales!A4")
        check(ok.ok and ok.sheet == "Sales" and (ok.row0, ok.col0) == (3, 0),
              f"Sales!A4 -> row0 3, col0 0; got {ok}")
        check(ok.row_slice() == (3, 4) and ok.col_slice() == (0, 1),
              f"half-open slices for iloc: {ok.row_slice()}, {ok.col_slice()}")

        lower = res("sheet:sales!A1")
        check(lower.ok and lower.sheet == "Sales",
              "sheet lookup must be case-insensitive and report actual spelling")

        named = res("sheet:Sales!@Tuote")
        check(named.ok and named.col0 == 0, f"@Tuote should be col0 0: {named}")
        named2 = res("sheet:Sales!@Yhteensä")
        check(named2.ok and named2.col0 == 4, f"@Yhteensä should be col0 4: {named2}")

        nodecl = res("sheet:Sales!@Tuote", headers0=None)
        check(nodecl.reason == "header_row_not_declared",
              f"expected header_row_not_declared: {nodecl}")

        ambig = res("sheet:Dup!@Myynti")
        check(ambig.reason == "header_ambiguous",
              f"duplicate headers must REFUSE, not pick the first: {ambig}")
        check("[1, 2]" in (ambig.detail or ""),
              f"ambiguity should name the 0-based columns: {ambig.detail}")

        check(res("sheet:Sales!@Nope").reason == "header_not_found", "missing header")
        check(res("sheet:Nope!A1").reason == "sheet_not_found", "missing sheet")
        check(res("sheet:Sales!G1").reason == "out_of_bounds", "col0 6 >= 6 cols")
        check(res("sheet:Sales!A10").reason == "out_of_bounds", "row0 9 >= 9 rows")
        check(res("sheet:Sales!F9").ok, "F9 is the last cell of the used range")
        check(res("sheet:Sales!(").reason == "malformed", "unparseable span")

        spaced = res("sheet:'Myynti 2026'!B2")
        check(spaced.ok and spaced.sheet == "Myynti 2026", f"spaced sheet: {spaced}")

        check(res("sheetset:Months").reason == "sheetset_not_declared", "undeclared sheetset")
        declared = resolve("sheetset:Months", wb,
                           sheetsets={"Months": ["2026-01", "2026-02"]})
        check(declared.ok and declared.sheets == ("2026-01", "2026-02"),
              f"declared sheetset: {declared}")

        whole = resolve("workbook:", wb)
        check(whole.ok and len(whole.sheets or ()) == 6, f"workbook: -> {whole.sheets}")

        sheet_only = res("sheet:Sales")
        check(sheet_only.ok and (sheet_only.row0_last, sheet_only.col0_last) == (8, 5),
              f"sheet: spans the used range as 0-based inclusive: {sheet_only}")

        # Every declared failure reason must actually be exercised.
        untested = set(REASONS) - seen_reasons
        check(not untested, f"declared but untested failure reasons: {sorted(untested)}")

    if failures:
        sys.stderr.write("SELF-TEST FAILED:\n  " + "\n  ".join(failures) + "\n")
        return 1
    sys.stdout.write(
        "SELF-TEST PASSED (round-trips / A1<->0-based boundary / keys / quoting + "
        "apostrophes / range ordering / no row 0 in A1 / column maths / legacy "
        "pointers carry across with no arithmetic / resolution + iloc slices on a "
        "real 6-sheet workbook / all 7 failure reasons exercised)\n"
    )
    return 0


if __name__ == "__main__":
    argv = sys.argv[1:]
    if argv and argv[0] == "--self-test":
        raise SystemExit(_self_test())
    if argv and argv[0] == "--resolve":
        view = WorkbookView(argv[1])
        out = resolve(argv[2], view)
        print(out)
        raise SystemExit(0 if out.ok else 1)
    sys.stderr.write("usage: referents.py --self-test | --resolve <workbook.xlsx> <referent>\n")
    raise SystemExit(2)
