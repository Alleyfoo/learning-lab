#!/usr/bin/env python3
"""Cross-sheet law 3 — Ignored Means Ignored.

## The law

> Varying a sheet declared `role: ignore` must not change authoritative output,
> whatever that sheet contains. Ignored must mean ignored **even when the ignored
> material is maximally relevant** — same headers, same labels, same data, same
> layout.

An invariance law rather than a conservation law, so its shape is the mirror of
laws 1 and 2:

```text
baseline:   workbook where ignored sheet I holds X   -> O
mutation:   same workbook, I holds Y                 -> O'
required:   O == O'
never:      O != O'
```

## Why "nastily" is the whole point

An ignored sheet full of unrelated junk tests nothing: no plausible defect would
read it. The defects this law guards against are all *relevance* defects — a
referent resolving against the wrong sheet because another sheet answers to the
same `@Name`, a label rule matching a label that lives elsewhere, a union
sweeping in a sheet that looks like a member. So every case here makes the
ignored sheet as confusable with the data sheet as the format allows, and only
the DECLARED ROLE distinguishes them.

## The control runs the other direction, and matters as much

`must_differ`: varying the DATA sheet MUST change the output. Without it a
harness that returned a constant would pass every invariance case, and the law
would be measuring its own plumbing. Law 1 needed exactly this control for the
same reason.

## Reachability

Two workbooks that do not actually differ make every invariance case pass
vacuously. `assert_mutation_reached()` proves the ignored sheet's content really
changed between baseline and mutation before any verdict is believed. This is the
third law in a row where a fixture bug would have produced a clean green.

## Two canaries, because there are two leak paths

```text
role_counterfactual   declare the SAME sheet as `data`; the output must move.
                      Not a synthetic defect -- it establishes that the
                      observation channel can see the sheet at all, so the
                      invariance is attributable to the DECLARED ROLE rather
                      than to the harness being blind to it.
resolution_leak       `@Name` resolves against the first sheet in workbook
                      order carrying that header -- authority by accident,
                      the defect axis 5 is named after.
```

**Run 1 registered only the resolution canary and it did not fire, so the run was
VOID despite 6/6 cases passing.** The diagnosis is the reason the corpus has the
shape it does: with the ignored sheet's headers in the SAME column positions as
the data sheet's, a referent mis-resolved to the wrong sheet is unobservable,
because only the column INDEX crosses into the read — never the sheet. Hence
`same_headers_reordered`, which varies the positions and is the only case either
canary can be seen through.

The lesson generalises past this law: a corpus can pass completely while blind to
an entire class of defect, and the canary is what tells the two apart. Both must
fire AND be reached, or the run is VOID.

Usage
-----
    python definition_phase/harness/ignored_independence.py            # run + record
    python definition_phase/harness/ignored_independence.py --no-record
"""
from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path
from typing import Any, Callable, Optional

HERE = Path(__file__).resolve().parent
LAB = HERE.parent.parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(LAB / "experimentL" / "harness"))

import execute_recipe  # noqa: E402
from execute_recipe import execute  # noqa: E402
from recipe import recipe_from_json  # noqa: E402
from referents import WorkbookView  # noqa: E402
from validate_recipe import validate  # noqa: E402

RESULTS = LAB / "definition_phase" / "results"

# The data sheet is IDENTICAL in every case and every variant. Only the ignored
# sheet moves, so any output difference is attributable to it and nothing else.
DATA_SHEET = [["Tuote", "Myynti"], ["A-1", 10], ["SUMMA", 99], ["A-2", 20]]

# The ignored sheet is named to sort FIRST in workbook order, so a resolver that
# takes "the first sheet answering to this header" reaches it rather than the
# data sheet. A defect that only manifests on ordering would otherwise be
# invisible here -- and ordering is axis 4, deliberately not relied on.
IGNORED = "AAA_ignored"
DATA = "Sales"


def _wb(tmp: Path, tag: str, ignored_rows: list[list],
        data_rows: Optional[list[list]] = None,
        hide: bool = False) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = tmp / f"{tag}.xlsx"
    book = Workbook()
    ws = book.active
    ws.title = IGNORED
    for row in ignored_rows:
        ws.append(row)
    if hide:
        # White-on-white text and a hidden row: visible to the machine, invisible
        # on screen. The white-text asymmetry, pointed at an IGNORED sheet.
        for cell in ws[1]:
            cell.font = Font(color="FFFFFFFF")
            cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF",
                                    fill_type="solid")
        ws.row_dimensions[2].hidden = True

    ws2 = book.create_sheet()
    ws2.title = DATA
    for row in (data_rows if data_rows is not None else DATA_SHEET):
        ws2.append(row)
    book.save(path)
    return path


def _recipe(ignored_role: str = "ignore"):
    """The recipe under test.

    `ignored_role` exists for the counterfactual canary: declaring the same sheet
    as `data` proves the harness CAN see its content, so the invariance observed
    under `ignore` is attributable to the declared role rather than to the
    observation channel being blind.
    """
    if ignored_role == "data":
        ignored_entry = {
            "sheet": f"sheet:{IGNORED}", "role": "data",
            "header_row": f"sheet:{IGNORED}!1", "data_region": "remainder",
            "fields": [
                {"target": "tuote_i", "source": f"sheet:{IGNORED}!@Tuote",
                 "role": "id", "type": "string"},
                {"target": "myynti_i", "source": f"sheet:{IGNORED}!@Myynti",
                 "role": "measure", "type": "number"}],
            "exclude": [], "ambiguities": []}
    else:
        ignored_entry = {"sheet": f"sheet:{IGNORED}", "role": "ignore",
                         "fields": [], "exclude": [], "ambiguities": []}

    raw = {
        "recipe_version": 1, "recipe_id": "law3", "workbook": {},
        "sheets": [
            {"sheet": f"sheet:{DATA}", "role": "data",
             "header_row": f"sheet:{DATA}!1", "data_region": "remainder",
             "fields": [
                 {"target": "tuote", "source": f"sheet:{DATA}!@Tuote",
                  "role": "id", "type": "string"},
                 {"target": "myynti", "source": f"sheet:{DATA}!@Myynti",
                  "role": "measure", "type": "number"}],
             # A label rule, so one case can plant the SAME label on the ignored
             # sheet and ask whether the rule reaches across.
             "exclude": [{"rule": {"op": "label_in",
                                   "column": f"sheet:{DATA}!@Tuote",
                                   "values": ["SUMMA"]},
                          "reason": "aggregate row"}],
             "ambiguities": []},
            ignored_entry,
        ],
        "applicability": None,
        "provenance": {"proposed_by": "law3", "approved_by": "law3",
                       "approved_recipe_sha256": None},
    }
    r = recipe_from_json(raw)
    raw["provenance"]["approved_recipe_sha256"] = r.content_sha256()
    return recipe_from_json(raw)


def _outcome(path: Path, ignored_role: str = "ignore") -> dict:
    wb = WorkbookView(path)
    recipe = _recipe(ignored_role)
    report = validate(recipe, wb)
    if not report.valid:
        return {"refused": True, "codes": sorted(report.codes())}
    ex = execute(recipe, wb)
    return {"refused": False, "columns": list(ex.columns),
            "rows": [list(r) for r in ex.rows],
            "member_contribution": dict(ex.member_contribution)}


def _sheet_bytes(path: Path, sheet: str) -> str:
    """The machine-readable content of one sheet, for the reachability check."""
    from openpyxl import load_workbook

    ws = load_workbook(path)[sheet]
    return json.dumps([[("" if c.value is None else str(c.value)) for c in row]
                       for row in ws.iter_rows()], ensure_ascii=False)


def assert_mutation_reached(base: Path, mut: Path, sheet: str) -> tuple[bool, str]:
    """The two workbooks must genuinely differ on the sheet being varied.

    Otherwise every invariance case passes because nothing changed, which is a
    fixture bug wearing a clean result.
    """
    a, b = _sheet_bytes(base, sheet), _sheet_bytes(mut, sheet)
    if a == b:
        return False, f"{sheet} is IDENTICAL in both workbooks -- nothing was varied"
    return True, f"{sheet} differs between the two workbooks ({len(a)} vs {len(b)} chars)"


# ---------------------------------------------------------------------------
# the corpus -- each case varies the ignored sheet between two nasty extremes
# ---------------------------------------------------------------------------

H = ["Tuote", "Myynti"]

CASES: list[dict[str, Any]] = [
    {
        "case": "same_headers",
        "why": "the ignored sheet answers to the same @Name referents as the data "
               "sheet, and sits first in workbook order. A resolver that takes "
               "the first matching sheet reads it.",
        "baseline": {"ignored": [H, ["X-1", 1]]},
        "mutation": {"ignored": [H, ["X-9", 777]]},
        "required": "invariant",
    },
    {
        "case": "same_headers_reordered",
        "why": "the ignored sheet carries the same header NAMES in different "
               "COLUMN POSITIONS. Added after the first run of this law: with "
               "identical positions, a referent mis-resolved to the wrong SHEET "
               "is invisible, because only the column INDEX crosses into the "
               "read. Varying the order is what makes a resolution leak "
               "observable at all.",
        "baseline": {"ignored": [["Tuote", "Myynti"], ["X-1", 1]]},
        "mutation": {"ignored": [["Myynti", "Tuote"], [1, "X-1"]]},
        "required": "invariant",
    },
    {
        "case": "same_labels",
        "why": "the label the exclusion rule denotes (SUMMA) is planted on the "
               "ignored sheet. A rule that scans by label rather than by resolved "
               "column reaches across.",
        "baseline": {"ignored": [H, ["SUMMA", 1]]},
        "mutation": {"ignored": [H, ["SUMMA", 2], ["SUMMA", 3], ["SUMMA", 4]]},
        "required": "invariant",
    },
    {
        "case": "same_data_verbatim",
        "why": "the ignored sheet is a byte-identical copy of the data sheet. "
               "Identity by CONTENT rather than by declared role would union it, "
               "which is law 1's mistake arriving through a different door.",
        "baseline": {"ignored": [r[:] for r in DATA_SHEET]},
        "mutation": {"ignored": [r[:] for r in DATA_SHEET] + [["A-3", 30]]},
        "required": "invariant",
    },
    {
        "case": "larger_than_data_sheet",
        "why": "the ignored sheet has more rows and more columns than the data "
               "sheet. Any dimension taken from the workbook rather than from the "
               "declared sheet shows up here.",
        "baseline": {"ignored": [H + ["Extra"], ["X-1", 1, "p"]]},
        "mutation": {"ignored": [H + ["Extra", "More"]]
                                + [[f"X-{i}", i, "p", "q"] for i in range(1, 12)]},
        "required": "invariant",
    },
    {
        "case": "hidden_content",
        "why": "the ignored sheet carries white-on-white text and a hidden row -- "
               "visible to the machine, invisible on screen. The white-text "
               "asymmetry aimed at ignored material.",
        "baseline": {"ignored": [H, ["X-1", 1]], "hide": True},
        "mutation": {"ignored": [H, ["SUMMA", 999], ["A-1", 888]], "hide": True},
        "required": "invariant",
    },
    {
        "case": "CONTROL_vary_the_data_sheet",
        "why": "the other direction. Varying the DATA sheet MUST change the "
               "output, or this suite would pass by returning a constant and the "
               "invariance results would mean nothing.",
        "baseline": {"ignored": [H, ["X-1", 1]]},
        "mutation": {"ignored": [H, ["X-1", 1]],
                     "data": [["Tuote", "Myynti"], ["A-1", 10], ["SUMMA", 99],
                              ["A-2", 20], ["A-3", 30]]},
        "required": "must_differ",
    },
]


def _build(tmp: Path, tag: str, spec: dict) -> Path:
    return _wb(tmp, tag, spec["ignored"], spec.get("data"), spec.get("hide", False))


def _verdict(case: dict, base: dict, mut: dict,
             reached: bool, reach_detail: str) -> tuple[str, str]:
    if not reached:
        return "NON_EVIDENTIAL", reach_detail
    if base["refused"] or mut["refused"]:
        return "NON_EVIDENTIAL", (
            f"refused before execution (baseline={base.get('codes')}, "
            f"mutation={mut.get('codes')}) -- the case never reached the "
            f"observation point")

    same = base["rows"] == mut["rows"] and base["columns"] == mut["columns"]

    if case["required"] == "must_differ":
        if same:
            return "CONTROL_FAILED", (
                "varying the DATA sheet did not change the output; this suite is "
                "returning a constant and every invariance result above is void")
        return "HELD", f"data change moved the output: {len(base['rows'])} -> {len(mut['rows'])} rows"

    if same:
        return "HELD", f"output unchanged: {len(base['rows'])} rows, {base['rows']}"
    return "VIOLATED", (f"ignored material changed the output: "
                        f"{base['rows']} -> {mut['rows']}")


# ---------------------------------------------------------------------------
# canary -- authority by accident, the axis-5 defect used as a known violation
# ---------------------------------------------------------------------------

def _canary_role_counterfactual() -> dict:
    """Declare the SAME sheet as `data` and require the output to move.

    Not a synthetic defect — the counterfactual. It establishes that the
    observation channel can see this sheet's content at all, so the invariance
    observed under `role: ignore` is attributable to the DECLARED ROLE rather
    than to the harness being blind to that sheet. Without it, a law-3 pass is
    consistent with the executor never having been able to read the sheet under
    any circumstances, which would make every case above vacuous.
    """
    case = next(c for c in CASES if c["case"] == "same_headers")
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        base = _outcome(_build(tmp, "cf_base", case["baseline"]), ignored_role="data")
        mut = _outcome(_build(tmp, "cf_mut", case["mutation"]), ignored_role="data")

    if base["refused"] or mut["refused"]:
        return {"name": "role_counterfactual", "fired": False, "reached": False,
                "detail": f"never executed: {base.get('codes')} / {mut.get('codes')}"}
    moved = base["rows"] != mut["rows"]
    return {"name": "role_counterfactual", "fired": moved, "reached": True,
            "detail": (f"as data, varying the sheet moved the output: "
                       f"{base['rows']} -> {mut['rows']}" if moved else
                       "as DATA the sheet still did not move the output -- the "
                       "harness cannot see it under any role, so every "
                       "invariance case is vacuous")}


def _canary_resolution_leak() -> dict:
    """First-matching-sheet-wins resolution — authority by accident.

    Fires on `same_headers_reordered` specifically. It cannot fire on
    `same_headers`: with identical column positions a mis-resolved SHEET is
    unobservable, because only the column INDEX reaches the read. That is the
    finding from this law's first run and the reason the reordered case exists.
    """
    original = execute_recipe.resolve

    def leaky(text, wb, header_rows0=None, **kw):
        result = original(text, wb, header_rows0=header_rows0 or {}, **kw)
        if "!@" not in str(text):
            return result
        name_part = str(text).split("!@", 1)[1]
        label = name_part.strip().casefold()
        for name in wb.sheet_names:                    # workbook order
            row0 = (header_rows0 or {}).get(name, 0)
            values = [str(v).strip().casefold() for v in wb.row_values(name, row0)]
            if label in values:
                alt = original(f"sheet:{name}!@{name_part}", wb,
                               header_rows0={**(header_rows0 or {}), name: row0})
                return alt if alt.ok else result
        return result

    case = next(c for c in CASES if c["case"] == "same_headers_reordered")
    execute_recipe.resolve = leaky
    try:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            base = _outcome(_build(tmp, "leak_base", case["baseline"]))
            mut = _outcome(_build(tmp, "leak_mut", case["mutation"]))
    finally:
        execute_recipe.resolve = original

    if base["refused"] or mut["refused"]:
        return {"name": "resolution_leak", "fired": False, "reached": False,
                "detail": f"never executed: {base.get('codes')} / {mut.get('codes')}"}
    fired = base["rows"] != mut["rows"]
    return {"name": "resolution_leak", "fired": fired, "reached": True,
            "detail": (f"leaky resolver let the ignored sheet steer the read: "
                       f"{base['rows']} -> {mut['rows']}" if fired else
                       "leaky resolver changed nothing -- this law cannot detect "
                       "a resolution leak")}


def canaries() -> list[dict]:
    """Two leak paths, each with a case in the corpus that can see it.

    Both must fire and be reached. Registering only one would leave the law
    claiming a sensitivity it has not demonstrated -- and the first run of this
    law demonstrated exactly that failure, passing 6/6 while blind to the
    resolution path.
    """
    return [_canary_role_counterfactual(), _canary_resolution_leak()]


# ---------------------------------------------------------------------------
# run
# ---------------------------------------------------------------------------

def run_all() -> dict:
    results = []
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        for case in CASES:
            base_path = _build(tmp, f"{case['case']}_base", case["baseline"])
            mut_path = _build(tmp, f"{case['case']}_mut", case["mutation"])
            varied = DATA if case["required"] == "must_differ" else IGNORED
            reached, reach_detail = assert_mutation_reached(base_path, mut_path, varied)

            base, mut = _outcome(base_path), _outcome(mut_path)
            status, why = _verdict(case, base, mut, reached, reach_detail)
            results.append({
                "case": case["case"], "required": case["required"],
                "rationale": case["why"], "varied_sheet": varied,
                "reachability": reach_detail, "reachability_ok": reached,
                "baseline": base, "mutation": mut,
                "status": status, "detail": why,
            })

    canary_results = canaries()
    violations = [r for r in results if r["status"] == "VIOLATED"]
    control_failed = [r for r in results if r["status"] == "CONTROL_FAILED"]
    non_evidential = [r for r in results if r["status"] == "NON_EVIDENTIAL"]

    if not all(c["fired"] and c["reached"] for c in canary_results):
        outcome = "VOID"
    elif control_failed:
        outcome = "VOID"
    elif violations:
        outcome = "LAW_3_VIOLATED"
    elif non_evidential:
        outcome = "INCONCLUSIVE"
    else:
        outcome = "LAW_3_HELD"

    return {
        "law": "Ignored Means Ignored",
        "statement": ("varying a sheet declared role:ignore must not change "
                      "authoritative output, whatever it contains"),
        "canaries": canary_results,
        "cases": results,
        "outcome": outcome,
        "stated_limitation": (
            "six relevance shapes, chosen by their author, and TWO registered "
            "leak paths (role counterfactual, resolution). Other leak paths are "
            "undemonstrated. A sheet is 'ignored' "
            "here by DECLARED ROLE only; ignoring by any other mechanism is not "
            "covered. Sheet ORDERING is held fixed (ignored sheet first) rather "
            "than varied -- that is axis 4."),
    }


def main(argv: list[str]) -> int:
    result = run_all()

    for c in result["canaries"]:
        print(f"CANARY {c['name']:22} fired={str(c['fired']):5} "
              f"reached={str(c['reached']):5}  {c['detail']}")
    print()
    for r in result["cases"]:
        print(f"  {r['status']:15} {r['case']:28} {r['detail']}")
    print(f"\nOUTCOME: {result['outcome']}")

    if "--no-record" not in argv:
        RESULTS.mkdir(exist_ok=True)
        n = 1
        while (RESULTS / f"ignored_independence_run{n}.json").exists():
            n += 1
        path = RESULTS / f"ignored_independence_run{n}.json"
        path.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n",
                        encoding="utf-8")
        print(f"  written to {path.name}")

    return 0 if result["outcome"] == "LAW_3_HELD" else 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
