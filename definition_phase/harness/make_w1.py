#!/usr/bin/env python3
"""Generate `fixtures/W1_multisheet.xlsx` — the referent-grammar fixture.

The generator is the fixture's source of truth: regenerating produces the same
workbook, so the .xlsx can be rebuilt rather than trusted. Every sheet earns its
place (see design/referent_grammar_v1.md sec.8).

  Sales        merged title, timestamp, blank row, header on row 4, data rows,
               a total ROW and a total COLUMN -- the realistic messy case
  Myynti 2026  sheet name containing a space -> quoting
  Dup          two columns headed 'Myynti'   -> header_ambiguous
  Notes        free text, no table           -> a sheet whose role is 'ignore'
  2026-01/02   two monthly sheets            -> a sheetset is meaningful here
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "fixtures" / "W1_multisheet.xlsx"


def _rows(ws: Worksheet, rows: list[list[object]]) -> None:
    for row in rows:
        ws.append(row)


def build() -> Workbook:
    wb = Workbook()

    # --- Sales: the messy one -------------------------------------------------
    ws = wb.active
    ws.title = "Sales"
    _rows(ws, [
        ["Myyntiraportti 2026", None, None, None, None, None],
        ["Päivitetty", "3.2.2026", None, None, None, None],
        [],
        ["Tuote", "Tammi", "Helmi", "Maalis", "Yhteensä", "Kommentti"],
        ["ART-001", 10, 12, 8, 30, None],
        ["ART-002", 7, 9, 11, 27, "korjattu"],
        ["ART-003", 5, 7, 6, 18, None],
        ["ART-004", 8, 11, 7, 26, None],
        ["YHTEENSÄ", 30, 39, 32, 101, None],
    ])
    # A merged title band: a fact for the EVIDENCE layer, not the grammar.
    ws.merge_cells("A1:F1")

    # --- 'Myynti 2026': a sheet name with a space -----------------------------
    ws = wb.create_sheet("Myynti 2026")
    _rows(ws, [
        ["Tuote", "Myynti", "Kate"],
        ["ART-001", 100, 30],
        ["ART-002", 80, 25],
    ])

    # --- Dup: duplicate header -> header_ambiguous ----------------------------
    ws = wb.create_sheet("Dup")
    _rows(ws, [
        ["Tuote", "Myynti", "Myynti"],
        ["ART-001", 1, 2],
        ["ART-002", 3, 4],
    ])

    # --- Notes: no table ------------------------------------------------------
    ws = wb.create_sheet("Notes")
    _rows(ws, [
        ["Vapaata tekstiä. Ei taulukkoa."],
        ["Toimittaja vaihtoi järjestelmää 1.1.2026."],
    ])

    # --- Monthly sheets: a sheetset is meaningful on this workbook -------------
    for name, base in (("2026-01", 10), ("2026-02", 20)):
        ws = wb.create_sheet(name)
        _rows(ws, [
            ["Tuote", "Myynti"],
            ["ART-001", base],
            ["ART-002", base + 5],
        ])

    return wb


def main(force: bool = False) -> None:
    """Write the fixture.

    `.xlsx` output is NOT byte-reproducible -- openpyxl embeds zip timestamps --
    and this workbook's hash is frozen (`frozen_manifest.json`; Experiment K's
    C1/C2 are copies of it). Regenerating it therefore VOIDS those freezes, so
    overwriting requires --force and is a deliberate re-freeze.
    """
    OUT.parent.mkdir(parents=True, exist_ok=True)
    if OUT.exists() and not force:
        raise SystemExit(
            f"REFUSING to overwrite {OUT.name}: it is a frozen artifact and .xlsx\n"
            f"output is not byte-reproducible, so rewriting it voids every hash\n"
            f"that references it. Re-run with --force only as a deliberate\n"
            f"re-freeze, then update frozen_manifest.json in the same commit and\n"
            f"run: python scripts/verify_frozen.py")
    build().save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main(force="--force" in sys.argv[1:])
