#!/usr/bin/env python3
"""Approval provenance — prove what evidence a human actually had.

OPEN-2: approval currently binds a hash of the *recipe*, not of the region the
reviewer looked at, nor of what they were shown of it. White text, hidden rows
and `;;;` formats are visible to the machine and invisible on screen, so
"a human approved it" is only a control if the human was shown what the machine
read. This module records enough state to prove that afterwards.

Deliberately boring. No storage layer, no signing, no chain — a record, four
hashes, a version, and a verifier.

```text
raw source region ──hash──> source_region_hash
        │
        ▼
    detector ────findings──> detector_findings_hash
        │
        ▼
 review renderer ──output──> review_view_hash        + renderer_version
        │
        ▼
      human
        │
        ▼
 approved recipe ────────── recipe_hash
```

Two rules that keep the mechanism from recreating the hole it exists to close:

1. **`source_region_hash` hashes the MACHINE-READABLE region**, never a visual
   rendering. Hashing what Excel draws would rebuild OPEN-2 inside the provenance
   mechanism — the reviewer's blind spot would become the record's blind spot.
2. **Verification fails per relationship, not as one boolean.** "Approval
   invalid" is useless; you immediately want to know *which* link broke.

And two verdicts rather than one, because they answer different questions:

    historically_valid            every link still recomputes under the record's
                                  OWN renderer -- what actually happened
    meets_current_review_policy   ...and that renderer is the current one

A renderer upgrade must NOT invalidate history, and must NOT silently confer the
new renderer's protections on an old approval. Both would be lies, in opposite
directions.
"""
from __future__ import annotations

import hashlib
import json
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Callable, Optional

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from hidden_content import scan_sheet  # noqa: E402
from recipe import Recipe  # noqa: E402
from structure_view import render_structure, render_values, sheet_rows  # noqa: E402

CURRENT_RENDERER_VERSION = "review-v3"

# Two questions that must NOT be allowed to collapse into one:
#
#     Execution.degraded    did execution fail to deliver something the recipe
#                           DECLARED?
#     the review renderer   was that degradation actually AVAILABLE to the human
#                           granting authority?
#
# Conflating them is OPEN-2 in a new costume: "the execution object contained it,
# therefore the approval was informed". It was not, unless the renderer that
# produced the reviewed view put it on screen -- which is what `review_view_hash`
# plus `renderer_version` records, and what `meets_current_review_policy` refuses
# to grant retroactively.


# ---------------------------------------------------------------------------
# renderers — versioned, and kept available so history stays verifiable
# ---------------------------------------------------------------------------

def _render_v1(sheets: dict, findings: list[dict],
               recipe=None, wb_path=None) -> str:
    """review-v1: the naive rendering — values only.

    Kept because approvals were made under it, not because it is good. It shows
    a reviewer what Excel would show them, which is precisely the blind spot:
    hidden content is present in the values and absent from any warning.
    """
    out = []
    for name, data in sheets.items():
        out.append(f"# sheet {name}")
        out.append(render_values(data["rows"], data["n_cols"]))
    return "\n".join(out)


def _render_v2(sheets: dict, findings: list[dict],
               recipe=None, wb_path=None) -> str:
    """review-v2: values, plus the structure view, plus hidden-content findings.

    The difference that matters is the findings block: under v1 a reviewer could
    approve a workbook carrying invisible text with nothing on screen to say so.
    """
    out = []
    for name, data in sheets.items():
        out.append(f"# sheet {name}")
        out.append(render_values(data["rows"], data["n_cols"]))
        out.append(render_structure(data["rows"], data["n_cols"]))
    out.append("# hidden content")
    if findings:
        for f in findings:
            out.append(f"  {f['kind']} {f['referent']} {f['detail']}")
    else:
        out.append("  none detected")
    return "\n".join(out)


def _render_v3(sheets: dict, findings: list[dict], recipe=None, wb_path=None) -> str:
    """review-v3: v2, plus what the recipe DECLARED and what this workbook DELIVERS.

    Observable Error v1 crossing the human boundary. The rule is the same one the
    result surface enforces:

        if the system promises X and delivers something weaker than X, that
        weakening must travel with the result

    and a human granting authority is a consumer of that result like any other.
    Under v2 a reviewer could approve a recipe declaring `amount: number` against
    a workbook that delivers strings, with nothing on screen to say so.

    The observed column comes from the EXECUTOR, not from a reimplementation of
    its coercion rules. Two implementations of "would this be honoured" is the
    PRO-2 shape this programme keeps finding: they agree until they do not, and
    the reviewer sees the one that is wrong.
    """
    out = [_render_v2(sheets, findings)]

    declared: list[tuple[str, str]] = []
    if recipe is not None:
        for entry in recipe.data_sheets():
            for fld in entry.fields:
                declared.append((fld.target, fld.type or "(untyped)"))

    out.append("# declared types")
    if declared:
        for target, dtype in declared:
            out.append(f"    {target}: {dtype}")
    else:
        out.append("    none declared")

    out.append("# observed on this workbook")
    if recipe is None or wb_path is None:
        # Not reachable through review_view(), which always passes both. Stated
        # rather than assumed: a renderer that silently omitted the section
        # would show the reviewer a clean-looking view of a degraded recipe.
        out.append("    UNAVAILABLE -- renderer was given no recipe or workbook")
        out.append("RESULT DEGRADATION UNKNOWN")
        return "\n".join(out)

    unhonoured, note = _observed_degradation(recipe, wb_path)
    if note:
        out.append(f"    UNAVAILABLE -- {note}")
        out.append("RESULT DEGRADATION UNKNOWN")
        return "\n".join(out)

    for target, dtype in declared:
        reason = unhonoured.get(target)
        if reason:
            out.append(f"    {target}: unhonoured -- {reason}")
        else:
            out.append(f"    {target}: honoured")
    out.append("RESULT DEGRADED" if unhonoured else "RESULT NOT DEGRADED")
    return "\n".join(out)


def _observed_degradation(recipe, wb_path) -> tuple[dict, Optional[str]]:
    """Per-target degradation as the EXECUTOR reports it.

    Returns ({target: short reason}, None) or ({}, why-it-could-not-be-determined).
    A recipe that cannot execute is not "not degraded" -- the reviewer is told
    the question could not be answered, which is different from being told no.
    """
    import sys as _sys

    _sys.path.insert(0, str(HERE.parent.parent / "experimentL" / "harness"))
    from execute_recipe import InsufficientRecipe, execute  # noqa: E402

    from referents import WorkbookView  # noqa: E402

    wb = WorkbookView(wb_path)
    try:
        ex = execute(recipe, wb)
    except InsufficientRecipe as exc:
        return {}, f"the recipe cannot execute on this workbook: {exc}"

    out: dict[str, str] = {}
    for item in ex.degradation:
        target = item.get("target", "?")
        gap = item.get("gap")
        declared_type = item.get("declared", "?")
        out[target] = (f"delivered as string (gap {gap})" if gap
                       else f"declared {declared_type}, {item.get('reason', '')}")
    return out, None


RENDERERS: dict[str, Callable[..., str]] = {
    "review-v1": _render_v1,
    "review-v2": _render_v2,
    "review-v3": _render_v3,
}


# ---------------------------------------------------------------------------
# the record
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class ApprovalRecord:
    recipe_id: str
    recipe_hash: str
    source_region_hash: str
    review_view_hash: str
    detector_findings_hash: str
    renderer_version: str
    approved_by: str
    approved_at: str

    def to_json(self) -> str:
        return json.dumps(asdict(self), ensure_ascii=False, indent=2, sort_keys=True)

    @staticmethod
    def from_json(text: str) -> "ApprovalRecord":
        return ApprovalRecord(**json.loads(text))


CHECKS = ("recipe", "source_region", "detector_findings", "review_view")
REASONS = ("RECIPE_CHANGED", "SOURCE_CHANGED", "FINDINGS_CHANGED",
           "REVIEW_VIEW_CHANGED", "RENDERER_UNAVAILABLE")


@dataclass
class VerifyResult:
    checks: dict[str, str] = field(default_factory=dict)
    failures: list[str] = field(default_factory=list)
    historically_valid: bool = False
    meets_current_review_policy: bool = False
    policy_reason: str = ""

    def as_dict(self) -> dict:
        return {"checks": self.checks, "failures": self.failures,
                "historically_valid": self.historically_valid,
                "meets_current_review_policy": self.meets_current_review_policy,
                "policy_reason": self.policy_reason}


# ---------------------------------------------------------------------------
# the four hashes
# ---------------------------------------------------------------------------

def _sha(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _declared_sheets(recipe: Recipe, wb_path: Path) -> dict:
    """The sheets the recipe makes a claim about, with their machine-readable
    values. Coverage totality means a data recipe claims EVERY row and column of
    a data sheet, so the whole used range is in scope."""
    from openpyxl import load_workbook

    wb = load_workbook(wb_path)
    out: dict[str, dict] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        out[name] = {"rows": sheet_rows(ws), "n_cols": ws.max_column or 0}
    return out


def source_region_hash(recipe: Recipe, wb_path: Path) -> str:
    """Machine-readable values, never a visual rendering (rule 1).

    The workbook's sheet NAMES are included because "which sheets matter" was
    part of what was approved: a new sheet appearing is a change to the thing
    the reviewer signed off, even if no existing cell moved.
    """
    sheets = _declared_sheets(recipe, wb_path)
    payload = {"sheet_names": list(sheets),
               "values": {name: [[("" if v is None else str(v)) for v in row]
                                 for row in data["rows"]]
                          for name, data in sheets.items()}}
    return _sha(json.dumps(payload, ensure_ascii=False, sort_keys=True))


def detector_findings(wb_path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(wb_path)
    out: list[dict] = []
    for name in wb.sheetnames:
        out.extend(f.as_dict() for f in scan_sheet(wb[name], name))
    return sorted(out, key=lambda f: (f["kind"], f["referent"]))


def detector_findings_hash(findings: list[dict]) -> str:
    return _sha(json.dumps(findings, ensure_ascii=False, sort_keys=True))


def review_view(recipe: Recipe, wb_path: Path, renderer_version: str) -> Optional[str]:
    renderer = RENDERERS.get(renderer_version)
    if renderer is None:
        return None
    return renderer(_declared_sheets(recipe, wb_path), detector_findings(wb_path),
                    recipe, wb_path)


# ---------------------------------------------------------------------------
# make and verify
# ---------------------------------------------------------------------------

def make_record(recipe: Recipe, wb_path: Path, approved_by: str, approved_at: str,
                renderer_version: str = CURRENT_RENDERER_VERSION) -> ApprovalRecord:
    view = review_view(recipe, wb_path, renderer_version)
    if view is None:
        raise ValueError(f"unknown renderer {renderer_version!r}")
    return ApprovalRecord(
        recipe_id=recipe.recipe_id,
        recipe_hash=recipe.content_sha256(),
        source_region_hash=source_region_hash(recipe, wb_path),
        review_view_hash=_sha(view),
        detector_findings_hash=detector_findings_hash(detector_findings(wb_path)),
        renderer_version=renderer_version,
        approved_by=approved_by,
        approved_at=approved_at,
    )


def verify(record: ApprovalRecord, recipe: Recipe, wb_path: Path) -> VerifyResult:
    """Recompute every link, under the record's OWN renderer (rule 2)."""
    result = VerifyResult()

    result.checks["recipe"] = ("OK" if recipe.content_sha256() == record.recipe_hash
                               else "RECIPE_CHANGED")
    result.checks["source_region"] = ("OK" if source_region_hash(recipe, wb_path)
                                      == record.source_region_hash else "SOURCE_CHANGED")
    findings = detector_findings(wb_path)
    result.checks["detector_findings"] = (
        "OK" if detector_findings_hash(findings) == record.detector_findings_hash
        else "FINDINGS_CHANGED")

    view = review_view(recipe, wb_path, record.renderer_version)
    if view is None:
        # The renderer that produced the evidence is gone. That is NOT the same
        # as the evidence having changed, and conflating them would let a
        # deleted renderer read as tampering.
        result.checks["review_view"] = "RENDERER_UNAVAILABLE"
    else:
        result.checks["review_view"] = ("OK" if _sha(view) == record.review_view_hash
                                        else "REVIEW_VIEW_CHANGED")

    result.failures = [v for v in result.checks.values() if v != "OK"]
    result.historically_valid = not result.failures

    if not result.historically_valid:
        result.meets_current_review_policy = False
        result.policy_reason = f"not historically valid: {result.failures}"
    elif record.renderer_version != CURRENT_RENDERER_VERSION:
        # The load-bearing case. History is intact and must not be rewritten,
        # and the approval must not inherit protections the reviewer never had.
        result.meets_current_review_policy = False
        result.policy_reason = (
            f"approved under {record.renderer_version}; current policy requires "
            f"{CURRENT_RENDERER_VERSION}. The approval remains historically valid "
            f"— it simply does not acquire the newer renderer's protections")
    else:
        result.meets_current_review_policy = True
        result.policy_reason = f"approved under the current renderer {CURRENT_RENDERER_VERSION}"
    return result


# ---------------------------------------------------------------------------
# Self-test: a control, four independent mutations, and the renderer upgrade
# ---------------------------------------------------------------------------

def _self_test() -> int:
    import copy
    import tempfile

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    sys.path.insert(0, str(HERE))
    from recipe import recipe_from_json

    failures: list[str] = []

    def check(cond: bool, msg: str) -> None:
        if not cond:
            failures.append(msg)

    lab = HERE.parent.parent
    src = lab / "definition_phase" / "fixtures" / "W1_multisheet.xlsx"
    raw = json.loads((lab / "experimentK" / "recipes" / "W1_sales_v13_approved.json")
                     .read_text(encoding="utf-8"))
    recipe = recipe_from_json(raw)

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        wb_path = tmp / "w.xlsx"
        wb_path.write_bytes(src.read_bytes())

        record = make_record(recipe, wb_path, "designer", "2026-08-14T12:00:00Z")

        # --- control ---------------------------------------------------------
        r = verify(record, recipe, wb_path)
        check(r.historically_valid and r.meets_current_review_policy,
              f"control must verify: {r.as_dict()}")

        # --- 1. recipe changed after approval --------------------------------
        edited = copy.deepcopy(raw)
        edited["sheets"][0]["exclude"].append(
            {"referent": "sheet:Sales!8", "reason": "added after approval"})
        r1 = verify(record, recipe_from_json(edited), wb_path)
        check(r1.checks["recipe"] == "RECIPE_CHANGED" and not r1.historically_valid,
              f"a recipe edited after approval must fail on the RECIPE link: {r1.checks}")
        check(r1.checks["source_region"] == "OK",
              "…and must NOT be reported as a source change")

        # --- 2. source changed, recipe identical -----------------------------
        moved = tmp / "moved.xlsx"
        moved.write_bytes(src.read_bytes())
        wb = load_workbook(moved)
        wb["Sales"]["B5"] = 999
        wb.save(moved)
        r2 = verify(record, recipe, moved)
        check(r2.checks["source_region"] == "SOURCE_CHANGED" and r2.checks["recipe"] == "OK",
              f"an edited source with an identical recipe must fail on SOURCE: {r2.checks}")

        # --- 3. findings changed, VALUES IDENTICAL ---------------------------
        # The sharpest case: nothing a value-hash can see. The document became
        # less honest without any cell value moving.
        hidden = tmp / "hidden.xlsx"
        hidden.write_bytes(src.read_bytes())
        wb = load_workbook(hidden)
        cell = wb["Sales"]["A5"]
        cell.font = Font(color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF",
                                fill_type="solid")
        wb.save(hidden)
        r3 = verify(record, recipe, hidden)
        check(r3.checks["source_region"] == "OK",
              f"values did not change, so the source link must hold: {r3.checks}")
        check(r3.checks["detector_findings"] == "FINDINGS_CHANGED",
              f"an existing cell made invisible must fail on FINDINGS: {r3.checks}")
        check(not r3.historically_valid,
              "a document that became less honest must not verify")

        # --- 4. undeclared renderer change -----------------------------------
        # The renderer's OUTPUT changes while its version string does not. A
        # version alone cannot catch this; the view hash can.
        # Patches whatever the CURRENT renderer is, not a hardcoded name: the
        # record above was made under CURRENT_RENDERER_VERSION, and pinning this
        # to "review-v2" made the case silently stop testing anything the moment
        # v3 became current.
        original = RENDERERS[CURRENT_RENDERER_VERSION]
        try:
            RENDERERS[CURRENT_RENDERER_VERSION] = (
                lambda sheets, findings, recipe=None, wb_path=None:
                original(sheets, findings, recipe, wb_path) + "\n# extra")
            r4 = verify(record, recipe, wb_path)
        finally:
            RENDERERS[CURRENT_RENDERER_VERSION] = original
        check(r4.checks["review_view"] == "REVIEW_VIEW_CHANGED",
              f"a renderer changed without bumping its version must be caught: {r4.checks}")

        # --- 5. renderer upgrade — the load-bearing case ----------------------
        old = make_record(recipe, wb_path, "designer", "2026-01-01T09:00:00Z",
                          renderer_version="review-v1")
        r5 = verify(old, recipe, wb_path)
        check(r5.historically_valid,
              f"a v1 approval must REMAIN historically valid: {r5.as_dict()}")
        check(not r5.meets_current_review_policy,
              "…and must NOT silently acquire the current renderer's protections")
        check("review-v1" in r5.policy_reason
              and CURRENT_RENDERER_VERSION in r5.policy_reason,
              f"the reason must name both renderers: {r5.policy_reason}")

        # v1 genuinely showed less: the two renderings must differ, or the
        # distinction above would be ceremony rather than substance.
        v1 = review_view(recipe, wb_path, "review-v1")
        v2 = review_view(recipe, wb_path, "review-v2")
        check(v1 != v2, "the two renderers must actually differ")
        check("hidden content" in v2 and "hidden content" not in v1,
              "v2 must expose hidden-content findings that v1 did not")

        # --- 7. review-v3: Observable Error crossing the human boundary -------
        # Execution.degraded answers "did execution fail to deliver something the
        # recipe declared?". The RENDERER answers "was that degradation actually
        # available to the human granting authority?". These must stay separate,
        # or it becomes tempting to argue that because the execution object
        # contained it, the approval was informed -- which is OPEN-2 exactly.
        sys.path.insert(0, str(lab / "experimentL" / "harness"))
        from execute_recipe import execute as _execute  # noqa: E402

        from referents import WorkbookView as _WBV  # noqa: E402

        ex = _execute(recipe, _WBV(wb_path))
        check(ex.degraded,
              "this fixture must BE degraded, or case 7 proves nothing about "
              "showing degradation")

        v3 = review_view(recipe, wb_path, "review-v3")
        check("RESULT DEGRADED" not in v2,
              "review-v2 must NOT show degradation -- that is the gap v3 closes")
        check("RESULT DEGRADED" in v3,
              f"review-v3 must show that the result is degraded:\n{v3[-400:]}")
        check("unhonoured" in v3,
              "v3 must name WHICH declaration was not honoured, not merely that "
              "something was")
        for tgt in ("paivitetty", "tuote", "myynti"):
            check(tgt in v3, f"v3 must list the declared field {tgt!r}")

        # A v2 approval of a degraded recipe: historically valid, and NOT covered
        # by v3's policy. The whole point -- an approval granted before degraded
        # types were visible does not acquire that protection retroactively.
        v2_record = make_record(recipe, wb_path, "designer", "2026-02-01T09:00:00Z",
                                renderer_version="review-v2")
        r7 = verify(v2_record, recipe, wb_path)
        check(r7.historically_valid,
              f"a v2 approval must remain historically valid: {r7.as_dict()}")
        check(not r7.meets_current_review_policy,
              "a v2 approval must NOT acquire v3's degraded-type protection")

        # --- 8. CONTROL: v3 must not cry degradation when nothing is ----------
        # Same workbook, same shape, with the one unhonourable declaration
        # (`date`, gap G1) declared as `string` instead. If v3 shouted here it
        # would be noise, and a reviewer who sees it every time stops reading it.
        clean_raw = copy.deepcopy(raw)
        for entry in clean_raw["sheets"]:
            for fld in entry.get("fields", []):
                if fld.get("type") == "date":
                    fld["type"] = "string"
        clean_recipe = recipe_from_json(clean_raw)
        clean_ex = _execute(clean_recipe, _WBV(wb_path))
        check(not clean_ex.degraded,
              f"the control must NOT be degraded, or it controls nothing: "
              f"{clean_ex.degradation}")
        v3_clean = review_view(clean_recipe, wb_path, "review-v3")
        check("RESULT NOT DEGRADED" in v3_clean,
              "v3 must say plainly that a clean result is clean")
        check("RESULT DEGRADED" not in v3_clean,
              f"v3 must not report degradation on a clean recipe:\n{v3_clean[-300:]}")
        check("unhonoured" not in v3_clean,
              "a clean control must name no unhonoured declaration")

        # --- 6. renderer gone ------------------------------------------------
        ghost = ApprovalRecord(**{**asdict(record), "renderer_version": "review-v0"})
        r6 = verify(ghost, recipe, wb_path)
        check(r6.checks["review_view"] == "RENDERER_UNAVAILABLE",
              f"a missing renderer must be distinguishable from tampering: {r6.checks}")

        # every declared reason must be reachable
        seen = {v for res in (r, r1, r2, r3, r4, r5, r6) for v in res.checks.values()}
        untested = sorted(set(REASONS) - seen)
        check(not untested, f"declared but unexercised verification reasons: {untested}")

    if failures:
        sys.stderr.write("SELF-TEST FAILED:\n  " + "\n  ".join(failures) + "\n")
        return 1
    sys.stdout.write(
        "SELF-TEST PASSED (control verifies / recipe, source, findings and review-view "
        "changes each fail on their OWN link / invisible-ink change caught with values "
        "identical / undeclared renderer change caught / v1 approval stays historically "
        "valid and does NOT meet current policy / missing renderer distinguished from "
        "tampering / review-v3 shows declared-vs-observed types and v2 does not / a v2 "
        "approval of a degraded recipe stays valid without acquiring v3's protection / "
        "v3 stays quiet on a clean recipe / all 5 reasons exercised)\n")
    return 0


if __name__ == "__main__":
    if sys.argv[1:2] == ["--self-test"]:
        raise SystemExit(_self_test())
    sys.stderr.write("usage: approval.py --self-test\n")
    raise SystemExit(2)
