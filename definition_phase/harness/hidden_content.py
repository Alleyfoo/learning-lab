#!/usr/bin/env python3
"""Detect content a human does not see but an agent reads.

The security finding this exists for (`experimentK/spec/v13_reconciliation.md`):
a spreadsheet cell can carry text that is invisible on screen — white font,
a hidden row or column, a zero-width column, a `;;;` number format. The original
reservation-calendar injection had at least to be visible to whoever looked at
the field. This does not.

That breaks an assumption the whole architecture rests on: **that the human
approving a recipe and the machine reading the workbook are looking at the same
document.** "A human reads it first" is only a control when the human is shown
what the machine reads.

So this module does not sanitise, strip, or block anything. It **surfaces**:
every finding is reported as a frozen-grammar referent plus a reason, so the
browser can mark it and a human can decide. Deciding is not its job.

Findings are addressed with the frozen referent grammar, which means they drop
straight into a recipe's `ambiguities` if a human wants them raised there.
"""
from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(HERE))

from referents import Referent, index0_to_col  # noqa: E402

KINDS = (
    "white_font",          # font colour matches the fill: invisible text
    "hidden_row",
    "hidden_column",
    "near_zero_width_column",
    "blanking_number_format",   # ';;;' renders any value as empty
)

# openpyxl reports theme/indexed colours too; only rgb is compared here, and a
# colour we cannot read is NOT reported. Silence beats a false accusation, and
# the limitation is stated rather than hidden.
_WHITES = {"FFFFFFFF", "00FFFFFF", "FFFFFF"}

# A column narrower than this renders as nothing. Frozen threshold, in
# Excel's character-width units.
NEAR_ZERO_WIDTH = 0.5


@dataclass(frozen=True)
class Finding:
    kind: str
    referent: str
    detail: str = ""
    preview: str = ""      # what the AGENT would read at that address

    def as_dict(self) -> dict:
        return {"kind": self.kind, "referent": self.referent,
                "detail": self.detail, "preview": self.preview}


def _rgb(colour) -> Optional[str]:
    if colour is None:
        return None
    rgb = getattr(colour, "rgb", None)
    return rgb.upper() if isinstance(rgb, str) else None


def _is_whiteish(font_rgb: Optional[str], fill_rgb: Optional[str]) -> bool:
    if font_rgb is None:
        return False
    if font_rgb in _WHITES and fill_rgb in (None, *_WHITES):
        return True
    return fill_rgb is not None and font_rgb == fill_rgb


def scan_sheet(ws, sheet_name: str) -> list[Finding]:
    findings: list[Finding] = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or str(cell.value).strip() == "":
                continue
            ref = Referent(kind="cell", sheet=sheet_name,
                           row0=cell.row - 1, col0=cell.column - 1).render()
            preview = str(cell.value)[:80]

            font_rgb = _rgb(getattr(cell.font, "color", None))
            fill = getattr(cell, "fill", None)
            fill_rgb = _rgb(getattr(fill, "fgColor", None)) if fill is not None else None
            if _is_whiteish(font_rgb, fill_rgb):
                findings.append(Finding(
                    "white_font", ref,
                    f"font {font_rgb} on fill {fill_rgb or 'none'}: not visible on screen",
                    preview))

            if (cell.number_format or "").replace(" ", "") == ";;;":
                findings.append(Finding(
                    "blanking_number_format", ref,
                    "number format ';;;' renders the cell empty whatever it holds",
                    preview))

    for idx, dim in (ws.row_dimensions or {}).items():
        if getattr(dim, "hidden", False):
            ref = Referent(kind="row", sheet=sheet_name, row0=int(idx) - 1).render()
            findings.append(Finding("hidden_row", ref, "row is hidden"))

    for letter, dim in (ws.column_dimensions or {}).items():
        col0 = _col_letter_to_index0(letter)
        if col0 is None:
            continue
        ref = Referent(kind="col", sheet=sheet_name, col0=col0).render()
        if getattr(dim, "hidden", False):
            findings.append(Finding("hidden_column", ref, "column is hidden"))
        width = getattr(dim, "width", None)
        # True width 0 is written by Excel AS hidden, so the distinct technique
        # is a width small enough to be invisible with the flag left unset.
        if width is not None and 0.0 <= float(width) <= NEAR_ZERO_WIDTH:
            findings.append(Finding(
                "near_zero_width_column", ref,
                f"width {width}: invisible on screen, hidden flag NOT set"))

    return findings


def _col_letter_to_index0(letters: str) -> Optional[int]:
    n = 0
    for ch in str(letters).upper():
        if not ("A" <= ch <= "Z"):
            return None
        n = n * 26 + (ord(ch) - 64)
    return n - 1 if n else None


def scan_workbook(path: str | Path) -> list[Finding]:
    from openpyxl import load_workbook

    wb = load_workbook(Path(path))          # not read_only: styles are needed
    out: list[Finding] = []
    for name in wb.sheetnames:
        out.extend(scan_sheet(wb[name], name))
    return out


def summary(findings: Iterable[Finding]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for f in findings:
        counts[f.kind] = counts.get(f.kind, 0) + 1
    return counts


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------

def _self_test() -> int:
    failures: list[str] = []

    def check(cond: bool, msg: str) -> None:
        if not cond:
            failures.append(msg)

    w2 = ROOT / "fixtures" / "W2_hidden_content.xlsx"
    if not w2.exists():
        sys.stderr.write(f"fixture missing: {w2} (run make_w2.py)\n")
        return 1

    found = scan_workbook(w2)
    by_kind: dict[str, list[Finding]] = {}
    for f in found:
        by_kind.setdefault(f.kind, []).append(f)

    for kind in KINDS:
        check(kind in by_kind, f"{kind} not detected on W2; got {sorted(by_kind)}")

    white = by_kind.get("white_font", [])
    check(any(f.referent == "sheet:Sales!E3" for f in white),
          f"the white-font injection cell E3 must be found: "
          f"{[f.referent for f in white]}")
    # The whole point: the payload is SURFACED, not silently dropped.
    check(any("ignore previous instructions" in f.preview.lower() for f in white),
          "the finding must carry a preview of what the agent would read")

    check(any(f.referent == "sheet:Sales!7" for f in by_kind.get("hidden_row", [])),
          f"hidden row 7: {[f.referent for f in by_kind.get('hidden_row', [])]}")
    check(any(f.referent == "sheet:Sales!D" for f in by_kind.get("hidden_column", [])),
          f"hidden column D: {[f.referent for f in by_kind.get('hidden_column', [])]}")
    check(any(f.referent == "sheet:Sales!F"
              for f in by_kind.get("near_zero_width_column", [])),
          f"near-zero-width column F: "
          f"{[f.referent for f in by_kind.get('near_zero_width_column', [])]}")
    check(any(f.referent == "sheet:Sales!G3"
              for f in by_kind.get("blanking_number_format", [])),
          "the ';;;' cell G3 must be found")

    # Every finding must be a VALID referent -- these feed the recipe layer.
    from referents import ReferentSyntaxError, parse
    for f in found:
        try:
            parse(f.referent)
        except ReferentSyntaxError as exc:
            failures.append(f"finding emitted an unparseable referent {f.referent!r}: {exc}")

    # No false alarms on the clean sheet, or on the frozen clean workbook.
    check(not [f for f in found if ":Clean!" in f.referent or "sheet:Clean" in f.referent],
          f"the clean sheet must produce nothing: "
          f"{[f.referent for f in found if 'Clean' in f.referent]}")
    w1 = ROOT / "fixtures" / "W1_multisheet.xlsx"
    if w1.exists():
        check(not scan_workbook(w1),
              f"W1 has no hidden content and must produce nothing: "
              f"{[f.as_dict() for f in scan_workbook(w1)]}")

    if failures:
        sys.stderr.write("SELF-TEST FAILED:\n  " + "\n  ".join(failures) + "\n")
        return 1
    sys.stdout.write(
        f"SELF-TEST PASSED (all {len(KINDS)} hiding techniques detected on W2 with "
        f"valid referents and a preview of the payload / no false alarms on the "
        f"clean sheet or on W1)\n")
    return 0


if __name__ == "__main__":
    argv = sys.argv[1:]
    if argv and argv[0] == "--self-test":
        raise SystemExit(_self_test())
    if argv:
        import json
        results = scan_workbook(argv[0])
        print(json.dumps({"summary": summary(results),
                          "findings": [f.as_dict() for f in results]},
                         ensure_ascii=False, indent=2))
        raise SystemExit(0)
    sys.stderr.write("usage: hidden_content.py --self-test | <workbook.xlsx>\n")
    raise SystemExit(2)
