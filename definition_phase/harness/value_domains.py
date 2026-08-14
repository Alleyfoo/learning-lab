#!/usr/bin/env python3
"""Value domains — vary what flows through an unchanging structure.

Expansion 2, taken alone. Multiplicity is frozen; cross-sheet interaction is not
opened. **The schema is held utterly fixed** so that anything found here is
attributable to interpretation rather than to composition again:

```text
id     -> string
value  -> string
```

Only the source value changes.

## The observation boundary

    the machine-readable TYPED VALUE ADMITTED BY INGESTION, against the
    TYPED VALUE EMITTED BY EXECUTION, under the accepted recipe

Not Excel's visual rendering. Ingestion here means what the system itself can
see — `openpyxl` with `data_only=True`, the same world the executor reads — so
this tests interpretation, not ingestion policy.

A `string` declaration authorises **representation as text and nothing else**:
not whitespace removal, not numeric parsing, not case folding, not Unicode
normalisation. Whether normalisation *should* be authorised is a question for
the language, and this corpus exists partly to force that decision rather than
let it be made accidentally by a helpful library.
"""
from __future__ import annotations

import json
import sys
import tempfile
import unicodedata
from pathlib import Path

HERE = Path(__file__).resolve().parent
LAB = HERE.parent.parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(LAB / "experimentL" / "harness"))

from primitive_invariants import interpretation_violation  # noqa: E402
from recipe import recipe_from_json  # noqa: E402
from referents import WorkbookView  # noqa: E402
from validate_recipe import validate  # noqa: E402
from execute_recipe import InsufficientRecipe, execute  # noqa: E402

# One boring schema, held fixed for every case.
SCHEMA = [
    {"target": "id", "source": "sheet:S!@Tuote", "role": "id", "type": "string"},
    {"target": "value", "source": "sheet:S!@V", "role": "measure", "type": "string"},
]

VALUES = [
    ("empty_string", ""),
    ("single_space", " "),
    ("leading_zeroes", "00123"),
    ("negative", "-1"),
    ("float_text", "1.0"),
    ("comma_decimal", "1,0"),
    ("bool_upper", "TRUE"),
    ("bool_padded", " true "),
    ("null_word", "null"),
    ("none_word", "None"),
    ("date_slash", "03/04/2026"),
    ("date_dots", "3.2.2026"),
    ("exponent", "1e6"),
    ("nan", "NaN"),
    ("infinity", "∞"),
    ("unicode_composed", "é"),          # e-acute, single code point
    ("unicode_decomposed", "é"),       # e + combining acute
    ("formula_text", "=SUM(A1:A2)"),
    ("big_number_text", "9" * 30),
    ("trailing_space", "abc "),
    ("real_int", 123),                       # not a string at all
    ("real_float", 1.5),
]


def _recipe() -> dict:
    return {"recipe_version": 1, "recipe_id": "values", "workbook": {},
            "sheets": [{"sheet": "sheet:S", "role": "data",
                        "header_row": "sheet:S!1", "data_region": "remainder",
                        "fields": SCHEMA, "exclude": [], "ambiguities": []}],
            "applicability": None,
            "provenance": {"proposed_by": "values", "approved_by": "values",
                           "approved_recipe_sha256": None}}


def run_all() -> dict:
    from openpyxl import Workbook, load_workbook

    findings: list[dict] = []
    observations: list[dict] = []
    checked = executed = 0

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        for name, value in VALUES:
            path = tmp / f"{name}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "S"
            ws.append(["Tuote", "V"])
            ws.append(["E1", value])
            wb.save(path)

            # INGESTION: what the system itself can see.
            admitted = load_workbook(path, data_only=True)["S"].cell(row=2, column=2).value

            raw = _recipe()
            r = recipe_from_json(raw)
            raw["provenance"]["approved_recipe_sha256"] = r.content_sha256()
            r = recipe_from_json(raw)

            view = WorkbookView(path)
            report = validate(r, view)
            checked += 1
            if not report.valid:
                observations.append({"case": name, "written": repr(value),
                                     "note": f"refused: {sorted(report.codes())}"})
                continue
            try:
                ex = execute(r, view)
            except InsufficientRecipe as exc:
                observations.append({"case": name, "written": repr(value),
                                     "note": f"executor refused: {exc}"})
                continue
            executed += 1
            emitted = ex.rows[0][ex.columns.index("value")]

            why = interpretation_violation(admitted, emitted, "string")
            record = {"case": name, "written": repr(value),
                      "admitted": f"{type(admitted).__name__}({admitted!r})",
                      "emitted": f"{type(emitted).__name__}({emitted!r})"}
            if why:
                findings.append({**record, "detail": why})

            # Unicode identity is a SPECIFICATION question, not a bug report:
            # same visible glyph, different code points. Recorded either way.
            if isinstance(admitted, str) and isinstance(emitted, str):
                if admitted != emitted and \
                        unicodedata.normalize("NFC", admitted) == unicodedata.normalize("NFC", emitted):
                    observations.append({**record,
                                         "note": "differs only by Unicode normalisation"})

    return {"checked": checked, "executed": executed,
            "findings": findings, "n_findings": len(findings),
            "observations": observations}


def _self_test() -> int:
    out = run_all()
    sys.stdout.write(f"  value cases       : {out['checked']}\n"
                     f"  executed          : {out['executed']}\n"
                     f"  interpretation    : {out['n_findings']} violation(s)\n\n")
    for f in out["findings"]:
        sys.stdout.write(f"  VIOLATION {f['case']:20} admitted {f['admitted']:26} "
                         f"emitted {f['emitted']}\n")
    for o in out["observations"]:
        sys.stdout.write(f"  note      {o['case']:20} {o.get('note','')[:70]}\n")

    # Never overwrite a previous run. Run 1 recorded PRO-2 instance 9 and is
    # evidence; a re-run after the fix is a different measurement, not a
    # replacement for it.
    results = HERE.parent / "results"
    results.mkdir(parents=True, exist_ok=True)
    n = 1
    while (results / f"value_domains_run{n}.json").exists():
        n += 1
    path = results / f"value_domains_run{n}.json"
    path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    sys.stdout.write(f"\n  written to {path.name}\n")
    # Findings are reported, not thrown: this is a measurement run.
    return 0


if __name__ == "__main__":
    raise SystemExit(_self_test())
