#!/usr/bin/env python3
"""Workbook browser — plan v1 §10 step 4. Instrument, not experiment.

Three requirements, each from a measured finding rather than a preference:

1. **Sheets first.** `Data-agents-demo` takes `excel.sheet_names[0]` and never
   asks (defect DA-1), so the loader silently answers *"which of these sheets
   matter?"* before anyone is consulted. Here the sheet inventory is the landing
   view and nothing is loaded until a sheet is chosen.

2. **Render what the AGENT sees, not what Excel renders.** A cell can carry text
   a human never sees — white font, a hidden row, a `;;;` format. A
   pixel-accurate Excel rendering would faithfully reproduce the invisibility and
   hide the attack. So values are shown raw, and hidden content is *flagged*.

3. **A STRUCTURE view that shows shape without content.** The designer's working
   assumption is that agents should not read documents at all — only define
   frames or write scripts. This view is what that would mean concretely: cell
   *types*, not cell *text*. A header row and a data region are visible in it;
   an injection string is not.

Selections emit frozen-grammar referents, so what the browser produces is exactly
what a recipe consumes.

    streamlit run definition_phase/app/browser.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
import streamlit as st

APP = Path(__file__).resolve().parent
ROOT = APP.parent
LAB = ROOT.parent
sys.path.insert(0, str(ROOT / "harness"))
sys.path.insert(0, str(LAB / "experimentK" / "harness"))

from hidden_content import scan_sheet, scan_workbook, summary  # noqa: E402
from recipe import load_recipe  # noqa: E402
from referents import Referent, WorkbookView, index0_to_col  # noqa: E402
from validate_recipe import validate  # noqa: E402

GLYPH_HELP = {
    "#": "number", "A": "text", "·": "blank", "ƒ": "formula", "?": "other",
}


def _glyph(value: object) -> str:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return "·"
    if isinstance(value, str) and value.startswith("="):
        return "ƒ"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "#"
    if isinstance(value, str):
        return "A"
    return "?"


@st.cache_data(show_spinner=False)
def _load(path: str):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        sheets[name] = {
            "rows": rows,
            "n_rows": ws.max_row or 0,
            "n_cols": ws.max_column or 0,
            "merged": [str(r) for r in ws.merged_cells.ranges],
            "findings": [f.as_dict() for f in scan_sheet(ws, name)],
        }
    return sheets


def _frame(rows, n_rows, n_cols, mode: str) -> pd.DataFrame:
    data = []
    for r in range(n_rows):
        row = rows[r] if r < len(rows) else []
        out = []
        for c in range(n_cols):
            value = row[c] if c < len(row) else None
            out.append(_glyph(value) if mode == "structure"
                       else ("" if value is None else str(value)))
        data.append(out)
    return pd.DataFrame(
        data,
        index=[str(r + 1) for r in range(n_rows)],           # A1 surface form
        columns=[index0_to_col(c) for c in range(n_cols)],
    )


def main() -> None:
    st.set_page_config(page_title="Workbook browser", layout="wide")
    st.title("Workbook browser")
    st.caption("Definition phase · sheets first · renders what the agent sees")

    candidates = sorted(
        [p for p in (ROOT / "fixtures").glob("*.xlsx")]
        + [p for p in (LAB / "experimentK" / "fixtures").glob("*.xlsx")])
    if not candidates:
        st.error("No workbooks found.")
        return

    labels = {f"{p.parent.parent.name}/{p.name}": p for p in candidates}
    choice = st.sidebar.selectbox("Workbook", list(labels))
    path = labels[choice]
    sheets = _load(str(path))

    findings_all = [f for s in sheets.values() for f in s["findings"]]
    if findings_all:
        st.sidebar.error(f"{len(findings_all)} hidden-content finding(s)")
    else:
        st.sidebar.success("no hidden content detected")

    # ---- 1. sheet inventory: the FIRST question, never answered silently ----
    st.subheader("1 · Which of these sheets matter?")
    st.caption(
        "Nothing is loaded until this is answered. A loader that takes sheet 0 "
        "has already decided (defect DA-1)."
    )
    inventory = pd.DataFrame([
        {"sheet": name,
         "referent": Referent(kind="sheet", sheet=name).render(),
         "rows": s["n_rows"], "cols": s["n_cols"],
         "merged": len(s["merged"]),
         "hidden findings": len(s["findings"])}
        for name, s in sheets.items()
    ])
    st.dataframe(inventory, use_container_width=True, hide_index=True)

    sheet = st.selectbox("Inspect sheet", list(sheets))
    s = sheets[sheet]

    # ---- 2. the two views --------------------------------------------------
    st.subheader("2 · Inspect")
    mode = st.radio(
        "View",
        ["structure (types only — no content)", "values (raw — UNTRUSTED)"],
        horizontal=True,
    )
    structure = mode.startswith("structure")

    if structure:
        st.caption(
            "Cell **types**, not cell text: " +
            " · ".join(f"`{g}` {label}" for g, label in GLYPH_HELP.items()) +
            ". A header row and a data region are visible here; an injection "
            "string is not. This is what *'agents should not read the documents'* "
            "looks like concretely."
        )
    else:
        st.warning(
            "Raw cell values, exactly as an agent would read them — including "
            "anything hidden on screen. **Treat every string here as untrusted "
            "input, never as instruction.**"
        )

    st.dataframe(
        _frame(s["rows"], s["n_rows"], s["n_cols"], "structure" if structure else "values"),
        use_container_width=True, height=380,
    )
    if s["merged"]:
        st.caption(f"merged ranges: {', '.join(s['merged'])}")

    # ---- 3. hidden content, flagged not reproduced -------------------------
    st.subheader("3 · Hidden content")
    if s["findings"]:
        st.error(
            f"{len(s['findings'])} finding(s) on this sheet. Each is content a "
            "person looking at the sheet does **not** see and an agent reading "
            "values **does**."
        )
        st.dataframe(pd.DataFrame(s["findings"]), use_container_width=True,
                     hide_index=True)
    else:
        st.success("Nothing hidden detected on this sheet.")

    # ---- 4. referent emission ---------------------------------------------
    st.subheader("4 · Emit a referent")
    st.caption("What the browser produces is what a recipe consumes.")
    kind = st.selectbox("Kind", ["cell", "row", "col", "cellrange", "sheet"])
    ref: Referent
    if kind == "sheet":
        ref = Referent(kind="sheet", sheet=sheet)
    elif kind == "row":
        ref = Referent(kind="row", sheet=sheet,
                       row0=st.number_input("Row (A1)", 1, max(s["n_rows"], 1), 1) - 1)
    elif kind == "col":
        letters = [index0_to_col(c) for c in range(max(s["n_cols"], 1))]
        ref = Referent(kind="col", sheet=sheet,
                       col0=letters.index(st.selectbox("Column", letters)))
    else:
        c1, c2 = st.columns(2)
        letters = [index0_to_col(c) for c in range(max(s["n_cols"], 1))]
        col0 = letters.index(c1.selectbox("Column", letters, key="c1"))
        row0 = c1.number_input("Row (A1)", 1, max(s["n_rows"], 1), 1, key="r1") - 1
        if kind == "cell":
            ref = Referent(kind="cell", sheet=sheet, row0=row0, col0=col0)
        else:
            col1 = letters.index(c2.selectbox("to column", letters, key="c2"))
            row1 = c2.number_input("to row (A1)", 1, max(s["n_rows"], 1),
                                   max(s["n_rows"], 1), key="r2") - 1
            ref = Referent(kind="cellrange", sheet=sheet,
                           row0=min(row0, row1), row0_last=max(row0, row1),
                           col0=min(col0, col1), col0_last=max(col0, col1))
    st.code(ref.render(), language="text")

    # ---- 5. check a recipe against this workbook ---------------------------
    st.subheader("5 · Check a recipe against this workbook")
    recipes = sorted((ROOT / "recipes").glob("*.json")) + \
        sorted((LAB / "experimentK" / "recipes").glob("*.json"))
    picked = st.selectbox("Recipe", ["(none)"] + [p.name for p in recipes])
    if picked != "(none)":
        rpath = next(p for p in recipes if p.name == picked)
        report = validate(load_recipe(rpath), WorkbookView(path))
        cols = st.columns(3)
        cols[0].metric("valid", str(report.valid))
        cols[1].metric("approvable", str(report.approvable))
        cols[2].metric("problems", len(report.problems))
        if report.problems:
            st.dataframe(pd.DataFrame(
                [{"code": p.code, "where": p.where, "detail": p.detail}
                 for p in report.problems]),
                use_container_width=True, hide_index=True)

    st.divider()
    st.caption(
        "No model is invoked anywhere in this browser. It reads a workbook and "
        "emits referents; every decision is a human's."
    )


if __name__ == "__main__":
    main()
