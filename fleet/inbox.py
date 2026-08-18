#!/usr/bin/env python3
"""A deterministic inbox around an established worker. No LLM anywhere.

```text
<worker>/
  inbox/        a file landing here is the trigger
  processed/    the run completed -- accepted, or refused by policy
  exceptions/   the run could not complete, or its effect did not land
  ledger.jsonl  append-only work-item state. This is the twice-protection.
```

## Why a ledger and not "did the file move"

A file's location is a consequence, not a record. If a process dies between
applying an effect and moving the file, the file is still in `inbox/` and a
naive poller reruns it — and for a worker with an effect that is a duplicate
booking, not a retry.

So an item is **claimed in the ledger before it is run**, and the ledger is what
the next poll consults. A crash leaves a visible `claimed` line rather than an
item that silently looks fresh.

## Payload digest and work-item identity are different things

They coincide for this worker and must not be assumed to in general.

```text
payload_digest   sha256 of the file's bytes. A fact ABOUT the content, always
                 recorded, useful for integrity regardless of policy.
item_id          what makes two arrivals the SAME WORK. A policy, declared per
                 worker in `work_item_identity`.
```

`room-reservation` declares `content_digest`, so identical bytes are the same
booking and must not be applied twice. That is right *for this worker*. Another
worker might identify work by an invoice number inside the file, or treat every
arrival as distinct even when the bytes repeat. Deliberately not generalised: the
only policy implemented is `content_digest`, and an undeclared or unknown one is
refused rather than defaulted.

## What lands where

```text
completed   accepted with its effect applied, OR refused by policy
            -> processed/    a policy refusal is a healthy, finished run
exception   the run failed, or an accepted decision's effect did not land
            -> exceptions/   retryable, deliberately, by moving it back
```
"""
from __future__ import annotations

import hashlib
import json
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

import fleet  # noqa: E402
import input_set  # noqa: E402

sys.path.insert(0, str(HERE.parent / "worker"))
import runtime  # noqa: E402

sys.path.insert(0, str(HERE.parent))


def _xlsx():
    """Imported on use, not on import.

    An optional input adapter must not be a hard dependency of the console: a
    fleet with no workbook worker should still be readable on a machine without
    openpyxl. Found the moment the console was opened.
    """
    import adapters.xlsx as module
    return module


def _sheet_specs(xlsx, w: "fleet.Worker") -> list:
    """The SheetSpecs an xlsx arrival is validated against.

    v0.6 source of truth: the worker's version-bound `input_contract`
    (`input_contracts/v<N>.json`), via `specs_from_contract` -- a contract
    answers what data representation this version is allowed to do it with.
    Falls back to the legacy `identity["adapter_sheets"]` shape when no
    contract exists, so workers not yet migrated (and non-sheet workers never
    reaching this branch) keep working unchanged. The back-compat gate.
    """
    contract = w.input_contract
    if contract is not None:
        return list(xlsx.specs_from_contract(contract).values())
    return xlsx.specs_from(w.identity["adapter_sheets"])

LEDGER = "ledger.jsonl"
FOLDERS = ("inbox", "processed", "exceptions")

# The only work-item identity policy implemented. Not a default -- a worker that
# declares nothing, or something else, is refused.
IDENTITY_POLICIES = ("content_digest",)

TERMINAL = ("completed", "exception", "skipped_duplicate",
            "recovered_completed", "recovered_exception")


def retained_name(document: str, digest: str) -> str:
    """Digest-namespaced raw-retention name: ``<digest>.<ext>``.

    Two same-named different-bytes arrivals both survive in ``processed/``
    (design §6.2) -- the bare filename would overwrite, losing the exact raw
    evidence an earlier run recorded. The original filename is kept in the
    ledger ``file`` field, not on disk; the digest is the truth. This is raw
    retention only: the materialized JSON is still overwritten per run.
    """
    return f"{digest}{Path(document).suffix}"


class UnknownIdentityPolicy(Exception):
    """The worker declares a work-item identity this inbox cannot compute."""


class CrashInjected(Exception):
    """Raised only by the self-test, at a named point inside a pass."""


@dataclass(frozen=True)
class Item:
    path: Path
    payload_digest: str
    item_id: str
    request: str


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def ensure(w: fleet.Worker) -> None:
    for folder in FOLDERS:
        (w.directory / folder).mkdir(exist_ok=True)


def ledger(w: fleet.Worker) -> list[dict]:
    path = w.directory / LEDGER
    if not path.is_file():
        return []
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()
            if line.strip()]


def _append(w: fleet.Worker, record: dict) -> None:
    with (w.directory / LEDGER).open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def item_state(w: fleet.Worker, item_id: str,
               entries: Optional[list] = None) -> Optional[str]:
    """The LAST recorded state for an item, or None if never seen."""
    state = None
    for entry in entries if entries is not None else ledger(w):
        if entry["item_id"] == item_id:
            state = entry["state"]
    return state


# ---------------------------------------------------------------------------
# the worker's verifiable effect -- what recovery reconciles against
# ---------------------------------------------------------------------------

def effect_landed(w: fleet.Worker, request: str) -> Optional[bool]:
    """Whether this request's effect is present in worker state RIGHT NOW.

    `None` means the question could not be answered -- unreadable state, not a
    negative. Recovery treats those differently, which is the whole point.
    """
    if not w.committing:
        return None
    try:
        parsed = runtime.task_model.parse(w.model)
        _, items = runtime._read_state(parsed, w.base)
    except Exception:                              # noqa: BLE001
        # Deliberately NOT `runtime._landed`, which returns False when the
        # state cannot be read. For the commit path that is right -- an effect
        # that cannot be verified did not land. For recovery it is wrong:
        # "definitely absent" and "cannot tell" lead to opposite actions, and
        # collapsing them would retry a booking that may already exist.
        return None
    field = runtime.source_field(parsed, "reservations")
    return any((item.get(field) if isinstance(item, dict) else item) == request
               for item in items)


def _precondition(w: fleet.Worker, request: str) -> dict:
    """Recorded WITH the claim, so recovery has something to compare against.

    Without it, "the date is in state" is ambiguous: this run may have put it
    there, or it may have been there all along and the decision was a refusal.
    """
    landed = effect_landed(w, request)
    size = None
    if w.committing:
        try:
            parsed = runtime.task_model.parse(w.model)
            _, items = runtime._read_state(parsed, w.base)
            size = len(items)
        except Exception:                          # noqa: BLE001
            size = None
    return {"committing": w.committing, "effect_target_present": landed,
            "state_size": size}


def identity_policy(w: fleet.Worker) -> str:
    policy = w.identity.get("work_item_identity")
    if policy not in IDENTITY_POLICIES:
        raise UnknownIdentityPolicy(
            f"{w.name} declares work_item_identity={policy!r}; this inbox "
            f"implements {IDENTITY_POLICIES}")
    return policy


def read_item(w: fleet.Worker, path: Path) -> Item:
    raw = path.read_bytes()
    digest = hashlib.sha256(raw).hexdigest()
    if path.suffix == ".xlsx":
        # A workbook carries the DATA, not a request. It is converted into the
        # worker's declared sources and the worker then runs as it always does.
        payload = {"request": None}
    else:
        payload = json.loads(raw.decode("utf-8"))
    policy = identity_policy(w)
    # One policy today. The branch exists so a second one has somewhere to go
    # and so the coincidence stays visible rather than becoming an assumption.
    item_id = digest if policy == "content_digest" else None
    return Item(path, digest, item_id, payload["request"])


def waiting(w: fleet.Worker) -> list[Path]:
    """Files in the inbox, in a deterministic order.

    A worker declaring an input adapter also accepts that adapter's file type.
    The adapter converts; it does not decide anything.
    """
    patterns = ["*.json"]
    if w.identity.get("input_adapter") == "xlsx":
        patterns.append("*.xlsx")
    found: list[Path] = []
    for pattern in patterns:
        found += list((w.directory / "inbox").glob(pattern))
    return sorted(found)


def _role_sidecar(path: Path) -> Path:
    """The operator's explicit slot choice for a sole-slot file: a
    `<filename>.role` sidecar carrying the role name. Binding authority is an
    operator act, never filename inference (design §4)."""
    return path.with_name(path.name + ".role")


def _target_roles(w: fleet.Worker, path: Path) -> Optional[list[str]]:
    """Which roles an arriving file binds.

    shared-slot worker: every shared role -- one workbook fills them all
    (fazerish). sole-slot worker: the one role named in the `<file>.role`
    sidecar (acme) -- None if no sidecar was written, which becomes an
    exception rather than a guess from the filename. Mixed sole/shared slot
    workers are beyond v0.6's two workers and are refused here.
    """
    shared = input_set.shared_roles(w)
    sole = input_set.sole_roles(w)
    if shared and not sole:
        return shared
    if sole and not shared:
        side = _role_sidecar(path)
        if not side.is_file():
            return None
        role = side.read_text(encoding="utf-8").strip()
        if role not in sole:
            return None
        return [role]
    raise ValueError(
        f"{w.name} mixes sole and shared slots; v0.6 supports one kind only")


def _terminalize_set(w: fleet.Worker, completing: Item, run: dict) -> list[dict]:
    """A complete set's run fired. Terminalize EVERY bound document: one
    ledger line per bound file (carrying every role it fills), and each raw
    file drained from inbox/ to processed/ or exceptions/. Then clear the set.

    The run is ONE record_run for the whole set; the ledger is per-item, so
    every bound document reaches a terminal state. The completing item's
    claim was written by `poll`; staged members were claimed+staged earlier.
    A shared workbook that fills several roles is ONE outcome carrying all of
    them (one file in, one terminal record out); a multi-sole set is one
    outcome per document.
    """
    doc = input_set.load(w) or {"roles": {}}
    healthy = run["ok"]
    state = "completed" if healthy else "exception"
    destination = "processed" if healthy else "exceptions"
    # group roles by the bound document (a shared workbook fills several)
    by_doc: dict[str, dict] = {}
    for role, binding in doc.get("roles", {}).items():
        entry = by_doc.setdefault(binding["document"],
                                  {"digest": binding["digest"], "roles": []})
        entry["roles"].append(role)
    outcomes: list[dict] = []
    for document, info in by_doc.items():
        is_completing = info["digest"] == completing.payload_digest
        record = {"at": _now(), "item_id": info["digest"],
                  "payload_digest": info["digest"], "file": document,
                  "state": state, "roles": sorted(info["roles"]),
                  "request": completing.request if is_completing else None,
                  "decision": run.get("decision") if is_completing else None,
                  "reason": run.get("reason") if is_completing else None,
                  "effect_applied": (run.get("effect_applied")
                                     if is_completing else None),
                  "problems": (run.get("problems", []) if is_completing
                               else [])}
        _append(w, record)
        src = w.directory / "inbox" / document
        if src.is_file():
            shutil.move(str(src),
                        w.directory / destination / retained_name(document, info["digest"]))
        side = _role_sidecar(src)
        if side.is_file():
            side.unlink()
        outcomes.append(record)
    input_set.clear(w)
    return outcomes


def _poll_contract_item(w: fleet.Worker, path: Path, item: Item,
                        crash_at: Optional[str]) -> list[dict]:
    """The v0.6 per-item flow for a worker with an input_contract.

    bind -> (partial: stage) | (complete: run + terminalize the set). A shape
    mismatch at bind is an exception with no partial mutation. The claim was
    already written by `poll`; this owns everything after it.
    """
    roles = _target_roles(w, path)
    if roles is None:
        record = {"at": _now(), "item_id": item.item_id,
                  "payload_digest": item.payload_digest, "file": path.name,
                  "state": "exception", "request": item.request,
                  "roles": [], "reason": "no slot assigned to this file; "
                                         "bind it to a role explicitly "
                                         "(filename is not authority)"}
        _append(w, record)
        shutil.move(str(path),
                    w.directory / "exceptions" / retained_name(path.name, item.payload_digest))
        side = _role_sidecar(path)
        if side.is_file():
            side.unlink()
        return [record]

    result = input_set.bind(w, roles, path)
    if result["problems"]:
        record = {"at": _now(), "item_id": item.item_id,
                  "payload_digest": item.payload_digest, "file": path.name,
                  "state": "exception", "request": item.request, "roles": roles,
                  "problems": result["problems"],
                  "reason": "the workbook could not be converted faithfully "
                            "against the slot's contract"}
        _append(w, record)
        shutil.move(str(path),
                    w.directory / "exceptions" / retained_name(path.name, item.payload_digest))
        side = _role_sidecar(path)
        if side.is_file():
            side.unlink()
        return [record]

    if not result["complete"]:
        # Stage: the set is partial. The file stays in inbox/ (the set, not the
        # file location, is the authority); the sidecar is consumed (the
        # binding is recorded in input_set.json). A re-poll sees "staged" and
        # skips it until the remaining slots arrive.
        record = {"at": _now(), "item_id": item.item_id,
                  "payload_digest": item.payload_digest, "file": path.name,
                  "state": "staged", "roles": roles, "request": item.request}
        _append(w, record)
        side = _role_sidecar(path)
        if side.is_file():
            side.unlink()
        if crash_at == "after_bind_stage":
            raise CrashInjected("after_bind_stage")
        return [record]

    # Complete -> run the worker over the whole set, then terminalize.
    # run_input (versions + the input-set fingerprint + per-slot provenance) is
    # passed INTO record_run so it is written atomically with the run line
    # (design §6.1). The fingerprint in that line is what recovery matches on.
    if crash_at == "after_bind_complete":
        raise CrashInjected("after_bind_complete")
    run = fleet.record_run(w, request=item.request,
                           run_input=input_set.run_input(w))
    if crash_at == "after_effect":
        raise CrashInjected("after_effect")
    return _terminalize_set(w, item, run)


def poll(w: fleet.Worker, crash_at: Optional[str] = None) -> list[dict]:
    """One deterministic pass over the inbox. Contains no LLM and no clock logic.

    Every item is claimed before it is run, so an interrupted pass leaves
    evidence rather than a fresh-looking file.

    `crash_at` exists ONLY for the self-test, to stop a pass at one of the two
    windows a real interruption can fall in: `after_claim`, before the effect,
    and `after_effect`, after it but before the terminal ledger line and the
    file move. Nothing in production passes it, and it raises rather than
    returning so the ledger is left exactly as a killed process would leave it.
    """
    ensure(w)
    outcomes = []
    for path in waiting(w):
        # A file may have been drained mid-pass: a prior iteration's set-
        # completion (v0.6 contract path) terminalizes every bound document in
        # the open set, including ones still in this snapshot. Skip anything
        # already gone rather than treating a drain as an unreadable item.
        if not path.is_file():
            continue
        try:
            item = read_item(w, path)
        except (OSError, ValueError, KeyError) as exc:
            record = {"at": _now(), "item_id": f"unreadable:{path.name}",
                      "file": path.name, "state": "exception",
                      "reason": f"unreadable work item: {type(exc).__name__}"}
            _append(w, record)
            shutil.move(str(path), w.directory / "exceptions" / path.name)
            outcomes.append(record)
            continue

        seen = item_state(w, item.item_id)
        if seen == "completed":
            # ALREADY DONE. Not re-run, and above all its effect is not
            # re-applied. The file is filed away so the inbox drains.
            record = {"at": _now(), "item_id": item.item_id,
                      "payload_digest": item.payload_digest, "file": path.name,
                      "state": "skipped_duplicate", "request": item.request,
                      "reason": "an item with identical content already completed"}
            _append(w, record)
            shutil.move(str(path), w.directory / "processed" / path.name)
            outcomes.append(record)
            continue

        # A staged item is already bound into the worker's open input set
        # (v0.6 contract path). It stays in inbox/ waiting for the remaining
        # slots; do not re-claim, do not move it -- the set, not the file
        # location, is the authority for partial state.
        if seen == "staged":
            continue

        _append(w, {"at": _now(), "item_id": item.item_id,
                    "payload_digest": item.payload_digest, "file": path.name,
                    "state": "claimed", "request": item.request,
                    "precondition": _precondition(w, item.request)})
        if crash_at == "after_claim":
            raise CrashInjected("after_claim")

        # --- v0.6 contract path: bind into the open input set ---------------
        # A worker with a version-bound input_contract binds each xlsx arrival
        # to a source ROLE (explicit operator choice for sole slots; all
        # shared roles at once for a shared-slot workbook), validates it
        # against that role's contract, materializes it, and runs ONLY when
        # the set is complete. This recasts the per-file run as "one document
        # completes an N-slot set" and lets a multi-document reconciliation
        # stage until both sides are bound. Non-contract workers fall through
        # to the legacy per-file path unchanged.
        if w.input_contract is not None and path.suffix == ".xlsx":
            outcomes += _poll_contract_item(w, path, item, crash_at)
            continue

        if path.suffix == ".xlsx":
            xlsx = _xlsx()
            conversion = xlsx.convert(path, _sheet_specs(xlsx, w))
            if not conversion.ok:
                record = {"at": _now(), "item_id": item.item_id,
                          "payload_digest": item.payload_digest,
                          "file": path.name, "state": "exception",
                          "request": None, "problems": conversion.problems,
                          "reason": "the workbook could not be converted "
                                    "faithfully"}
                _append(w, record)
                shutil.move(str(path), w.directory / "exceptions" / path.name)
                outcomes.append(record)
                continue
            # Written to exactly the paths the MODEL declares, so the adapter
            # cannot quietly relocate a source the worker depends on.
            for collection, items in conversion.collections.items():
                spec = next((v for v in w.model["sources"].values()
                             if v["collection"] == collection), None)
                if spec is None:
                    continue
                target = w.base / spec["path"]
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_text(
                    json.dumps({"_note": f"converted from {path.name}",
                                collection: items}, indent=2,
                               ensure_ascii=False) + chr(10), encoding="utf-8")

        run = fleet.record_run(w, request=item.request)
        if crash_at == "after_effect":
            raise CrashInjected("after_effect")
        healthy = run["ok"]
        state = "completed" if healthy else "exception"
        destination = "processed" if healthy else "exceptions"
        record = {"at": _now(), "item_id": item.item_id,
                  "payload_digest": item.payload_digest, "file": path.name,
                  "state": state, "request": item.request,
                  "decision": run.get("decision"), "reason": run.get("reason"),
                  "effect_applied": run.get("effect_applied"),
                  "problems": run.get("problems", [])}
        _append(w, record)
        shutil.move(str(path), w.directory / destination / path.name)
        outcomes.append(record)
    return outcomes


# ---------------------------------------------------------------------------
# recovery
# ---------------------------------------------------------------------------

def dangling(w: fleet.Worker) -> list[dict]:
    """Claims with no terminal line -- one per interrupted item."""
    last: dict[str, dict] = {}
    for entry in ledger(w):
        last[entry["item_id"]] = entry
    return [e for e in last.values() if e["state"] == "claimed"]


def reconcile(w: fleet.Worker, claim: dict) -> tuple[str, str]:
    """Decide what an interrupted item's effect actually did.

    already_landed   the effect is present and the claim recorded it ABSENT,
                     so this run applied it. Complete WITHOUT re-executing.
    safe_to_retry    the effect is definitely absent, or none was earnable.
                     Re-execute.
    indeterminate    the question cannot be answered. Exception -- never
                     guessed in either direction.

    The precondition recorded with the claim is what makes this answerable.
    Without it, "the date is in state" is ambiguous: this run may have put it
    there, or it may have been there all along and the decision was a refusal.
    """
    precondition = claim.get("precondition")
    if precondition is None:
        return "indeterminate", ("the claim recorded no precondition, so "
                                 "'present now' cannot be attributed to it")
    if not precondition.get("committing"):
        return "safe_to_retry", ("this worker commits no effect, so "
                                 "re-execution cannot duplicate one")

    before = precondition.get("effect_target_present")
    if before is None:
        return "indeterminate", "worker state was unreadable when it was claimed"
    now = effect_landed(w, claim["request"])
    if now is None:
        return "indeterminate", "worker state is unreadable now"

    if before:
        # Already present before the run, so the decision could not have been
        # an acceptance -- it was a refusal, which earns no effect. Re-running
        # reproduces that refusal deterministically and applies nothing.
        return "safe_to_retry", ("the effect target was already present before "
                                 "the run, so no effect was earnable")
    if now:
        return "already_landed", ("the effect was absent at claim time and is "
                                  "present now, so this run applied it")
    return "safe_to_retry", "the effect is definitely absent"


def _run_with_fingerprint(w: fleet.Worker, fp: str) -> Optional[dict]:
    """The run record in runs.jsonl whose run_input.fingerprint equals `fp`,
    or None. Recovery matches on this to decide run-exactly-once."""
    if not fp:
        return None
    for r in reversed(fleet.load(w.directory).runs):
        if (r.get("run_input") or {}).get("fingerprint") == fp:
            return r
    return None


def _recover_terminalize(w: fleet.Worker, run: dict) -> list[dict]:
    """Terminalize a complete set that recovery resolved: one recovered_
    ledger record per bound document, each raw file drained from inbox/ to
    processed/ or exceptions/, then the set cleared. `run` is either the run
    recovery just recorded (no matching run existed) or the already-recorded
    run recovery matched (so it terminalizes WITHOUT re-running)."""
    healthy = run["ok"]
    state = "recovered_completed" if healthy else "recovered_exception"
    destination = "processed" if healthy else "exceptions"
    doc = input_set.load(w) or {"roles": {}}
    by_doc: dict[str, dict] = {}
    for role, binding in doc.get("roles", {}).items():
        entry = by_doc.setdefault(binding["document"],
                                  {"digest": binding["digest"], "roles": []})
        entry["roles"].append(role)
    outcomes: list[dict] = []
    for document, info in by_doc.items():
        record = {"at": _now(), "item_id": info["digest"],
                  "payload_digest": info["digest"], "file": document,
                  "state": state, "roles": sorted(info["roles"]),
                  "request": None, "decision": run.get("decision"),
                  "reason": run.get("reason"),
                  "effect_applied": run.get("effect_applied"),
                  "problems": run.get("problems", []),
                  "recovered": True}
        _append(w, record)
        src = w.directory / "inbox" / document
        if src.is_file():
            shutil.move(str(src),
                        w.directory / destination / retained_name(document, info["digest"]))
        side = _role_sidecar(src)
        if side.is_file():
            side.unlink()
        outcomes.append(record)
    input_set.clear(w)
    return outcomes


def _recover_contract(w: fleet.Worker) -> list[dict]:
    """The v0.6 set-wise recovery contract (design §5.2). Four cases:

      complete set + no matching run  -> run ONCE, then terminalize  (canary 2)
      complete set + matching run      -> terminalize WITHOUT another run
                                         (canary 3: run exactly once)
      partial set                      -> PRESERVE; wait for remaining slots
                                         (canary 1; no run)
      corrupt/ambiguous                -> exception, never guess

    A complete set's fingerprint is matched against runs.jsonl. This OVERRIDES
    the noncommitting safe_to_retry re-run (reconcile returns safe_to_retry
    for a worker that commits no effect): a completed run is durable history,
    so re-execution would duplicate a runs.jsonl line even though no external
    effect duplicates.
    """
    doc = input_set.load(w)
    if doc is None or not doc.get("complete"):
        # partial or no set: preserve. A partial set waits for the remaining
        # slots; a mid-bind crash's dangling claim is re-bound on the next
        # poll. Recovery runs nothing here.
        return []
    fp = input_set.fingerprint(w)
    matched = _run_with_fingerprint(w, fp)
    if matched is not None:
        # The run is already in runs.jsonl (crash after record_run, before
        # terminalize). Complete WITHOUT another run.
        return _recover_terminalize(w, matched)
    # Complete set, no matching run (crash after materialize, before
    # record_run). Run ONCE -- atomically, with run_input carrying the
    # fingerprint -- then terminalize.
    run = fleet.record_run(w, request=None, run_input=input_set.run_input(w))
    return _recover_terminalize(w, run)


def recover(w: fleet.Worker) -> list[dict]:
    """Resolve every interrupted item. Deterministic; no LLM, no clock logic."""
    ensure(w)
    # v0.6 contract workers recover SET-wise, not item-wise. A complete input
    # set runs exactly once: the fingerprint guard overrides the noncommitting
    # safe_to_retry re-run that would otherwise duplicate a runs.jsonl line
    # (design §5.2). Routed here before the legacy per-item recovery.
    if w.input_contract is not None:
        return _recover_contract(w)
    outcomes = []
    for claim in dangling(w):
        verdict, why = reconcile(w, claim)
        filename = claim.get("file", "")
        path = w.directory / "inbox" / filename
        common = {"at": _now(), "item_id": claim["item_id"],
                  "payload_digest": claim.get("payload_digest"),
                  "file": filename, "request": claim["request"],
                  "verdict": verdict, "reason": why}

        if verdict == "already_landed":
            # Completed WITHOUT re-execution. Re-running would book it twice.
            record = {**common, "state": "recovered_completed",
                      "decision": "accepted", "effect_applied": True}
            _append(w, record)
            if path.is_file():
                shutil.move(str(path), w.directory / "processed" / filename)
            outcomes.append(record)
            continue

        if verdict == "safe_to_retry":
            if not path.is_file():
                record = {**common, "state": "recovered_exception",
                          "verdict": "indeterminate",
                          "reason": "safe to retry, but the work item is no "
                                    "longer in the inbox"}
                _append(w, record)
                outcomes.append(record)
                continue
            run = fleet.record_run(w, request=claim["request"])
            healthy = run["ok"]
            record = {**common,
                      "state": "recovered_completed" if healthy
                               else "recovered_exception",
                      "decision": run.get("decision"),
                      "reason": run.get("reason") or why,
                      "effect_applied": run.get("effect_applied")}
            _append(w, record)
            shutil.move(str(path),
                        w.directory / ("processed" if healthy else "exceptions")
                        / filename)
            outcomes.append(record)
            continue

        record = {**common, "state": "recovered_exception"}
        _append(w, record)
        if path.is_file():
            shutil.move(str(path), w.directory / "exceptions" / filename)
        outcomes.append(record)
    return outcomes


def retry(w: fleet.Worker, filename: str) -> None:
    """Move one exception back to the inbox, deliberately.

    Retrying is a decision a person makes, so it is an explicit act rather than
    something the poller does on a timer. An item whose effect never landed will
    apply it on the retry; one that completed is caught by the ledger.
    """
    source = w.directory / "exceptions" / filename
    shutil.move(str(source), w.directory / "inbox" / filename)


def summary(w: fleet.Worker) -> dict:
    ensure(w)
    entries = ledger(w)
    final: dict[str, str] = {}
    # An item that completed and was later RESENT ends on skipped_duplicate, so
    # "did it ever complete" is a different question from "what happened last".
    # Counting only the last state reported 0 completed for a worker that had
    # done its job -- found by resending a workbook.
    ever_completed: set = set()
    for entry in entries:
        final[entry["item_id"]] = entry["state"]
        if entry["state"] in ("completed", "recovered_completed"):
            ever_completed.add(entry["item_id"])

    def _files(folder: str) -> int:
        directory = w.directory / folder
        return len([p for p in directory.iterdir() if p.is_file()])             if directory.is_dir() else 0

    return {
        "waiting": len(waiting(w)),
        "processed": _files("processed"),
        "exceptions": _files("exceptions"),
        "items_seen": len(final),
        "completed": len(ever_completed),
        "in_flight": sum(1 for s in final.values() if s == "claimed"),
        "staged": sum(1 for s in final.values() if s == "staged"),
        "recovered": sum(1 for e in entries
                         if str(e["state"]).startswith("recovered_")),
        "duplicates_skipped": sum(1 for e in entries
                                  if e["state"] == "skipped_duplicate"),
        "ledger_lines": len(entries),
    }


def _self_test() -> int:
    """Runs inside the lab, not the system temp dir.

    A worker's base is stored relative to the lab root, so a scratch worker has
    to live where that path can be expressed. Cleaned up either way.
    """
    failures: list[str] = []

    def check(cond: bool, msg: str) -> None:
        if not cond:
            failures.append(msg)

    model = json.loads((fleet.LAB / "reservation" / "models" /
                        "reservation_v1.json").read_text(encoding="utf-8"))
    rc = _run_in_lab(check, failures, model)
    if rc:
        return rc
    # The v0.6 hard gate: the inbox spec-source moved from
    # identity["adapter_sheets"] to the version-bound input_contract. A scratch
    # xlsx worker whose contract is the ONLY place its sheet shape lives must
    # still convert, run, and file to processed/ via the ordinary poll. This is
    # the Fazerish-migration canary: if the contract path is not wired, the
    # live xlsx inbox breaks.
    return _run_xlsx_in_lab(check, failures)


def _run_xlsx_in_lab(check, failures) -> int:
    scratch = fleet.LAB / "fleet" / ".selftest-xlsx"
    if scratch.exists():
        shutil.rmtree(scratch)
    try:
        root = scratch / "workers"
        root.mkdir(parents=True)
        fazerish_model = json.loads(
            (fleet.LAB / "fleet" / "workers" / "fazerish-invoicing" /
             "versions" / "v1.json").read_text(encoding="utf-8"))
        base_rel = "fleet/.selftest-xlsx/workers/xlsx-worker/state"
        w = fleet.establish(root, "xlsx-worker", "Xlsx inbox gate.", "enrichment",
                            base_rel, fazerish_model)
        # Post-write the stable source roles + operational policy (establish
        # writes only the identity core; roles/policy are added the same way
        # `work_item_identity` is in the reservation self-test).
        ident = json.loads((w.directory / "worker.json").read_text(encoding="utf-8"))
        ident["input_adapter"] = "xlsx"
        ident["work_item_identity"] = "content_digest"
        ident["source_roles"] = {
            "order_lines": {"label": "order lines", "slot": "shared",
                            "required": True},
            "price_list": {"label": "price list", "slot": "shared",
                           "required": True},
        }
        (w.directory / "worker.json").write_text(
            json.dumps(ident, indent=2) + "\n", encoding="utf-8")
        # The contract is the ONLY place the sheet shape lives now -- no
        # adapter_sheets on identity. Same version as the model.
        contracts_dir = w.directory / "input_contracts"
        contracts_dir.mkdir()
        (contracts_dir / "v1.json").write_text(json.dumps({
            "roles": {
                "order_lines": {"sheet": "Order lines", "collection": "order_lines",
                                "header_row": 1},
                "price_list": {"sheet": "Price list", "collection": "price_list",
                               "header_row": 1},
            }}, indent=2) + "\n", encoding="utf-8")
        w = fleet.load(w.directory)
        ensure(w)
        check(w.input_contract is not None
              and "adapter_sheets" not in w.identity,
              "the xlsx gate worker sources shape from input_contract, not identity")

        shutil.copy(
            fleet.LAB / "data" / "xlsx-fazerish" / "may-order-lines.xlsx",
            w.directory / "inbox" / "may.xlsx")

        out = poll(w)
        check(len(out) == 1 and out[0]["state"] == "completed",
              f"the contract-sourced xlsx arrival completes via the ordinary "
              f"poll: {out}")
        _may_src = fleet.LAB / "data" / "xlsx-fazerish" / "may-order-lines.xlsx"
        check((w.directory / "processed" / retained_name(
                  "may.xlsx", hashlib.sha256(_may_src.read_bytes()).hexdigest())).is_file(),
              "…and is filed to processed/ by the ordinary path (digest-namespaced)")
        check(any(e["state"] == "completed" for e in ledger(w)),
              "…and wrote a ledger line")
        runs = fleet.load(w.directory).runs
        check(runs and runs[-1]["ok"] and runs[-1]["rows"] == 3,
              f"…and the enrichment ran clean on the materialized sources: "
              f"{runs[-1] if runs else None}")
        # the materialized source JSONs landed at the model-declared paths
        check((w.base / "sources" / "order_lines.json").is_file()
              and (w.base / "sources" / "price_list.json").is_file(),
              "the workbook's sheets were written to the model's source paths")

        # --- PHASE 6: collision-safe raw retention (design §6.2) ------------
        # The run above landed `may.xlsx` (digest_a) in processed/. Now arrive a
        # SAME-named, DIFFERENT-bytes workbook: it must NOT overwrite the first
        # -- both survive, distinguishable by digest, and the ledger carries
        # both. The bare filename would lose the exact raw evidence the earlier
        # run recorded; the digest is the truth (materialized JSON is still
        # overwritten per run -- this is raw retention only).
        from openpyxl import load_workbook
        src_a = fleet.LAB / "data" / "xlsx-fazerish" / "may-order-lines.xlsx"
        digest_a = hashlib.sha256(src_a.read_bytes()).hexdigest()
        twin = scratch / "twin.xlsx"
        wb = load_workbook(src_a)
        # bump one Qty cell: same shape, different bytes, still a healthy run
        cell = wb["Order lines"].cell(row=3, column=3)
        cell.value = int(cell.value or 0) + 1
        wb.save(twin)
        digest_b = hashlib.sha256(twin.read_bytes()).hexdigest()
        check(digest_a != digest_b, "the twin has different bytes (same shape)")
        shutil.copy(twin, w.directory / "inbox" / "may.xlsx")
        poll(w)
        check((w.directory / "processed" / retained_name("may.xlsx", digest_a)).is_file()
              and (w.directory / "processed" / retained_name("may.xlsx", digest_b)).is_file(),
              "BOTH same-named different-bytes arrivals survive in processed/ "
              "(distinguishable by digest, not overwrite)")
        completed = [e for e in ledger(w) if e["state"] == "completed"]
        by_file: dict[str, set] = {}
        for e in completed:
            by_file.setdefault(e["file"], set()).add(e["payload_digest"])
        check("may.xlsx" in by_file and digest_a in by_file["may.xlsx"]
              and digest_b in by_file["may.xlsx"],
              f"the ledger carries both arrivals under file may.xlsx with "
              f"distinct digests: { {k: sorted(v) for k, v in by_file.items()} }")
    finally:
        shutil.rmtree(scratch, ignore_errors=True)

    if failures:
        sys.stderr.write("XLSX GATE FAILED:\n  " + "\n  ".join(failures) + "\n")
        return 1
    print("XLSX GATE PASSED (a scratch worker whose sheet shape lives ONLY in "
          "input_contracts/v1.json still converts, runs the enrichment, files "
          "to processed/ and writes a ledger line via the ordinary poll -- the "
          "v0.6 spec-source move is atomic and non-breaking / "
          "Phase 6: two same-named different-bytes arrivals both survive in "
          "processed/ distinguishable by digest, and the ledger carries both)")
    return 0


def _run_in_lab(check, failures, model) -> int:
    import os
    import stat

    scratch = fleet.LAB / "fleet" / ".selftest"
    if scratch.exists():
        shutil.rmtree(scratch)
    try:
        root = scratch / "workers"
        root.mkdir(parents=True)
        base_rel = "fleet/.selftest/workers/inbox-worker/state"
        w = fleet.establish(root, "inbox-worker", "Inbox test.", "reservation",
                            base_rel, model)
        ident = json.loads((w.directory / "worker.json").read_text(encoding="utf-8"))
        ident["work_item_identity"] = "content_digest"
        (w.directory / "worker.json").write_text(
            json.dumps(ident, indent=2) + chr(10), encoding="utf-8")

        shutil.copytree(fleet.LAB / "reservation" / "fixtures",
                        w.directory / "state" / "fixtures")
        w = fleet.load(w.directory)
        ensure(w)

        def drop(name: str, request: str) -> Path:
            path = w.directory / "inbox" / name
            path.write_text(json.dumps({"request": request}) + "\n",
                            encoding="utf-8")
            return path

        def reservations() -> list:
            return json.loads((w.directory / "state" / "fixtures" /
                               "reservations.json").read_text(encoding="utf-8"))["reservations"]

        start = len(reservations())

        # --- a policy refusal COMPLETES normally --------------------------
        drop("a.json", "2026-12-25")
        out = poll(w)
        check(out[0]["state"] == "completed" and out[0]["reason"] == "HOLIDAY",
              f"a policy refusal must complete, not except: {out}")
        check((w.directory / "processed" / "a.json").is_file(),
              "…and be filed as processed")
        check(len(reservations()) == start, "…and change no state")

        # --- an acceptance applies its effect ------------------------------
        drop("b.json", "2026-04-02")
        out = poll(w)
        check(out[0]["state"] == "completed" and out[0]["effect_applied"] is True,
              f"an acceptance must complete with its effect applied: {out}")
        check(len(reservations()) == start + 1,
              f"…and worker state must grow: {reservations()}")

        # --- THE RETRY: identical content must not double-apply ------------
        after_first = list(reservations())
        drop("b-again.json", "2026-04-02")          # same bytes, new filename
        out = poll(w)
        check(out[0]["state"] == "skipped_duplicate",
              f"CANARY: identical content is the same work item: {out}")
        check(reservations() == after_first,
              f"CANARY: the effect must NOT be applied twice: {reservations()}")
        check(len(fleet.load(w.directory).runs) == 2,
              "CANARY: a duplicate must not even produce a run")

        # --- an effect that cannot land becomes an EXCEPTION ---------------
        target = w.directory / "state" / "fixtures" / "reservations.json"
        drop("c.json", "2026-05-05")
        os.chmod(target, stat.S_IREAD)
        try:
            out = poll(w)
        finally:
            os.chmod(target, stat.S_IWRITE | stat.S_IREAD)
        check(out[0]["state"] == "exception"
              and out[0]["effect_applied"] is False,
              f"an accepted decision whose effect failed must except: {out}")
        check((w.directory / "exceptions" / "c.json").is_file(),
              "…and land in the exception queue")
        check(len(reservations()) == start + 1,
              f"…and have changed nothing: {reservations()}")

        # --- retrying an exception DOES apply it ---------------------------
        retry(w, "c.json")
        out = poll(w)
        check(out[0]["state"] == "completed"
              and out[0]["effect_applied"] is True,
              f"a retried exception must apply once the cause is gone: {out}")
        check(len(reservations()) == start + 2,
              f"…and state must grow exactly once: {reservations()}")

        # --- and now IT is a duplicate too ---------------------------------
        after_retry = list(reservations())
        drop("c-again.json", "2026-05-05")
        poll(w)
        check(reservations() == after_retry,
              f"CANARY: a completed retry cannot be re-applied: {reservations()}")

        # --- an unreadable item excepts rather than stopping the pass ------
        (w.directory / "inbox" / "bad.json").write_text("{not json",
                                                        encoding="utf-8")
        drop("d.json", "2026-06-01")
        out = poll(w)
        states = {o["file"]: o["state"] for o in out}
        check(states.get("bad.json") == "exception",
              f"an unreadable item must except: {states}")
        check(states.get("d.json") == "completed",
              f"CANARY: and must not stop the rest of the pass: {states}")

        # --- CRASH IMMEDIATELY BEFORE THE EFFECT ---------------------------
        # The claim is written; the process dies before the run. Nothing was
        # applied, so recovery must RETRY -- losing the work would be the bug.
        before_crash = list(reservations())
        drop("e.json", "2026-07-14")
        try:
            poll(w, crash_at="after_claim")
            failures.append("the injected crash must actually interrupt")
        except CrashInjected:
            pass
        check(reservations() == before_crash,
              f"nothing may have been applied yet: {reservations()}")
        check((w.directory / "inbox" / "e.json").is_file(),
              "the work item is still in the inbox, exactly as a kill leaves it")
        stuck = dangling(w)
        check(len(stuck) == 1 and stuck[0]["request"] == "2026-07-14",
              f"the interrupted item is visible as claimed: {stuck}")

        verdict, why = reconcile(w, stuck[0])
        check(verdict == "safe_to_retry",
              f"a crash BEFORE the effect must reconcile as retryable: "
              f"{verdict} -- {why}")
        out = recover(w)
        check(out[0]["state"] == "recovered_completed"
              and out[0]["effect_applied"] is True,
              f"recovery must apply the work that was lost: {out}")
        check(reservations() == before_crash + ["2026-07-14"],
              f"…exactly once: {reservations()}")
        check(not dangling(w), "…and nothing is left claimed")

        # --- CRASH IMMEDIATELY AFTER THE EFFECT ----------------------------
        # The effect landed but the terminal line was never written. Recovery
        # must complete WITHOUT re-executing; re-running would book it twice.
        before_crash = list(reservations())
        runs_before = len(fleet.load(w.directory).runs)
        drop("f.json", "2026-08-20")
        try:
            poll(w, crash_at="after_effect")
            failures.append("the injected crash must actually interrupt")
        except CrashInjected:
            pass
        check(reservations() == before_crash + ["2026-08-20"],
              f"the effect DID land before the interruption: {reservations()}")
        check((w.directory / "inbox" / "f.json").is_file(),
              "and the file was never moved, so a naive poller would rerun it")
        stuck = dangling(w)
        verdict, why = reconcile(w, stuck[0])
        check(verdict == "already_landed",
              f"a crash AFTER the effect must reconcile as already landed: "
              f"{verdict} -- {why}")

        landed_state = list(reservations())
        out = recover(w)
        check(out[0]["state"] == "recovered_completed",
              f"recovery must complete it: {out}")
        check(reservations() == landed_state,
              f"CANARY: and must NOT apply the effect a second time: "
              f"{reservations()}")
        check(len(fleet.load(w.directory).runs) == runs_before + 1,
              "CANARY: recovery must not re-execute -- no extra run")
        check((w.directory / "processed" / "f.json").is_file(),
              "…and the item is filed as processed")

        # --- a naive poll would have duplicated it -------------------------
        # The date is now in state, so even a rerun would be refused. The
        # ledger is what stops it BEFORE that, and both matter.
        check(item_state(w, read_item(w, w.directory / "processed" /
                                      "f.json").item_id) == "recovered_completed",
              "the item is terminal in the ledger, so it is not work any more")

        # --- INDETERMINATE: unreadable state -> exception, never a guess ----
        drop("g.json", "2026-09-01")
        try:
            poll(w, crash_at="after_claim")
        except CrashInjected:
            pass
        stuck = dangling(w)[0]
        state_file = w.directory / "state" / "fixtures" / "reservations.json"
        good = state_file.read_text(encoding="utf-8")
        state_file.write_text("{ this is not json", encoding="utf-8")
        try:
            verdict, why = reconcile(w, stuck)
            check(verdict == "indeterminate",
                  f"CANARY: unreadable state must be indeterminate, not a "
                  f"guess: {verdict} -- {why}")
            out = recover(w)
            check(out[0]["state"] == "recovered_exception",
                  f"…and must go to the exception queue: {out}")
        finally:
            state_file.write_text(good, encoding="utf-8")
        check((w.directory / "exceptions" / "g.json").is_file(),
              "the indeterminate item is queued for a person")

        # --- a claim with no precondition is indeterminate too -------------
        check(reconcile(w, {"item_id": "x", "request": "2026-01-01",
                            "state": "claimed"})[0] == "indeterminate",
              "CANARY: a claim carrying no precondition cannot be attributed")

        # --- payload digest and work-item identity stay distinct -----------
        sample = read_item(w, w.directory / "processed" / "b.json")
        check(sample.payload_digest == sample.item_id,
              "for THIS worker the two coincide, by its declared policy")
        entry = [e for e in ledger(w) if e.get("file") == "b.json"][0]
        check("payload_digest" in entry and "item_id" in entry,
              f"…but both are recorded separately: {sorted(entry)}")
        w.identity["work_item_identity"] = "invoice_number"
        try:
            read_item(w, w.directory / "processed" / "b.json")
            failures.append("CANARY: an unimplemented identity policy must be "
                            "refused, not defaulted to content_digest")
        except UnknownIdentityPolicy:
            pass
        w.identity["work_item_identity"] = "content_digest"

        # --- the ledger is append-only and the summary agrees --------------
        s = summary(w)
        check(s["waiting"] == 0 and s["in_flight"] == 0,
              f"the inbox must drain: {s}")
        check(s["duplicates_skipped"] == 2, f"two duplicates skipped: {s}")
        check(s["recovered"] == 3, f"three items resolved by recovery: {s}")
        entries = ledger(w)
        # The invariant, restated once recovery exists: every run is preceded
        # by a claim, and the only claims WITHOUT a run are the ones recovery
        # could not attribute. A duplicate produces neither; an unreadable item
        # produces neither.
        claims = [e["state"] for e in entries].count("claimed")
        runs = len(fleet.load(w.directory).runs)
        unattributed = sum(1 for e in entries
                           if e["state"] == "recovered_exception"
                           and e.get("verdict") == "indeterminate")
        check(claims == runs + unattributed,
              f"{claims} claim(s) = {runs} run(s) + {unattributed} "
              f"unattributed; every run was claimed first")
        check(unattributed == 1,
              f"exactly one claim was never attributed -- the unreadable-state "
              f"item: {unattributed}")
        check(s["ledger_lines"] == len(entries), "the summary reads the ledger")

    finally:
        if scratch.exists():
            shutil.rmtree(scratch, ignore_errors=True)

    if failures:
        sys.stderr.write("SELF-TEST FAILED:\n  " + "\n  ".join(failures) + "\n")
        return 1
    print("SELF-TEST PASSED (a policy refusal completes and changes no state / "
          "an acceptance applies its effect and grows state / identical content "
          "is the same work item, produces no run and does NOT apply twice / a "
          "failed effect excepts, queues and changes nothing / retrying applies "
          "it exactly once / an unreadable item excepts without stopping the "
          "pass / A CRASH BEFORE THE EFFECT reconciles as retryable and "
          "recovery applies the lost work exactly once / A CRASH AFTER THE "
          "EFFECT reconciles as already landed and recovery completes it "
          "WITHOUT re-executing and WITHOUT a second effect / unreadable state "
          "and a precondition-less claim are indeterminate and queue for a "
          "person rather than being guessed / payload digest and work-item "
          "identity are recorded separately and an unimplemented identity "
          "policy is refused / one claim per run and the inbox drains)")
    return 0


if __name__ == "__main__":
    raise SystemExit(_self_test() if sys.argv[1:2] == ["--self-test"] else 2)
